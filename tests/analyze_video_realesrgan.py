#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RealESRGAN 超分处理结果检测分析脚本
=================================
依据 RealESRGAN 流水线输出日志及所有分段/超分文件，生成以下检测分析文件：
  1. ffmpeg_outputX.txt      — ffmpeg 逐段 pipe 解码日志（含丢帧/错误信息）
  2. ffprob_outputX.txt      — ffprobe 逐段帧数统计（segments + processed + 输出视频）
  3. 帧数差值统计汇总_outputX.xlsx — 帧数差值汇总 Excel 表格
     （含：日志提示输出帧数/差值1/ffprob count_packets/差值2/ffmpeg拆帧/差值3/
           异常检测/性能统计(FPS/平均ms/总时间)/GFPGAN模式）

与 IFRNet 版本的区别：
  - 超分不改变帧数，期望输出帧数 = 原始帧数
  - 处理后文件为 upscaled_segment_XXX.mp4（非 interpolated_）
  - 解析 RealESRGAN 特有日志字段：[性能统计]、GFPGAN 模式、已用时/预计剩余、模型加载耗时等

用法:
    python analyze_video_realesrgan.py <output_log> [--suffix X] ...

示例:
    python analyze_video_realesrgan.py temp/esrgan_video/output.txt --suffix esrgan \
        --segments-dir ./segments --processed-dir ./processed \
        --output-video ./output.mp4
"""

import argparse
import re
import os
import subprocess
import sys
import shlex
from pathlib import Path
from collections import OrderedDict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl 未安装，将不会生成 xlsx 文件。安装: pip install openpyxl")


# ═══════════════════════════════════════════════
# 1. 日志解析
# ═══════════════════════════════════════════════

def parse_output_log(log_path: str) -> dict:
    """解析 RealESRGAN 超分流水线日志，提取每段帧数及性能信息"""
    with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    info = {
        'num_segments': 0,
        'segments': [],
        'segment_duration': 300,
        'source_video': '',
        'output_video': '',
        'total_input_duration': '',
        'total_output_duration': '',
        'total_output_size': '',
        'total_elapsed': '',
        'total_input_frames': None,
        'total_output_frames': None,
        'gfpgan_mode': 'unknown',
        'total_perf_time': None,         # 总时间(秒)，来自[性能统计]
        'total_perf_avg_ms': None,       # 平均 ms
        'total_perf_fps': None,          # FPS
        'total_model_load_time': '',     # 含模型加载总耗时
        'total_elapsed_estimate': '',    # 已用时（最后一行）
        'total_remaining_estimate': '',  # 预计剩余
    }

    # 分段时长
    m = re.search(r'分段时长\s*[:：]\s*(\d+)', content)
    if m:
        info['segment_duration'] = int(m.group(1))

    # 源/输出视频
    m = re.search(r'输入\s*[:：]\s*(.+\.mp4)', content)
    if m:
        info['source_video'] = m.group(1).strip()
    # RealESRGAN 输出文件路径形如：✅ 输出: /path/upscaled_segment_000.mp4 (964.7 MB)
    # 全局输出视频另行匹配
    # 匹配多种输出视频格式：
    #   📤 输出: /path/to/video.mp4
    #   📤 输出文件: /path/to/video.mp4
    #   最终输出: /path/to/video.mp4
    m = re.search(r'(?:📤\s*输出(?:文件)?|最终输出)\s*[:：]\s*(.+\.mp4)', content)
    if m:
        info['output_video'] = m.group(1).strip()

    # 分段总数
    m = re.search(r'✅\s*共\s*(\d+)\s*个片段', content)
    if m:
        info['num_segments'] = int(m.group(1))
    else:
        segs = re.findall(r'片段\s*(\d+)/(\d+):', content)
        if segs:
            info['num_segments'] = max(int(s[1]) for s in segs)

    # GFPGAN 模式（全局，可能在每段重复打印）
    m = re.search(r'\[性能统计\]\s*GFPGAN\s*模式\s*[:：]\s*(\S+)', content)
    if m:
        info['gfpgan_mode'] = m.group(1).strip()

    # ── 每段解析 ──
    # 片段头：🎨 片段 2/6: interpolated_segment_001.mp4 或 segment_001.mp4
    segment_pattern = re.compile(
        r'🎨\s*片段\s*(\d+)/(\d+):\s*\S*segment_(\d+)\.mp4'
    )
    # 原始帧: 7223 → 输出帧: 7223
    frame_pattern = re.compile(
        r'原始帧:\s*(\d+)\s*→\s*输出帧:\s*(\d+)'
    )
    # 总耗时（含模型加载）: 3分50秒
    model_load_time_pattern = re.compile(
        r'总耗时（含模型加载）\s*[:：]\s*([\d分秒]+)'
    )
    # ✅ 输出: /path/upscaled_interpolated_segment_000.mp4 (964.7 MB)
    output_file_pattern = re.compile(
        r'✅\s*输出:\s*(.+?segment_\d+\.mp4)\s*\(([\d.]+\s*[KMGT]B)\)'
    )
    # [性能统计] 总时间: 230.8秒 | 平均: 705.7ms | FPS: 31.29
    perf_pattern = re.compile(
        r'\[性能统计\]\s*总时间\s*[:：]\s*([\d.]+)秒\s*\|\s*平均\s*[:：]\s*([\d.]+)ms\s*\|\s*FPS\s*[:：]\s*([\d.]+)'
    )
    # [性能统计] GFPGAN 模式: disabled
    gfpgan_pattern = re.compile(
        r'\[性能统计\]\s*GFPGAN\s*模式\s*[:：]\s*(\S+)'
    )
    # 输出时长 5:01.259 | 耗时 3:51.947
    time_pattern = re.compile(
        r'输出时长\s*([\d:.]+)\s*\|\s*耗时\s*([\d:.]+)'
    )
    # 已用时: 3:56.667, 预计剩余: 19:43.336
    estimate_pattern = re.compile(
        r'已用时\s*[:：]\s*([\d:.]+)\s*,\s*预计剩余\s*[:：]\s*([\d:.]+)'
    )

    seg_starts = [(m.start(), int(m.group(1)), int(m.group(3)))
                  for m in segment_pattern.finditer(content)]

    for idx, (start_pos, seg_num, seg_id) in enumerate(seg_starts):
        end_pos = seg_starts[idx + 1][0] if idx + 1 < len(seg_starts) else len(content)
        seg_text = content[start_pos:end_pos]

        seg_info = {
            'index': seg_id,
            'number': seg_num,
            'raw_frames': None,
            'output_frames': None,        # 日志提示输出帧数
            'expected_frames': None,      # 超分: 期望 = 原始帧
            'frame_diff': None,           # 差值1 = 日志输出 - 期望
            'output_duration': '',
            'elapsed': '',
            'output_file': '',
            'output_size': '',
            'model_load_time': '',
            'perf_total_sec': None,
            'perf_avg_ms': None,
            'perf_fps': None,
            'gfpgan_mode': 'unknown',
            'elapsed_estimate': '',
            'remaining_estimate': '',
        }

        fm = frame_pattern.search(seg_text)
        if fm:
            seg_info['raw_frames'] = int(fm.group(1))
            seg_info['output_frames'] = int(fm.group(2))
            # 超分：期望输出帧数 = 原始帧数
            seg_info['expected_frames'] = seg_info['raw_frames']
            seg_info['frame_diff'] = seg_info['output_frames'] - seg_info['expected_frames']

        mlm = model_load_time_pattern.search(seg_text)
        if mlm:
            seg_info['model_load_time'] = mlm.group(1)

        ofm = output_file_pattern.search(seg_text)
        if ofm:
            seg_info['output_file'] = ofm.group(1).strip()
            seg_info['output_size'] = ofm.group(2).strip()

        pm = perf_pattern.search(seg_text)
        if pm:
            seg_info['perf_total_sec'] = float(pm.group(1))
            seg_info['perf_avg_ms'] = float(pm.group(2))
            seg_info['perf_fps'] = float(pm.group(3))

        gm = gfpgan_pattern.search(seg_text)
        if gm:
            seg_info['gfpgan_mode'] = gm.group(1).strip()

        tm = time_pattern.search(seg_text)
        if tm:
            seg_info['output_duration'] = tm.group(1)
            seg_info['elapsed'] = tm.group(2)

        em = estimate_pattern.search(seg_text)
        if em:
            seg_info['elapsed_estimate'] = em.group(1)
            seg_info['remaining_estimate'] = em.group(2)

        info['segments'].append(seg_info)

    # ── 全局汇总信息 ──
    # 最后一行预计剩余可反映整体进度
    all_estimates = estimate_pattern.findall(content)
    if all_estimates:
        info['total_elapsed_estimate'] = all_estimates[-1][0]
        info['total_remaining_estimate'] = all_estimates[-1][1]

    # 全局输出大小（如有合并后大小）
    m = re.search(r'📦\s*文件大小\s*[:：]\s*([\d.]+\s*[KMGT]B)', content)
    if m:
        info['total_output_size'] = m.group(1)
    m = re.search(r'总用时\s*[:：]\s*([\d:.]+)', content)
    if m:
        info['total_elapsed'] = m.group(1)

    # 全局性能统计汇总：把所有段的 perf 累加
    perf_matches = perf_pattern.findall(content)
    if perf_matches:
        info['total_perf_time'] = sum(float(x[0]) for x in perf_matches)
        # 平均 ms 与 FPS 取均值
        info['total_perf_avg_ms'] = sum(float(x[1]) for x in perf_matches) / len(perf_matches)
        info['total_perf_fps'] = sum(float(x[2]) for x in perf_matches) / len(perf_matches)

    # 含模型加载总耗时（取最后一段或汇总）
    mlm_all = model_load_time_pattern.findall(content)
    if mlm_all:
        info['total_model_load_time'] = mlm_all[-1]

    return info


# ═══════════════════════════════════════════════
# 2. ffmpeg image2pipe / null 解码 + 错误检测
# ═══════════════════════════════════════════════

def _ffmpeg_image2pipe_decode(video_path: str, timeout: int = 300) -> dict:
    """
    使用 ffmpeg -f image2pipe -vcodec png - 触发完整解码+编码管线，
    stdout 丢弃避免 OOM，流式捕获 stderr。
    返回: {'frame_count': int|None, 'errors': [str], 'dup_count': int|None, 'drop_count': int|None}
    """
    result = {
        'frame_count': None,
        'errors': [],
        'dup_count': None,
        'drop_count': None,
    }
    cmd = [
        "ffmpeg", "-hide_banner",
        "-i", str(video_path),
        "-f", "image2pipe",
        "-vcodec", "png",
        "-"
    ]
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
        try:
            for line in proc.stderr:
                stripped = line.strip()

                m = re.search(r'frame=\s*(\d+)', stripped)
                if m:
                    result['frame_count'] = int(m.group(1))

                m_dup = re.search(r'dup=(\d+)', stripped)
                m_drop = re.search(r'drop=(\d+)', stripped)
                if m_dup and m_drop and 'Lsize=' in stripped:
                    result['dup_count'] = int(m_dup.group(1))
                    result['drop_count'] = int(m_drop.group(1))

                lower = stripped.lower()
                if any(kw in lower for kw in [
                    'illegal', 'error', 'missing', 'corrupt',
                    'duplicated', 'duplicate', 'decoder',
                    'non-existing pps', 'no frame',
                    'more than 1000 frames duplicated',
                ]):
                    result['errors'].append(stripped)

            proc.wait(timeout=timeout)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg image2pipe 解码超时")

    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")

    return result


def _ffmpeg_null_decode(video_path: str, timeout: int = 180) -> dict:
    """
    使用 ffmpeg -f null - 轻量解码（不编码，速度快），
    流式捕获 stderr 中的帧数和错误信息。
    """
    result = {
        'frame_count': None,
        'errors': [],
        'dup_count': None,
        'drop_count': None,
    }
    cmd = [
        "ffmpeg", "-hide_banner",
        "-i", str(video_path),
        "-f", "null", "-"
    ]
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
        try:
            for line in proc.stderr:
                stripped = line.strip()

                m = re.search(r'frame=\s*(\d+)', stripped)
                if m:
                    result['frame_count'] = int(m.group(1))

                m_dup = re.search(r'dup=(\d+)', stripped)
                m_drop = re.search(r'drop=(\d+)', stripped)
                if m_dup and m_drop and 'Lsize=' in stripped:
                    result['dup_count'] = int(m_dup.group(1))
                    result['drop_count'] = int(m_drop.group(1))

                lower = stripped.lower()
                if any(kw in lower for kw in [
                    'illegal', 'error', 'missing', 'corrupt',
                    'duplicated', 'duplicate', 'decoder',
                    'non-existing pps', 'no frame',
                    'more than 1000 frames duplicated',
                ]):
                    result['errors'].append(stripped)

            proc.wait(timeout=timeout)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg null 解码超时")

    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")

    return result


def generate_ffmpeg_output(processed_dir: str, output_path: str, num_segments: int,
                           use_image2pipe: bool = False, processed_prefix: str = 'upscaled_segment'):
    """
    对每个 upscaled_segment_XXX.mp4 执行 ffmpeg 解码。
    use_image2pipe=True: -f image2pipe（完整解码+PNG编码管线，慢但能捕获 dup/drop）
    use_image2pipe=False: -f null（轻量解码，快）
    返回: (ffmpeg_frames: list, ffmpeg_errors: list, ffmpeg_dup_drop: list)
    """
    processed_path = Path(processed_dir)
    ffmpeg_frames = []
    ffmpeg_errors = []
    ffmpeg_dup_drop = []

    def _find_segment_file(proc_dir: Path, seg_idx: int) -> Path:
        """通过 *segment_XXX.mp4 模式匹配查找文件，适配多种命名前缀"""
        pattern = f"*segment_{seg_idx:03d}.mp4"
        matches = list(proc_dir.glob(pattern))
        if matches:
            return matches[0]
        # 回退到旧命名方式
        fallback = proc_dir / f"{processed_prefix}_{seg_idx:03d}.mp4"
        return fallback

    with open(output_path, 'w', encoding='utf-8') as out:
        for i in range(num_segments):
            seg_file = _find_segment_file(processed_path, i)
            seg_filename = seg_file.name

            if use_image2pipe:
                cmd_str = (f"PS> ffmpeg -hide_banner -i "
                           f".\\{seg_filename} "
                           f"-f image2pipe -vcodec png - | find /v \"\"\n")
            else:
                cmd_str = (f"PS> ffmpeg -hide_banner -i "
                           f".\\{seg_filename} "
                           f"-f null - | find /v \"\"\n")
            out.write(cmd_str)

            if not seg_file.exists():
                print(f"  [WARN] 文件不存在: {seg_file}")
                ffmpeg_frames.append(None)
                ffmpeg_errors.append([f"[FILE NOT FOUND] {seg_file}"])
                ffmpeg_dup_drop.append(None)
                out.write(f"[FILE NOT FOUND]\n")
                out.write(f"PS>\n")
                continue

            if use_image2pipe:
                result = _ffmpeg_image2pipe_decode(str(seg_file))
            else:
                result = _ffmpeg_null_decode(str(seg_file))
            ffmpeg_frames.append(result['frame_count'])
            ffmpeg_errors.append(result['errors'])
            ffmpeg_dup_drop.append({
                'dup': result['dup_count'],
                'drop': result['drop_count'],
            })

            print(f"  [ffmpeg] {seg_filename}: frame={result['frame_count']}, "
                  f"dup={result['dup_count']}, drop={result['drop_count']}, "
                  f"errors={len(result['errors'])}")

            if result['frame_count'] is not None:
                out.write(f"frame={result['frame_count']}\n")
            for err in result['errors']:
                out.write(err + '\n')
            out.write(f"PS>\n")

    print(f"✅ ffmpeg_output 已生成: {output_path}")
    return ffmpeg_frames, ffmpeg_errors, ffmpeg_dup_drop


# ═══════════════════════════════════════════════
# 3. ffprobe 帧数统计
# ═══════════════════════════════════════════════

def _ffprobe_count(video_path: str, count_mode: str = 'packets', timeout: int = 60) -> str:
    """ffprobe 获取帧/包计数。count_mode: 'packets' | 'frames'"""
    if count_mode == 'packets':
        cmd = [
            "ffprobe", "-v", "error",
            "-select_streams", "v:0",
            "-count_packets",
            "-show_entries", "stream=nb_read_packets",
            "-of", "csv=p=0",
            video_path
        ]
    else:
        cmd = [
            "ffprobe", "-v", "error",
            "-select_streams", "v:0",
            "-count_frames",
            "-show_entries", "stream=nb_read_frames",
            "-of", "csv=p=0",
            video_path
        ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        stdout = result.stdout.strip()
        stderr = result.stderr.strip()
        m = re.search(r'(\d+)', stdout)
        if m:
            return m.group(1)
        m = re.search(r'(\d+)', stderr)
        if m:
            return m.group(1)
        return stdout if stdout else "[ERROR]"
    except subprocess.TimeoutExpired:
        return "[TIMEOUT]"
    except FileNotFoundError:
        return "[SKIP: ffprobe not found]"


def _ffprobe_count_int(video_path: str, count_mode: str = 'packets', timeout: int = 60) -> int:
    """返回 int 版本，失败返回 -1"""
    val = _ffprobe_count(video_path, count_mode, timeout)
    try:
        return int(val)
    except (ValueError, TypeError):
        return -1


def generate_ffprob_output(
    segments_dir: str,
    processed_dir: str,
    output_video: str,
    output_path: str,
    num_segments: int,
    processed_prefix: str = 'upscaled_segment',
):
    """
    对所有 segments/processed 文件执行 ffprobe 统计，输出到 ffprob_outputX.txt
    同时返回结构化数据供 xlsx 使用
    """
    segments_path = Path(segments_dir)
    processed_path = Path(processed_dir)

    def _find_segment_file(proc_dir: Path, seg_idx: int, fallback_prefix: str = None) -> Path:
        """通过 *segment_XXX.mp4 模式匹配查找文件，适配多种命名前缀"""
        pattern = f"*segment_{seg_idx:03d}.mp4"
        matches = list(proc_dir.glob(pattern))
        if matches:
            return matches[0]
        # 回退到旧命名方式
        if fallback_prefix:
            fallback = proc_dir / f"{fallback_prefix}_{seg_idx:03d}.mp4"
        else:
            fallback = proc_dir / f"segment_{seg_idx:03d}.mp4"
        return fallback

    data = {
        'seg_packets': [],
        'seg_frames': [],
        'proc_packets': [],
        'proc_frames': [],
        'src_packets': None,
        'src_frames': None,
    }

    with open(output_path, 'w', encoding='utf-8') as out:
        out.write("=" * 60 + "\n")
        out.write(" ffprobe 帧数统计 — RealESRGAN 超分处理\n")
        out.write(f" 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        out.write(f" segments_dir: {segments_dir}\n")
        out.write(f" processed_dir: {processed_dir}\n")
        out.write(f" output_video: {output_video}\n")
        out.write(f" num_segments: {num_segments}\n")
        out.write("=" * 60 + "\n\n")

        # ── segments: nb_read_packets ──
        out.write("── segments: nb_read_packets ──\n")
        for i in range(num_segments):
            seg_file = _find_segment_file(segments_path, i, fallback_prefix='segment')
            seg_filename = seg_file.name
            out.write(f"ffprobe -v error -select_streams v:0 "
                      f"-count_packets -show_entries stream=nb_read_packets "
                      f"-of csv=p=0 {seg_filename}\n")
            if seg_file.exists():
                count = _ffprobe_count(str(seg_file), 'packets', timeout=30)
                print(f"  [segments/packets] {seg_filename}: {count}")
                data['seg_packets'].append(count)
                out.write(count + '\n')
            else:
                data['seg_packets'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        out.write("\n── segments: nb_read_frames ──\n")
        for i in range(num_segments):
            seg_file = _find_segment_file(segments_path, i, fallback_prefix='segment')
            seg_filename = seg_file.name
            out.write(f"ffprobe -v error -select_streams v:0 "
                      f"-count_frames -show_entries stream=nb_read_frames "
                      f"-of csv=p=0 {seg_filename}\n")
            if seg_file.exists():
                count = _ffprobe_count(str(seg_file), 'frames', timeout=300)
                print(f"  [segments/frames]  {seg_filename}: {count}")
                data['seg_frames'].append(count)
                out.write(count + '\n')
            else:
                data['seg_frames'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        out.write(f"\n── processed ({processed_prefix}): nb_read_packets ──\n")
        for i in range(num_segments):
            proc_file = _find_segment_file(processed_path, i, fallback_prefix=processed_prefix)
            proc_filename = proc_file.name
            out.write(f"ffprobe -v error -select_streams v:0 "
                      f"-count_packets -show_entries stream=nb_read_packets "
                      f"-of csv=p=0 {proc_filename}\n")
            if proc_file.exists():
                count = _ffprobe_count(str(proc_file), 'packets', timeout=30)
                print(f"  [processed/packets] {proc_filename}: {count}")
                data['proc_packets'].append(count)
                out.write(count + '\n')
            else:
                data['proc_packets'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        out.write(f"\n── processed ({processed_prefix}): nb_read_frames ──\n")
        for i in range(num_segments):
            proc_file = _find_segment_file(processed_path, i, fallback_prefix=processed_prefix)
            proc_filename = proc_file.name
            out.write(f"ffprobe -v error -select_streams v:0 "
                      f"-count_frames -show_entries stream=nb_read_frames "
                      f"-of csv=p=0 {proc_filename}\n")
            if proc_file.exists():
                count = _ffprobe_count(str(proc_file), 'frames', timeout=300)
                print(f"  [processed/frames]  {proc_filename}: {count}")
                data['proc_frames'].append(count)
                out.write(count + '\n')
            else:
                data['proc_frames'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        # ── 输出视频 ──
        if output_video and Path(output_video).exists():
            out_name = Path(output_video).name
            out.write(f"\n── output video: {out_name} ──\n")
            out.write(f"ffprobe -v error -select_streams v:0 "
                      f"-count_packets -show_entries stream=nb_read_packets "
                      f"-of csv=p=0 '{out_name}'\n")
            count = _ffprobe_count(str(output_video), 'packets', timeout=30)
            print(f"  [output/packets] {out_name}: {count}")
            data['src_packets'] = count
            out.write(count + '\n')

            out.write(f"ffprobe -v error -select_streams v:0 "
                      f"-count_frames -show_entries stream=nb_read_frames "
                      f"-of csv=p=0 '{out_name}'\n")
            count = _ffprobe_count(str(output_video), 'frames', timeout=1800)
            print(f"  [output/frames]  {out_name}: {count}")
            data['src_frames'] = count
            out.write(count + '\n')

        out.write("\n" + "=" * 60 + "\n")

    print(f"✅ ffprob_output 已生成: {output_path}")
    return data


# ═══════════════════════════════════════════════
# 4. xlsx 生成
# ═══════════════════════════════════════════════

def _to_int(val):
    """安全转 int"""
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def generate_xlsx(
    log_info: dict,
    ffprobe_data: dict,
    ffmpeg_frames: list,
    ffmpeg_errors: list,
    ffmpeg_dup_drop: list,
    output_path: str,
):
    """
    生成 RealESRGAN 超分 帧数差值统计汇总 Excel

    核心列:
      分段序号 | 分段编号 | 原始帧数(segment) | 期望输出帧数(=原始) |
      日志提示输出帧数 | 差值1 |
      ffprob -count_packets | 差值2 |
      ffprob -count_frames | 差值3 |
      ffmpeg 拆帧 | 差值4 |
      ffmpeg dup | ffmpeg drop |
      异常检测(花屏/丢帧/重复帧) |
      输出时长 | 处理耗时 |
      性能统计(总秒/平均ms/FPS) | GFPGAN模式 | 输出大小 | 模型加载耗时
    """
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl 未安装，跳过 xlsx 生成")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "帧数差值统计汇总"

    # ── 样式 ──
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='微软雅黑', size=14, bold=True, color='1F3864')
    normal_font = Font(name='微软雅黑', size=10)
    warn_font = Font(name='微软雅黑', size=10, color='CC6600')
    bad_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
    ok_font = Font(name='微软雅黑', size=10, color='008000')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bad_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')

    # 列定义（共 22 列）
    headers = [
        '分段序号',
        '分段编号',
        '原始帧数\n(segment)',
        '期望输出帧数\n(=原始帧)',
        '日志提示\n输出帧数',
        '差值1\n(日志-期望)',
        'ffprob\n-count_packets',
        '差值2\n(ffprob包-期望)',
        'ffprob\n-count_frames',
        '差值3\n(ffprob帧-期望)',
        'ffmpeg\n拆帧',
        '差值4\n(ffmpeg-期望)',
        'ffmpeg\ndup',
        'ffmpeg\ndrop',
        '异常检测\n(花屏/丢帧/重复帧)',
        '输出时长',
        '处理耗时',
        '性能统计\n(总秒/平均ms/FPS)',
        'GFPGAN\n模式',
        '输出大小',
        '模型加载\n耗时',
        '输出文件名',
    ]
    col_widths = [8, 12, 12, 12, 12, 12, 14, 13, 14, 13, 12, 13, 9, 9, 28, 13, 12, 22, 12, 12, 12, 36]
    num_cols = len(headers)

    # ── 标题 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = 'RealESRGAN 超分处理 — 帧数差值统计汇总'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # ── 概览信息 ──
    row = 4
    perf_summary = ''
    if log_info.get('total_perf_time') is not None:
        perf_summary = (f"总{log_info['total_perf_time']:.1f}s | "
                        f"平均{log_info.get('total_perf_avg_ms', 0):.1f}ms | "
                        f"FPS {log_info.get('total_perf_fps', 0):.2f}")
    info_items = [
        ('源视频', log_info.get('source_video', 'N/A')),
        ('输出视频', log_info.get('output_video', 'N/A')),
        ('分段数', str(log_info.get('num_segments', 0))),
        ('分段时长', f"{log_info.get('segment_duration', 300)}秒"),
        ('GFPGAN模式', log_info.get('gfpgan_mode', 'N/A')),
        ('输出大小', log_info.get('total_output_size', 'N/A')),
        ('总耗时(含加载)', log_info.get('total_model_load_time', 'N/A')),
        ('已用时(估算)', log_info.get('total_elapsed_estimate', 'N/A')),
        ('预计剩余', log_info.get('total_remaining_estimate', 'N/A')),
        ('性能统计汇总', perf_summary if perf_summary else 'N/A'),
    ]
    for key, val in info_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws[f'A{row}'] = key
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws[f'C{row}'] = val
        ws[f'C{row}'].font = normal_font
        row += 1

    row += 1  # 空行

    # ── 核心表头 ──
    header_row = row
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 56
    row += 1

    # ── 数据行 ──
    segs = log_info.get('segments', [])

    totals = {
        'raw': 0, 'expected': 0, 'log_output': 0,
        'ffprob_packets': 0, 'ffprob_frames': 0, 'ffmpeg_frames': 0,
        'diff1_sum': 0, 'diff2_sum': 0, 'diff3_sum': 0, 'diff4_sum': 0,
        'dup_sum': 0, 'drop_sum': 0,
        'perf_time_sum': 0.0, 'perf_fps_list': [],
        'bad_segments': 0, 'warn_segments': 0,
    }

    for i, seg in enumerate(segs):
        # 原始帧数：优先使用 ffprobe 统计的 segment 帧数，回退到日志解析值
        seg_frames_ffprobe = _to_int(ffprobe_data['seg_packets'][i]) if i < len(ffprobe_data['seg_packets']) else None
        raw_from_log = seg.get('raw_frames') or 0
        raw = seg_frames_ffprobe if seg_frames_ffprobe is not None else raw_from_log
        # 超分：期望输出帧数 = 原始帧数
        expected = raw
        log_output = seg.get('output_frames') or 0
        diff1 = log_output - expected

        ffprob_pkt = _to_int(ffprobe_data['proc_packets'][i]) if i < len(ffprobe_data['proc_packets']) else None
        diff2 = (ffprob_pkt - expected) if ffprob_pkt is not None else None

        ffprob_frm = _to_int(ffprobe_data['proc_frames'][i]) if i < len(ffprobe_data['proc_frames']) else None
        diff3 = (ffprob_frm - expected) if ffprob_frm is not None else None

        ffmpeg_val = ffmpeg_frames[i] if i < len(ffmpeg_frames) else None
        diff4 = (ffmpeg_val - expected) if ffmpeg_val is not None else None

        dup_val = None
        drop_val = None
        if i < len(ffmpeg_dup_drop):
            dd = ffmpeg_dup_drop[i]
            if dd:
                dup_val = dd.get('dup')
                drop_val = dd.get('drop')
                if dup_val is not None:
                    totals['dup_sum'] += dup_val
                if drop_val is not None:
                    totals['drop_sum'] += drop_val

        # ── 异常检测 ──
        errors = ffmpeg_errors[i] if i < len(ffmpeg_errors) else []
        has_illegal = any('illegal' in e.lower() for e in errors)
        has_missing = any('missing' in e.lower() for e in errors)
        has_duplicate = any('more than 1000 frames duplicated' in e.lower() for e in errors)
        has_no_frame = any('no frame' in e.lower() for e in errors)
        has_decode_error = any('decode_slice_header error' in e.lower() for e in errors)
        has_pps_error = any('non-existing pps' in e.lower() for e in errors)

        anomaly_parts = []

        if has_illegal:
            anomaly_parts.append('⚠ illegal buffer(花屏)')
        if has_pps_error:
            anomaly_parts.append('⚠ PPS丢失')
        if has_decode_error:
            anomaly_parts.append('⚠ 解码错误')

        if has_no_frame:
            anomaly_parts.append('❌ no frame!(丢帧)')
        if has_missing:
            anomaly_parts.append('⚠ missing ref(丢帧)')

        if has_duplicate:
            anomaly_parts.append('🔁 大量重复帧(>1000)')
        if dup_val is not None and dup_val > 100:
            anomaly_parts.append(f'🔁 dup={dup_val}')

        # 帧数偏差（超分场景，期望帧数=原始帧数，偏差应严格为 0）
        if diff1 is not None and abs(diff1) > 0:
            anomaly_parts.append(f'📉 日志偏差={diff1}')
        if diff2 is not None and abs(diff2) > 2:
            anomaly_parts.append(f'📉 包偏差={diff2}')
        if diff3 is not None and abs(diff3) > 2:
            anomaly_parts.append(f'📉 帧偏差={diff3}')
        if diff4 is not None and abs(diff4) > 2:
            anomaly_parts.append(f'📉 拆帧偏差={diff4}')

        is_bad = any(kw in ''.join(anomaly_parts) for kw in ['❌', '⚠ illegal', '⚠ PPS'])
        is_warn = (not is_bad) and len(anomaly_parts) > 0
        anomaly_str = '\n'.join(anomaly_parts) if anomaly_parts else '✅ 正常'

        if is_bad:
            totals['bad_segments'] += 1
        elif is_warn:
            totals['warn_segments'] += 1

        # 性能统计字符串
        perf_str = ''
        if seg.get('perf_total_sec') is not None:
            perf_str = (f"{seg['perf_total_sec']:.1f}s / "
                        f"{seg.get('perf_avg_ms', 0):.1f}ms / "
                        f"FPS {seg.get('perf_fps', 0):.2f}")

        # 输出文件名（仅 basename）
        out_file_name = ''
        if seg.get('output_file'):
            out_file_name = Path(seg['output_file']).name

        row_data = [
            i + 1,
            f"segment_{seg['index']:03d}",
            raw,
            expected,
            log_output,
            f"{diff1}" if diff1 is not None else 'N/A',
            ffprob_pkt if ffprob_pkt is not None else 'N/A',
            f"{diff2}" if diff2 is not None else 'N/A',
            ffprob_frm if ffprob_frm is not None else 'N/A',
            f"{diff3}" if diff3 is not None else 'N/A',
            ffmpeg_val if ffmpeg_val is not None else 'N/A',
            f"{diff4}" if diff4 is not None else 'N/A',
            dup_val if dup_val is not None else '',
            drop_val if drop_val is not None else '',
            anomaly_str,
            seg.get('output_duration', ''),
            seg.get('elapsed', ''),
            perf_str,
            seg.get('gfpgan_mode', 'N/A'),
            seg.get('output_size', ''),
            seg.get('model_load_time', ''),
            out_file_name,
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = center_align if ci != 15 else Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border

            # 异常列着色
            if ci == 15:
                if is_bad:
                    cell.font = bad_font
                    cell.fill = bad_fill
                elif is_warn:
                    cell.font = warn_font
                    cell.fill = warn_fill
                else:
                    cell.font = ok_font
            # 差值列着色 (6=差值1, 8=差值2, 10=差值3, 12=差值4)
            elif ci in (6, 8, 10, 12):
                try:
                    dv = float(str(val).split()[0]) if val not in ('', 'N/A', None) else 0
                except (ValueError, TypeError):
                    dv = 0
                if abs(dv) > 2:
                    cell.font = bad_font
                    cell.fill = bad_fill
                elif abs(dv) > 0.5:
                    cell.font = warn_font
                    cell.fill = warn_fill
                else:
                    cell.font = ok_font
            else:
                cell.font = normal_font

        totals['raw'] += raw
        totals['expected'] += expected
        totals['log_output'] += log_output
        if ffprob_pkt is not None:
            totals['ffprob_packets'] += ffprob_pkt
        if ffprob_frm is not None:
            totals['ffprob_frames'] += ffprob_frm
        if ffmpeg_val is not None:
            totals['ffmpeg_frames'] += ffmpeg_val
        if diff1 is not None:
            totals['diff1_sum'] += diff1
        if diff2 is not None:
            totals['diff2_sum'] += diff2
        if diff3 is not None:
            totals['diff3_sum'] += diff3
        if diff4 is not None:
            totals['diff4_sum'] += diff4
        if seg.get('perf_total_sec') is not None:
            totals['perf_time_sum'] += seg['perf_total_sec']
        if seg.get('perf_fps') is not None:
            totals['perf_fps_list'].append(seg['perf_fps'])

        ws.row_dimensions[row].height = max(30, len(anomaly_parts) * 16 + 14)
        row += 1

    # ── 汇总行 ──
    sum_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    sum_font = Font(name='微软雅黑', size=10, bold=True)

    diff1_total = totals['diff1_sum']
    diff2_total = totals['diff2_sum']
    diff3_total = totals['diff3_sum']
    diff4_total = totals['diff4_sum']

    anomaly_summary_parts = []
    if totals['bad_segments'] > 0:
        anomaly_summary_parts.append(f"❌ 花屏/丢帧: {totals['bad_segments']}/{len(segs)}")
    if totals['warn_segments'] > 0:
        anomaly_summary_parts.append(f"⚠ 警告: {totals['warn_segments']}/{len(segs)}")
    if totals['dup_sum'] > 0:
        anomaly_summary_parts.append(f"总dup={totals['dup_sum']}")
    if totals['drop_sum'] > 0:
        anomaly_summary_parts.append(f"总drop={totals['drop_sum']}")
    anomaly_summary = '\n'.join(anomaly_summary_parts) if anomaly_summary_parts else '✅ 全部正常'

    perf_avg_fps = (sum(totals['perf_fps_list']) / len(totals['perf_fps_list'])) if totals['perf_fps_list'] else 0
    perf_sum_str = (f"总{totals['perf_time_sum']:.1f}s | FPS均值{perf_avg_fps:.2f}"
                    if totals['perf_time_sum'] > 0 else '')

    sum_data = [
        '汇总', '',
        totals['raw'],
        totals['expected'],
        totals['log_output'],
        f"{diff1_total}",
        totals['ffprob_packets'] if totals['ffprob_packets'] else 'N/A',
        str(diff2_total) if diff2_total is not None else 'N/A',
        totals['ffprob_frames'] if totals['ffprob_frames'] else 'N/A',
        str(diff3_total) if diff3_total is not None else 'N/A',
        totals['ffmpeg_frames'] if totals['ffmpeg_frames'] else 'N/A',
        str(diff4_total) if diff4_total is not None else 'N/A',
        totals['dup_sum'],
        totals['drop_sum'],
        anomaly_summary,
        log_info.get('total_elapsed_estimate', ''),
        log_info.get('total_model_load_time', ''),
        perf_sum_str,
        log_info.get('gfpgan_mode', 'N/A'),
        log_info.get('total_output_size', ''),
        log_info.get('total_model_load_time', ''),
        '',
    ]
    for ci, val in enumerate(sum_data, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = sum_font
        cell.fill = sum_fill
        cell.alignment = center_align if ci != 15 else Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row].height = max(30, len(anomaly_summary_parts) * 16 + 14)

    ws.freeze_panes = f'A{header_row + 1}'

    # ── Sheet 2: 详细错误日志 ──
    ws2 = wb.create_sheet("ffmpeg错误详情")

    ws2.merge_cells('A1:E1')
    ws2['A1'] = 'ffmpeg 解码 — 逐段错误详情 (RealESRGAN)'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    headers2 = ['分段', 'ffmpeg拆帧帧数', 'dup', 'drop', '错误/警告信息']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 80
    ws2.row_dimensions[3].height = 28

    r = 4
    for i in range(len(segs)):
        seg_name = f"upscaled_{i:03d}"
        ffmpeg_val = ffmpeg_frames[i] if i < len(ffmpeg_frames) else 'N/A'
        dd = ffmpeg_dup_drop[i] if i < len(ffmpeg_dup_drop) else {}
        dup_val = dd.get('dup', '') if dd else ''
        drop_val = dd.get('drop', '') if dd else ''
        errors = ffmpeg_errors[i] if i < len(ffmpeg_errors) else []
        err_text = '\n'.join(errors[:30]) if errors else '✅ 无错误'

        data2 = [seg_name, ffmpeg_val, dup_val, drop_val, err_text]
        for ci, val in enumerate(data2, 1):
            cell = ws2.cell(row=r, column=ci, value=val)
            cell.font = bad_font if (errors and ci == 5) else normal_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True) if ci == 5 else center_align
            cell.border = thin_border
        ws2.row_dimensions[r].height = max(30, min(300, len(errors) * 16))
        r += 1

    ws2.freeze_panes = 'A4'

    # ── Sheet 3: 性能统计明细 ──
    ws3 = wb.create_sheet("性能统计明细")

    ws3.merge_cells('A1:H1')
    ws3['A1'] = 'RealESRGAN 性能统计明细'
    ws3['A1'].font = title_font
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 28

    headers3 = ['分段', '总时间(秒)', '平均(ms)', 'FPS', '输出时长',
                '处理耗时', '模型加载耗时', 'GFPGAN模式']
    widths3 = [14, 14, 12, 12, 14, 14, 16, 14]
    for ci, (h, w) in enumerate(zip(headers3, widths3), 1):
        cell = ws3.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[3].height = 28

    r = 4
    for i, seg in enumerate(segs):
        data3 = [
            f"upscaled_{seg['index']:03d}",
            seg.get('perf_total_sec', ''),
            seg.get('perf_avg_ms', ''),
            seg.get('perf_fps', ''),
            seg.get('output_duration', ''),
            seg.get('elapsed', ''),
            seg.get('model_load_time', ''),
            seg.get('gfpgan_mode', ''),
        ]
        for ci, val in enumerate(data3, 1):
            cell = ws3.cell(row=r, column=ci, value=val)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border
        r += 1

    # 汇总行
    avg_fps = (sum(totals['perf_fps_list']) / len(totals['perf_fps_list'])) if totals['perf_fps_list'] else 0
    sum3 = ['汇总', f"{totals['perf_time_sum']:.1f}", '', f"{avg_fps:.2f}", '', '', '', '']
    for ci, val in enumerate(sum3, 1):
        cell = ws3.cell(row=r, column=ci, value=val)
        cell.font = sum_font
        cell.fill = sum_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws3.freeze_panes = 'A4'

    wb.save(output_path)
    print(f"✅ xlsx 已生成: {output_path}")


# ═══════════════════════════════════════════════
# 5. 主函数
# ═══════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='RealESRGAN 超分处理结果检测分析脚本',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('log_file', help='RealESRGAN 流水线输出日志文件')
    parser.add_argument('--suffix', default='esrgan', help='输出文件后缀 (默认: esrgan)')
    parser.add_argument('--segments-dir', default=None,
                        help='segments 目录 (默认: 日志同目录下的 segments/)')
    parser.add_argument('--processed-dir', default=None,
                        help='processed 目录 (默认: 日志同目录下的 processed/)')
    parser.add_argument('--processed-prefix', default='upscaled_segment',
                        help='处理后文件前缀 (默认: upscaled_segment)')
    parser.add_argument('--output-video', default=None,
                        help='输出视频文件路径（由 processed 分段合成而来）')
    parser.add_argument('--output-dir', default=None,
                        help='输出目录 (默认: 日志同目录)')
    parser.add_argument('--image2pipe', action='store_true',
                        help='使用 -f image2pipe 模式（完整解码+编码管线，慢但能捕获 dup/drop；默认用 -f null 快速模式）')
    parser.add_argument('--skip-ffmpeg', action='store_true',
                        help='跳过 ffmpeg 解码')
    parser.add_argument('--skip-ffprobe', action='store_true',
                        help='跳过 ffprobe 帧数统计')
    parser.add_argument('--skip-xlsx', action='store_true',
                        help='跳过 xlsx 生成')

    args = parser.parse_args()

    log_path = Path(args.log_file).resolve()
    if not log_path.exists():
        print(f"[ERROR] 日志文件不存在: {log_path}")
        sys.exit(1)

    log_dir = log_path.parent
    output_dir = Path(args.output_dir) if args.output_dir else log_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    segments_dir = args.segments_dir or str(log_dir / 'segments')
    processed_dir = args.processed_dir or str(log_dir / 'processed')
    suffix = args.suffix
    processed_prefix = args.processed_prefix

    # ── 解析日志 ──
    print(f"📖 解析日志: {log_path}")
    log_info = parse_output_log(str(log_path))
    print(f"   分段数: {log_info['num_segments']}")
    print(f"   GFPGAN 模式: {log_info['gfpgan_mode']}")
    if log_info.get('total_perf_time') is not None:
        print(f"   性能统计: 总{log_info['total_perf_time']:.1f}s | "
              f"FPS {log_info.get('total_perf_fps', 0):.2f}")

    output_video = args.output_video or log_info.get('output_video', '')
    if not output_video:
        # 在日志目录或上一级查找可能的最终输出视频
        for cand in [log_dir, log_dir.parent]:
            if cand.exists():
                for f in cand.iterdir():
                    if (f.suffix.lower() in ('.mp4', '.avi', '.mkv')
                            and f.is_file()
                            and 'segment' not in f.name.lower()
                            and 'upscaled_segment' not in f.name.lower()):
                        output_video = str(f)
                        break
            if output_video:
                break

    num_segs = log_info['num_segments']

    ffmpeg_frames = []
    ffmpeg_errors = []
    ffmpeg_dup_drop = []
    ffprobe_data = {
        'seg_packets': [], 'seg_frames': [],
        'proc_packets': [], 'proc_frames': [],
        'src_packets': None, 'src_frames': None,
    }

    # ── 生成 ffmpeg_output ──
    if not args.skip_ffmpeg:
        ffmpeg_path = output_dir / f"ffmpeg_output_{suffix}.txt"
        mode_label = "image2pipe" if args.image2pipe else "null"
        print(f"\n🎬 生成 ffmpeg 输出 ({mode_label} 解码模式): {ffmpeg_path}")
        try:
            ffmpeg_frames, ffmpeg_errors, ffmpeg_dup_drop = generate_ffmpeg_output(
                processed_dir, str(ffmpeg_path), num_segs,
                use_image2pipe=args.image2pipe,
                processed_prefix=processed_prefix,
            )
        except Exception as e:
            print(f"[WARN] ffmpeg 生成异常: {e}，继续后续步骤")

    # ── 生成 ffprob_output ──
    if not args.skip_ffprobe:
        ffprob_path = output_dir / f"ffprob_output_{suffix}.txt"
        print(f"\n🔍 生成 ffprobe 输出: {ffprob_path}")
        try:
            ffprobe_data = generate_ffprob_output(
                segments_dir, processed_dir, output_video,
                str(ffprob_path), num_segs,
                processed_prefix=processed_prefix,
            )
        except Exception as e:
            print(f"[WARN] ffprobe 生成异常: {e}，继续后续步骤")

    # ── 生成 xlsx ──
    if not args.skip_xlsx:
        xlsx_path = output_dir / f"帧数差值统计汇总_output_{suffix}.xlsx"
        print(f"\n📊 生成 xlsx: {xlsx_path}")
        try:
            generate_xlsx(
                log_info, ffprobe_data,
                ffmpeg_frames, ffmpeg_errors, ffmpeg_dup_drop,
                str(xlsx_path)
            )
        except Exception as e:
            print(f"[ERROR] xlsx 生成失败: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "=" * 60)
    print("  ✅ 所有检测分析文件生成完毕！(RealESRGAN)")
    print("=" * 60)


if __name__ == '__main__':
    main()
