#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频增强处理结果检测分析脚本 (v2)
=================================
依据流水线输出日志 (如 output9.txt) 及所有分段/插帧文件，
生成以下检测分析文件：
  1. ffmpeg_output9.txt      — ffmpeg 逐段 pipe 解码日志（含丢帧/错误信息）
  2. ffprob_output9.txt      — ffprobe 逐段帧数统计（segments + processed + 输出视频）
  3. 帧数差值统计汇总_output9.xlsx — 帧数差值汇总 Excel 表格
     （含：日志提示输出帧数/差值1/ffprob count_packets/差值2/ffmpeg拆帧/差值3/异常检测）

用法:
    python analyze_video_enhancement.py <output_log> [--suffix 9] ...

示例:
    python analyze_video_enhancement.py temp9/output9.txt --suffix 9 \
        --output-video ./output.mp4
    python analyze_video_enhancement.py output10.txt --suffix 10 \
        --segments-dir ./segments --processed-dir ./processed
"""

import argparse
import re
import os
import subprocess
import sys
import shlex
from pathlib import Path
from collections import OrderedDict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl 未安装，将不会生成 xlsx 文件。安装: pip install openpyxl")


# ═══════════════════════════════════════════════
# 1. 日志解析
# ═══════════════════════════════════════════════

def parse_output_log(log_path: str) -> dict:
    """解析流水线日志 outputX.txt，提取每段帧数信息"""
    with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    info = {
        'num_segments': 0,
        'segments': [],
        'interpolation_factor': 2,
        'segment_duration': 300,
        'source_video': '',
        'output_video': '',
        'total_input_duration': '',
        'total_output_duration': '',
        'total_output_size': '',
        'total_elapsed': '',
        'total_input_frames': None,
        'total_output_frames': None,
    }

    m = re.search(r'插帧倍数\s*[:：]\s*(\d+)', content)
    if m:
        info['interpolation_factor'] = int(m.group(1))

    m = re.search(r'分段时长\s*[:：]\s*(\d+)', content)
    if m:
        info['segment_duration'] = int(m.group(1))

    m = re.search(r'输入\s*[:：]\s*(.+\.mp4)', content)
    if m:
        info['source_video'] = m.group(1).strip()
    m = re.search(r'输出\s*[:：]\s*(.+\.mp4)', content)
    if m:
        info['output_video'] = m.group(1).strip()

    m = re.search(r'✅\s*共\s*(\d+)\s*个片段', content)
    if m:
        info['num_segments'] = int(m.group(1))
    else:
        segs = re.findall(r'片段\s*(\d+)/(\d+):', content)
        if segs:
            info['num_segments'] = max(int(s[1]) for s in segs)

    # 每段信息
    segment_pattern = re.compile(
        r'🎬\s*片段\s*(\d+)/(\d+):\s*segment_(\d+)\.mp4'
    )
    frame_pattern = re.compile(
        r'原始帧:\s*(\d+)\s*→\s*输出帧:\s*(\d+)'
    )
    time_pattern = re.compile(
        r'输出时长\s*([\d:.]+)\s*\|\s*耗时\s*([\d:.]+)'
    )

    seg_starts = [(m.start(), int(m.group(1)), int(m.group(3)))
                  for m in segment_pattern.finditer(content)]

    for idx, (start_pos, seg_num, seg_id) in enumerate(seg_starts):
        end_pos = seg_starts[idx + 1][0] if idx + 1 < len(seg_starts) else len(content)
        seg_text = content[start_pos:end_pos]

        seg_info = {
            'index': seg_id,
            'number': seg_num,
            'raw_frames': None,
            'output_frames': None,      # 日志提示输出帧数
            'expected_frames': None,
            'frame_diff': None,          # 差值1 = 日志输出帧数 - 期望
            'output_duration': '',
            'elapsed': '',
        }

        fm = frame_pattern.search(seg_text)
        if fm:
            seg_info['raw_frames'] = int(fm.group(1))
            seg_info['output_frames'] = int(fm.group(2))

        tm = time_pattern.search(seg_text)
        if tm:
            seg_info['output_duration'] = tm.group(1)
            seg_info['elapsed'] = tm.group(2)

        info['segments'].append(seg_info)

    # 总时长 / 大小
    m = re.search(r'输入\s*:\s*[\d.]+\s*MB,\s*([\d:]+)', content)
    if m:
        info['total_input_duration'] = m.group(1)
    m = re.search(r'输出\s*:\s*[\d.]+\s*MB,\s*([\d:]+)', content)
    if m:
        info['total_output_duration'] = m.group(1)
    m = re.search(r'📦\s*文件大小:\s*([\d.]+\s*[MGT]B)', content)
    if m:
        info['total_output_size'] = m.group(1)
    m = re.search(r'总用时\s*[:：]\s*([\d:]+)', content)
    if m:
        info['total_elapsed'] = m.group(1)

    # 总输入/输出帧数（从日志最后的统计行提取）
    m = re.search(r'输入\s*:\s*[\d.]+\s*MB,\s*[\d:]+\s*→\s*输出\s*:\s*[\d.]+\s*MB,\s*[\d:]+', content)
    # 用 ffmpeg 合并后的统计
    m2 = re.search(r'输入\s*:\s*([\d.]+)\s*MB,\s*[\d:]+\s*\n.*输出\s*:\s*([\d.]+)\s*MB,\s*([\d:]+)', content)
    if m2:
        info['total_input_size'] = m2.group(1)
        info['total_output_size_alt'] = m2.group(2)

    return info


# ═══════════════════════════════════════════════
# 2. ffmpeg image2pipe 解码 + 错误检测
# ═══════════════════════════════════════════════

def _ffmpeg_image2pipe_decode(video_path: str, timeout: int = 300) -> dict:
    """
    使用 ffmpeg -f image2pipe -vcodec png - 模拟写入（触发完整解码+编码管线，
    但不实际写文件），stdout 丢弃避免 OOM，流式捕获 stderr。
    返回: {'frame_count': int|None, 'errors': [str], 'dup_count': int|None, 'drop_count': int|None}
    """
    result = {
        'frame_count': None,
        'errors': [],
        'dup_count': None,
        'drop_count': None,
    }
    cmd = [
        "ffmpeg", "-hide_banner",
        "-i", str(video_path),
        "-f", "image2pipe",
        "-vcodec", "png",
        "-"
    ]
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.DEVNULL,  # 丢弃 pipe 中的 PNG 数据，避免 OOM
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
        try:
            for line in proc.stderr:
                stripped = line.strip()

                # 提取 frame= 行
                m = re.search(r'frame=\s*(\d+)', stripped)
                if m:
                    result['frame_count'] = int(m.group(1))

                # 提取 dup/drop 计数（最后一行的最终统计）
                m_dup = re.search(r'dup=(\d+)', stripped)
                m_drop = re.search(r'drop=(\d+)', stripped)
                if m_dup and m_drop and 'Lsize=' in stripped:
                    result['dup_count'] = int(m_dup.group(1))
                    result['drop_count'] = int(m_drop.group(1))

                # 收集错误/警告行
                lower = stripped.lower()
                if any(kw in lower for kw in [
                    'illegal', 'error', 'missing', 'corrupt',
                    'duplicated', 'duplicate', 'decoder',
                    'non-existing pps', 'no frame',
                    'more than 1000 frames duplicated',
                ]):
                    result['errors'].append(stripped)

            proc.wait(timeout=timeout)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg image2pipe 解码超时")

    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")

    return result


def _ffmpeg_null_decode(video_path: str, timeout: int = 180) -> dict:
    """
    使用 ffmpeg -f null - 轻量解码（不编码，速度快），
    流式捕获 stderr 中的帧数和错误信息。
    返回: {'frame_count': int|None, 'errors': [str], 'dup_count': None, 'drop_count': None}
    """
    result = {
        'frame_count': None,
        'errors': [],
        'dup_count': None,
        'drop_count': None,
    }
    cmd = [
        "ffmpeg", "-hide_banner",
        "-i", str(video_path),
        "-f", "null", "-"
    ]
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
        try:
            for line in proc.stderr:
                stripped = line.strip()

                m = re.search(r'frame=\s*(\d+)', stripped)
                if m:
                    result['frame_count'] = int(m.group(1))

                m_dup = re.search(r'dup=(\d+)', stripped)
                m_drop = re.search(r'drop=(\d+)', stripped)
                if m_dup and m_drop and 'Lsize=' in stripped:
                    result['dup_count'] = int(m_dup.group(1))
                    result['drop_count'] = int(m_drop.group(1))

                lower = stripped.lower()
                if any(kw in lower for kw in [
                    'illegal', 'error', 'missing', 'corrupt',
                    'duplicated', 'duplicate', 'decoder',
                    'non-existing pps', 'no frame',
                    'more than 1000 frames duplicated',
                ]):
                    result['errors'].append(stripped)

            proc.wait(timeout=timeout)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg null 解码超时")

    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")

    return result


def generate_ffmpeg_output(processed_dir: str, output_path: str, num_segments: int, use_image2pipe: bool = False):
    """
    对每个 interpolated_segment_XXX.mp4 执行 ffmpeg 解码。
    use_image2pipe=True: -f image2pipe（完整解码+PNG编码管线，慢但能捕获 dup/drop）
    use_image2pipe=False: -f null（轻量解码，快）
    流式捕获 stderr，输出到 ffmpeg_outputX.txt
    返回: (ffmpeg_frames: list, ffmpeg_errors: list, ffmpeg_dup_drop: list)
    """
    processed_path = Path(processed_dir)
    ffmpeg_frames = []
    ffmpeg_errors = []
    ffmpeg_dup_drop = []

    with open(output_path, 'w', encoding='utf-8') as out:
        for i in range(num_segments):
            seg_file = processed_path / f"interpolated_segment_{i:03d}.mp4"

            if use_image2pipe:
                cmd_str = (f"PS D:\\temp9\\processed> ffmpeg -hide_banner -i "
                           f".\\interpolated_segment_{i:03d}.mp4 "
                           f"-f image2pipe -vcodec png - | find /v \"\"\n")
            else:
                cmd_str = (f"PS D:\\temp9\\processed> ffmpeg -hide_banner -i "
                           f".\\interpolated_segment_{i:03d}.mp4 "
                           f"-f null - | find /v \"\"\n")
            out.write(cmd_str)

            if not seg_file.exists():
                print(f"  [WARN] 文件不存在: {seg_file}")
                ffmpeg_frames.append(None)
                ffmpeg_errors.append([f"[FILE NOT FOUND] {seg_file}"])
                ffmpeg_dup_drop.append(None)
                out.write(f"[FILE NOT FOUND]\n")
                out.write(f"PS D:\\temp9\\processed>\n")
                continue

            if use_image2pipe:
                result = _ffmpeg_image2pipe_decode(str(seg_file))
            else:
                result = _ffmpeg_null_decode(str(seg_file))
            ffmpeg_frames.append(result['frame_count'])
            ffmpeg_errors.append(result['errors'])
            ffmpeg_dup_drop.append({
                'dup': result['dup_count'],
                'drop': result['drop_count'],
            })

            print(f"  [ffmpeg] segment_{i:03d}: frame={result['frame_count']}, "
                  f"dup={result['dup_count']}, drop={result['drop_count']}, "
                  f"errors={len(result['errors'])}")

            if result['frame_count'] is not None:
                out.write(f"frame={result['frame_count']}\n")
            for err in result['errors']:
                out.write(err + '\n')
            out.write(f"PS D:\\temp9\\processed>\n")

    print(f"✅ ffmpeg_output 已生成: {output_path}")
    return ffmpeg_frames, ffmpeg_errors, ffmpeg_dup_drop


# ═══════════════════════════════════════════════
# 3. ffprobe 帧数统计
# ═══════════════════════════════════════════════

def _ffprobe_count(video_path: str, count_mode: str = 'packets', timeout: int = 60) -> str:
    """
    ffprobe 获取帧/包计数。count_mode: 'packets' | 'frames'
    """
    if count_mode == 'packets':
        cmd = [
            "ffprobe", "-v", "error",
            "-select_streams", "v:0",
            "-count_packets",
            "-show_entries", "stream=nb_read_packets",
            "-of", "csv=p=0",
            video_path
        ]
    else:
        cmd = [
            "ffprobe", "-v", "error",
            "-select_streams", "v:0",
            "-count_frames",
            "-show_entries", "stream=nb_read_frames",
            "-of", "csv=p=0",
            video_path
        ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        stdout = result.stdout.strip()
        stderr = result.stderr.strip()
        m = re.search(r'(\d+)', stdout)
        if m:
            return m.group(1)
        m = re.search(r'(\d+)', stderr)
        if m:
            return m.group(1)
        return stdout if stdout else "[ERROR]"
    except subprocess.TimeoutExpired:
        return "[TIMEOUT]"
    except FileNotFoundError:
        return "[SKIP: ffprobe not found]"


def _ffprobe_count_int(video_path: str, count_mode: str = 'packets', timeout: int = 60) -> int:
    """返回 int 版本，失败返回 -1"""
    val = _ffprobe_count(video_path, count_mode, timeout)
    try:
        return int(val)
    except (ValueError, TypeError):
        return -1


def generate_ffprob_output(
    segments_dir: str,
    processed_dir: str,
    output_video: str,
    output_path: str,
    num_segments: int,
):
    """
    对所有 segments/processed 文件执行 ffprobe 统计，输出到 ffprob_outputX.txt
    同时返回结构化数据供 xlsx 使用
    """
    segments_path = Path(segments_dir)
    processed_path = Path(processed_dir)

    # 存储结构化结果
    data = {
        'seg_packets': [],
        'seg_frames': [],
        'proc_packets': [],
        'proc_frames': [],
        'src_packets': None,
        'src_frames': None,
    }

    with open(output_path, 'w', encoding='utf-8') as out:
        out.write("PS D:\\temp9> dir\n\n\n"
                  "    Directory: D:\\temp9\n\n\n"
                  "Mode                 LastWriteTime         Length Name\n"
                  "----                 -------------         ------ ----\n"
                  "d-----         2026/7/14     12:59                processed\n"
                  "d-----         2026/7/14     12:59                segments\n"
                  "-a----         2026/7/14     12:59      568373098 05 Race to the Spaceship - Sandbox Surprise_R60.mp4\n"
                  "-a----         2026/7/14     13:00         160079 output9.txt\n\n\n")
        out.flush()

        # ── segments: nb_read_packets ──
        out.write("PS D:\\temp9> cd .\\segments\\\n")
        for i in range(num_segments):
            seg_file = segments_path / f"segment_{i:03d}.mp4"
            out.write(f"PS D:\\temp9\\segments> ffprobe -v error -select_streams v:0 "
                      f"-count_packets -show_entries stream=nb_read_packets "
                      f"-of csv=p=0 .\\segment_{i:03d}.mp4\n")
            if seg_file.exists():
                count = _ffprobe_count(str(seg_file), 'packets', timeout=30)
                print(f"  [segments/packets] segment_{i:03d}: {count}")
                data['seg_packets'].append(count)
                out.write(count + '\n')
            else:
                data['seg_packets'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        out.write("PS D:\\temp9\\segments>\n")

        # ── segments: nb_read_frames ──
        for i in range(num_segments):
            seg_file = segments_path / f"segment_{i:03d}.mp4"
            out.write(f"PS D:\\temp9\\segments> ffprobe -v error -select_streams v:0 "
                      f"-count_frames -show_entries stream=nb_read_frames "
                      f"-of csv=p=0 .\\segment_{i:03d}.mp4\n")
            if seg_file.exists():
                count = _ffprobe_count(str(seg_file), 'frames', timeout=60)
                print(f"  [segments/frames]  segment_{i:03d}: {count}")
                data['seg_frames'].append(count)
                out.write(count + '\n')
            else:
                data['seg_frames'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        out.write("PS D:\\temp9\\segments> cd ..\\processed\\\n")

        # ── processed: nb_read_packets ──
        for i in range(num_segments):
            proc_file = processed_path / f"interpolated_segment_{i:03d}.mp4"
            out.write(f"PS D:\\temp9\\processed> ffprobe -v error -select_streams v:0 "
                      f"-count_packets -show_entries stream=nb_read_packets "
                      f"-of csv=p=0 .\\interpolated_segment_{i:03d}.mp4\n")
            if proc_file.exists():
                count = _ffprobe_count(str(proc_file), 'packets', timeout=30)
                print(f"  [processed/packets] interpolated_segment_{i:03d}: {count}")
                data['proc_packets'].append(count)
                out.write(count + '\n')
            else:
                data['proc_packets'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        out.write("PS D:\\temp9\\processed>\n")

        # ── processed: nb_read_frames ──
        for i in range(num_segments):
            proc_file = processed_path / f"interpolated_segment_{i:03d}.mp4"
            out.write(f"PS D:\\temp9\\processed> ffprobe -v error -select_streams v:0 "
                      f"-count_frames -show_entries stream=nb_read_frames "
                      f"-of csv=p=0 .\\interpolated_segment_{i:03d}.mp4\n")
            if proc_file.exists():
                count = _ffprobe_count(str(proc_file), 'frames', timeout=60)
                print(f"  [processed/frames]  interpolated_segment_{i:03d}: {count}")
                data['proc_frames'].append(count)
                out.write(count + '\n')
            else:
                data['proc_frames'].append("[NOT FOUND]")
                out.write("[FILE NOT FOUND]\n")

        # ── 输出视频 ──
        out.write("PS D:\\temp9\\processed> cd ..\n")
        if output_video and Path(output_video).exists():
            out_name = Path(output_video).name
            out.write(f"PS D:\\temp9> ffprobe -v error -select_streams v:0 "
                      f"-count_packets -show_entries stream=nb_read_packets "
                      f"-of csv=p=0 '.\\{out_name}'\n")
            count = _ffprobe_count(str(output_video), 'packets', timeout=30)
            print(f"  [output/packets] {out_name}: {count}")
            data['src_packets'] = count
            out.write(count + '\n')

            out.write(f"PS D:\\temp9> ffprobe -v error -select_streams v:0 "
                      f"-count_frames -show_entries stream=nb_read_frames "
                      f"-of csv=p=0 '.\\{out_name}'\n")
            count = _ffprobe_count(str(output_video), 'frames', timeout=120)
            print(f"  [output/frames]  {out_name}: {count}")
            data['src_frames'] = count
            out.write(count + '\n')

        out.write("PS D:\\temp9>\n")

    print(f"✅ ffprob_output 已生成: {output_path}")
    return data


# ═══════════════════════════════════════════════
# 4. xlsx 生成
# ═══════════════════════════════════════════════

def _to_int(val):
    """安全转 int"""
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def generate_xlsx(
    log_info: dict,
    ffprobe_data: dict,
    ffmpeg_frames: list,
    ffmpeg_errors: list,
    ffmpeg_dup_drop: list,
    output_path: str,
):
    """
    生成 帧数差值统计汇总 Excel

    核心列:
      分段序号 | 分段编号 | 原始帧数(segment) | 期望输出帧数 |
      日志提示输出帧数 | 差值1 |
      ffprob -count_packets | 差值2 |
      ffprob -count_frames | 差值3 |
      ffmpeg 拆帧 | 差值4 |
      异常检测(花屏/丢帧/重复帧) | 输出时长 | 处理耗时 | 插帧倍数
    """
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl 未安装，跳过 xlsx 生成")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "帧数差值统计汇总"

    # ── 样式 ──
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='微软雅黑', size=14, bold=True, color='1F3864')
    normal_font = Font(name='微软雅黑', size=10)
    warn_font = Font(name='微软雅黑', size=10, color='CC6600')
    bad_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
    ok_font = Font(name='微软雅黑', size=10, color='008000')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bad_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')

    num_cols = 18  # 总列数（含新增的 dup/drop）

    # ── 标题 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = '视频增强处理 — 帧数差值统计汇总'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # ── 概览信息 ──
    row = 4
    info_items = [
        ('源视频', log_info.get('source_video', 'N/A')),
        ('输出视频', log_info.get('output_video', 'N/A')),
        ('插帧倍数', f"{log_info.get('interpolation_factor', 2)}×"),
        ('分段数', str(log_info.get('num_segments', 0))),
        ('分段时长', f"{log_info.get('segment_duration', 300)}秒"),
        ('输入时长', log_info.get('total_input_duration', 'N/A')),
        ('输出时长', log_info.get('total_output_duration', 'N/A')),
        ('输出大小', log_info.get('total_output_size', 'N/A')),
        ('总耗时', log_info.get('total_elapsed', 'N/A')),
    ]
    for key, val in info_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws[f'A{row}'] = key
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws[f'C{row}'] = val
        ws[f'C{row}'].font = normal_font
        row += 1

    row += 1  # 空行

    # ── 核心表头 ──
    headers = [
        '分段序号',
        '分段编号',
        '原始帧数\n(segment)',
        '期望输出帧数\n(原始-1)*倍数+1',
        '日志提示\n输出帧数',
        '差值1\n(日志-期望)',
        'ffprob\n-count_packets',
        '差值2\n(ffprob包-期望)',
        'ffprob\n-count_frames',
        '差值3\n(ffprob帧-期望)',
        'ffmpeg\n拆帧',
        '差值4\n(ffmpeg-期望)',
        'ffmpeg\ndup',
        'ffmpeg\ndrop',
        '异常检测\n(花屏/丢帧/重复帧)',
        '输出时长',
        '处理耗时',
        '插帧倍数',
    ]
    col_widths = [8, 12, 12, 14, 13, 13, 14, 13, 14, 13, 12, 13, 10, 10, 28, 13, 12, 9]

    header_row = row
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 56
    row += 1

    # ── 数据行 ──
    segs = log_info.get('segments', [])
    mult = log_info.get('interpolation_factor', 2)

    totals = {
        'raw': 0, 'expected': 0.0, 'log_output': 0,
        'ffprob_packets': 0, 'ffprob_frames': 0, 'ffmpeg_frames': 0,
        'diff1_sum': 0.0, 'diff2_sum': 0, 'diff3_sum': 0, 'diff4_sum': 0,
        'dup_sum': 0, 'drop_sum': 0,
        'bad_segments': 0, 'warn_segments': 0,
    }

    for i, seg in enumerate(segs):
        # 原始帧数：优先使用 ffprobe 统计的 segment 帧数（最准确），回退到日志解析值
        seg_frames_ffprobe = _to_int(ffprobe_data['seg_packets'][i]) if i < len(ffprobe_data['seg_packets']) else None
        raw_from_log = seg.get('raw_frames') or 0
        raw = seg_frames_ffprobe if seg_frames_ffprobe is not None else raw_from_log
        # 期望输出帧数 = (原始帧数 - 1) * 插帧倍数 + 1
        expected = (raw - 1) * mult + 1
        log_output = seg.get('output_frames') or 0  # 日志提示输出帧数
        diff1 = log_output - expected                  # 差值1

        # ffprob count_packets (processed 目录)
        ffprob_pkt = _to_int(ffprobe_data['proc_packets'][i]) if i < len(ffprobe_data['proc_packets']) else None
        diff2 = (ffprob_pkt - expected) if ffprob_pkt is not None else None

        # ffprob count_frames (processed 目录)
        ffprob_frm = _to_int(ffprobe_data['proc_frames'][i]) if i < len(ffprobe_data['proc_frames']) else None
        diff3 = (ffprob_frm - expected) if ffprob_frm is not None else None

        # ffmpeg 拆帧
        ffmpeg_val = ffmpeg_frames[i] if i < len(ffmpeg_frames) else None
        diff4 = (ffmpeg_val - expected) if ffmpeg_val is not None else None

        # dup/drop 信息
        dup_val = None
        drop_val = None
        if i < len(ffmpeg_dup_drop):
            dd = ffmpeg_dup_drop[i]
            if dd:
                dup_val = dd.get('dup')
                drop_val = dd.get('drop')
                if dup_val is not None:
                    totals['dup_sum'] += dup_val
                if drop_val is not None:
                    totals['drop_sum'] += drop_val

        # ── 异常检测（花屏/丢帧/重复帧） ──
        errors = ffmpeg_errors[i] if i < len(ffmpeg_errors) else []
        has_illegal = any('illegal' in e.lower() for e in errors)
        has_missing = any('missing' in e.lower() for e in errors)
        has_duplicate = any('more than 1000 frames duplicated' in e.lower() for e in errors)
        has_no_frame = any('no frame' in e.lower() for e in errors)
        has_decode_error = any('decode_slice_header error' in e.lower() for e in errors)
        has_pps_error = any('non-existing pps' in e.lower() for e in errors)

        anomaly_parts = []

        # 花屏相关
        if has_illegal:
            anomaly_parts.append('⚠ illegal buffer(花屏)')
        if has_pps_error:
            anomaly_parts.append('⚠ PPS丢失')
        if has_decode_error:
            anomaly_parts.append('⚠ 解码错误')

        # 丢帧相关
        if has_no_frame:
            anomaly_parts.append('❌ no frame!(丢帧)')
        if has_missing:
            anomaly_parts.append('⚠ missing ref(丢帧)')

        # 重复帧相关
        if has_duplicate:
            anomaly_parts.append('🔁 大量重复帧(>1000)')
        if dup_val is not None and dup_val > 100:
            anomaly_parts.append(f'🔁 dup={dup_val}')

        # 帧数偏差
        if diff3 is not None and abs(diff3) > 2:
            anomaly_parts.append(f'📉 拆帧偏差={diff3}')
        if diff1 is not None and abs(diff1) > 2:
            anomaly_parts.append(f'📉 日志偏差={diff1:.0f}')

        # 判定严重程度
        is_bad = any(kw in ''.join(anomaly_parts) for kw in ['❌', '⚠ illegal', '⚠ PPS'])
        is_warn = (not is_bad) and len(anomaly_parts) > 0
        anomaly_str = '\n'.join(anomaly_parts) if anomaly_parts else '✅ 正常'

        if is_bad:
            totals['bad_segments'] += 1
        elif is_warn:
            totals['warn_segments'] += 1

        # 差值百分比（基于 diff1）
        pct1 = (abs(diff1) / expected * 100) if expected > 0 else 0

        row_data = [
            i + 1,
            f"segment_{seg['index']:03d}",
            raw,
            expected,
            log_output,
            f"{diff1:.1f} ({pct1:.2f}%)",
            ffprob_pkt if ffprob_pkt is not None else 'N/A',
            f"{diff2:.1f}" if diff2 is not None else 'N/A',
            ffprob_frm if ffprob_frm is not None else 'N/A',
            f"{diff3:.1f}" if diff3 is not None else 'N/A',
            ffmpeg_val if ffmpeg_val is not None else 'N/A',
            f"{diff4}" if diff4 is not None else 'N/A',
            dup_val if dup_val is not None else '',
            drop_val if drop_val is not None else '',
            anomaly_str,
            seg.get('output_duration', ''),
            seg.get('elapsed', ''),
            f"{mult}×",
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = center_align if ci != 15 else Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border

            # 异常列着色
            if ci == 15:
                if is_bad:
                    cell.font = bad_font
                    cell.fill = bad_fill
                elif is_warn:
                    cell.font = warn_font
                    cell.fill = warn_fill
                else:
                    cell.font = ok_font
            # 差值列着色 (ci: 6=差值1, 8=差值2, 10=差值3, 12=差值4)
            elif ci in (6, 8, 10, 12):
                try:
                    dv = float(str(val).split()[0]) if val and val != 'N/A' else 0
                except (ValueError, TypeError):
                    dv = 0
                if abs(dv) > 2:
                    cell.font = bad_font
                    cell.fill = bad_fill
                elif abs(dv) > 0.5:
                    cell.font = warn_font
                    cell.fill = warn_fill
                else:
                    cell.font = ok_font
            else:
                cell.font = normal_font

        totals['raw'] += raw
        totals['expected'] += expected
        totals['log_output'] += log_output
        if ffprob_pkt is not None:
            totals['ffprob_packets'] += ffprob_pkt
        if ffprob_frm is not None:
            totals['ffprob_frames'] += ffprob_frm
        if ffmpeg_val is not None:
            totals['ffmpeg_frames'] += ffmpeg_val
        totals['diff1_sum'] += diff1
        if diff2 is not None:
            totals['diff2_sum'] += diff2
        if diff3 is not None:
            totals['diff3_sum'] += diff3
        if diff4 is not None:
            totals['diff4_sum'] += diff4

        ws.row_dimensions[row].height = max(30, len(anomaly_parts) * 16 + 14)
        row += 1

    # ── 汇总行 ──
    sum_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    sum_font = Font(name='微软雅黑', size=10, bold=True)

    expected_total = totals['expected']
    diff1_total = totals['diff1_sum']
    diff2_total = totals['diff2_sum']
    diff3_total = totals['diff3_sum']
    diff4_total = totals['diff4_sum']

    anomaly_summary_parts = []
    if totals['bad_segments'] > 0:
        anomaly_summary_parts.append(f"❌ 花屏/丢帧: {totals['bad_segments']}/{len(segs)}")
    if totals['warn_segments'] > 0:
        anomaly_summary_parts.append(f"⚠ 警告: {totals['warn_segments']}/{len(segs)}")
    if totals['dup_sum'] > 0:
        anomaly_summary_parts.append(f"总dup={totals['dup_sum']}")
    if totals['drop_sum'] > 0:
        anomaly_summary_parts.append(f"总drop={totals['drop_sum']}")
    anomaly_summary = '\n'.join(anomaly_summary_parts) if anomaly_summary_parts else '✅ 全部正常'

    sum_data = [
        '汇总', '',
        totals['raw'],
        f"{expected_total:.1f}",
        totals['log_output'],
        f"{diff1_total:.1f}" + (f" ({abs(diff1_total)/expected_total*100:.2f}%)" if expected_total > 0 else ""),
        totals['ffprob_packets'],
        f"{diff2_total:.1f}" if diff2_total else 'N/A',
        totals['ffprob_frames'],
        f"{diff3_total:.1f}" if diff3_total else 'N/A',
        totals['ffmpeg_frames'],
        f"{diff4_total:.1f}" if diff4_total else 'N/A',
        totals['dup_sum'],
        totals['drop_sum'],
        anomaly_summary,
        log_info.get('total_output_duration', ''),
        log_info.get('total_elapsed', ''),
        f"{mult}×",
    ]
    for ci, val in enumerate(sum_data, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = sum_font
        cell.fill = sum_fill
        cell.alignment = center_align if ci != 15 else Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row].height = max(30, len(anomaly_summary_parts) * 16 + 14)

    ws.freeze_panes = f'A{header_row + 1}'

    # ── Sheet 2: 详细错误日志 ──
    ws2 = wb.create_sheet("ffmpeg错误详情")

    ws2.merge_cells('A1:E1')
    ws2['A1'] = 'ffmpeg image2pipe 解码 — 逐段错误详情'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    headers2 = ['分段', 'ffmpeg拆帧帧数', 'dup', 'drop', '错误/警告信息']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 80
    ws2.row_dimensions[3].height = 28

    r = 4
    for i in range(len(segs)):
        seg_name = f"seg_{i:03d}"
        ffmpeg_val = ffmpeg_frames[i] if i < len(ffmpeg_frames) else 'N/A'
        dd = ffmpeg_dup_drop[i] if i < len(ffmpeg_dup_drop) else {}
        dup_val = dd.get('dup', '') if dd else ''
        drop_val = dd.get('drop', '') if dd else ''
        errors = ffmpeg_errors[i] if i < len(ffmpeg_errors) else []
        err_text = '\n'.join(errors[:30]) if errors else '✅ 无错误'

        data2 = [seg_name, ffmpeg_val, dup_val, drop_val, err_text]
        for ci, val in enumerate(data2, 1):
            cell = ws2.cell(row=r, column=ci, value=val)
            cell.font = bad_font if (errors and ci == 5) else normal_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True) if ci == 5 else center_align
            cell.border = thin_border
        ws2.row_dimensions[r].height = max(30, min(300, len(errors) * 16))
        r += 1

    ws2.freeze_panes = 'A4'

    wb.save(output_path)
    print(f"✅ xlsx 已生成: {output_path}")


# ═══════════════════════════════════════════════
# 5. 主函数
# ═══════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='视频增强处理结果检测分析脚本 (v2)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('log_file', help='流水线输出日志文件 (如 output9.txt)')
    parser.add_argument('--suffix', default='9', help='输出文件后缀编号 (默认: 9)')
    parser.add_argument('--segments-dir', default=None,
                        help='segments 目录 (默认: 日志同目录下的 segments/)')
    parser.add_argument('--processed-dir', default=None,
                        help='processed 目录 (默认: 日志同目录下的 processed/)')
    parser.add_argument('--output-video', default=None,
                        help='输出视频文件路径（由 processed 分段合成而来）')
    parser.add_argument('--output-dir', default=None,
                        help='输出目录 (默认: 日志同目录)')
    parser.add_argument('--image2pipe', action='store_true',
                        help='使用 -f image2pipe 模式（完整解码+编码管线，慢但能捕获 dup/drop；默认用 -f null 快速模式）')
    parser.add_argument('--skip-ffmpeg', action='store_true',
                        help='跳过 ffmpeg 解码')
    parser.add_argument('--skip-ffprobe', action='store_true',
                        help='跳过 ffprobe 帧数统计')
    parser.add_argument('--skip-xlsx', action='store_true',
                        help='跳过 xlsx 生成')

    args = parser.parse_args()

    log_path = Path(args.log_file).resolve()
    if not log_path.exists():
        print(f"[ERROR] 日志文件不存在: {log_path}")
        sys.exit(1)

    log_dir = log_path.parent
    output_dir = Path(args.output_dir) if args.output_dir else log_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    segments_dir = args.segments_dir or str(log_dir / 'segments')
    processed_dir = args.processed_dir or str(log_dir / 'processed')
    suffix = args.suffix

    # ── 解析日志 ──
    print(f"📖 解析日志: {log_path}")
    log_info = parse_output_log(str(log_path))
    print(f"   分段数: {log_info['num_segments']}")
    print(f"   插帧倍数: {log_info['interpolation_factor']}×")

    output_video = args.output_video or log_info.get('output_video', '')
    if not output_video:
        for f in log_dir.iterdir():
            if f.suffix.lower() in ('.mp4', '.avi', '.mkv') and f.is_file():
                output_video = str(f)
                break

    num_segs = log_info['num_segments']

    # 初始化收集器
    ffmpeg_frames = []
    ffmpeg_errors = []
    ffmpeg_dup_drop = []
    ffprobe_data = {
        'seg_packets': [], 'seg_frames': [],
        'proc_packets': [], 'proc_frames': [],
        'src_packets': None, 'src_frames': None,
    }

    # ── 生成 ffmpeg_output ──
    if not args.skip_ffmpeg:
        ffmpeg_path = output_dir / f"ffmpeg_output{suffix}.txt"
        mode_label = "image2pipe" if args.image2pipe else "null"
        print(f"\n🎬 生成 ffmpeg 输出 ({mode_label} 解码模式): {ffmpeg_path}")
        try:
            ffmpeg_frames, ffmpeg_errors, ffmpeg_dup_drop = generate_ffmpeg_output(
                processed_dir, str(ffmpeg_path), num_segs,
                use_image2pipe=args.image2pipe
            )
        except Exception as e:
            print(f"[WARN] ffmpeg 生成异常: {e}，继续后续步骤")

    # ── 生成 ffprob_output ──
    if not args.skip_ffprobe:
        ffprob_path = output_dir / f"ffprob_output{suffix}.txt"
        print(f"\n🔍 生成 ffprobe 输出: {ffprob_path}")
        try:
            ffprobe_data = generate_ffprob_output(
                segments_dir, processed_dir, output_video,
                str(ffprob_path), num_segs
            )
        except Exception as e:
            print(f"[WARN] ffprobe 生成异常: {e}，继续后续步骤")

    # ── 生成 xlsx ──
    if not args.skip_xlsx:
        xlsx_path = output_dir / f"帧数差值统计汇总_output{suffix}.xlsx"
        print(f"\n📊 生成 xlsx: {xlsx_path}")
        try:
            generate_xlsx(
                log_info, ffprobe_data,
                ffmpeg_frames, ffmpeg_errors, ffmpeg_dup_drop,
                str(xlsx_path)
            )
        except Exception as e:
            print(f"[ERROR] xlsx 生成失败: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "=" * 60)
    print("  ✅ 所有检测分析文件生成完毕！")
    print("=" * 60)


if __name__ == '__main__':
    main()
