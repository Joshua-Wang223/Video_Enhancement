#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频增强全流程检测分析脚本（合并版）
=====================================
合并 analyze_video_ifrnet.py 与 analyze_video_realesrgan.py，
支持 IFRNet 插帧 + Real-ESRGAN 超分 任意顺序、任意步数的全流程日志。

核心能力（全部根据 log_file 自动解析，可用命令行参数覆盖）：
  1. 自动识别插帧/超分步骤顺序（依据 "阶段 X: Step N/M — xxx" 步骤头，
     无步骤头时回退到 "🎬 IFRNet 插帧处理" / "🎨 Real-ESRGAN 超分处理" 标记）
  2. 自动反推初始分段目录：取第一个步骤首个输出分段路径
     （如 .../ifrnet_source_xxx/processed/interpolated_segment_000.mp4）
     的同级 ../segments 目录
  3. 自动串联中间分段位置：前一步骤的分段输出目录即后一步骤的分段输入目录
  4. 自动定位最终输出视频（📤 输出 / 🔗 合并 N 个最终分段 → / 配置摘要 输出）

生成文件（suffix 默认从日志文件名自动提取，如 output28-9.txt → 28-9）：
  ffmpeg_output{suffix}.txt          — 各步骤输出分段逐段 ffmpeg 解码日志
  ffprob_output{suffix}.txt          — 初始分段 + 各步骤输出分段 + 最终视频 ffprobe 帧数统计
  帧数差值统计汇总_output{suffix}.xlsx — 全流程概览 + 每步骤帧数差值明细 + 错误详情 + 性能明细

用法:
    python analyze_video_full.py <output_log> [--suffix X] [--image2pipe] ...

示例:
    python analyze_video_full.py temp/output28-9.txt
    python analyze_video_full.py temp/output28-9.txt --image2pipe --output-dir temp/analysis
"""

import argparse
import re
import subprocess
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl 未安装，将不会生成 xlsx 文件。安装: pip install openpyxl")


STEP_TYPE_NAMES = {
    'ifrnet': 'IFRNet 插帧',
    'esrgan': 'Real-ESRGAN 超分',
    'unknown': '未知步骤',
}


# ═══════════════════════════════════════════════
# 1. 全流程日志解析
# ═══════════════════════════════════════════════

def _detect_step_type(title: str) -> str:
    t = title.lower()
    if 'ifrnet' in t or '插帧' in title:
        return 'ifrnet'
    if 'esrgan' in t or '超分' in title:
        return 'esrgan'
    return 'unknown'


def _parse_step_segments(step_text: str, step_type: str, default_mult: int) -> list:
    """解析单个步骤区块内的所有分段信息"""
    seg_head = re.compile(
        r'[🎬🎨]\s*片段\s*(\d+)/(\d+)\s*[:：]\s*(\S*segment_(\d+)\.mp4)'
    )
    frame_pat = re.compile(
        r'原始帧\s*[:：]\s*(\d+)\s*→\s*输出帧\s*[:：]\s*(\d+)(?:\s*\(×([\d.]+))?'
    )
    gpu_done_pat = re.compile(
        r'\[GPU\d+\]\s*完成\s*\|\s*原始帧=(\d+)\s*→\s*输出帧=(\d+)\s*\|\s*([\d.]+)\s*原始帧/s'
    )
    out_file_pat = re.compile(
        r'(?:✅\s*)?输出\s*[:：]\s*(.+?segment_\d+\.mp4)\s*\(([\d.]+\s*[KMGT]B)\)'
    )
    time_pat = re.compile(r'输出时长\s*([\d:.]+)\s*\|\s*耗时\s*([\d:.]+)')
    model_load_pat = re.compile(r'总耗时（含模型加载）\s*[:：]\s*([\d分秒]+)')
    perf_pat = re.compile(
        r'\[性能统计\]\s*总时间\s*[:：]\s*([\d.]+)秒\s*\|\s*平均\s*[:：]\s*([\d.]+)ms\s*\|\s*FPS\s*[:：]\s*([\d.]+)'
    )
    gfpgan_pat = re.compile(r'\[性能统计\]\s*GFPGAN\s*模式\s*[:：]\s*(\S+)')
    estimate_pat = re.compile(r'已用时\s*[:：]\s*([\d:.]+)\s*,\s*预计剩余\s*[:：]\s*([\d:.]+)')

    seg_starts = [
        (m.start(), int(m.group(1)), m.group(3), int(m.group(4)))
        for m in seg_head.finditer(step_text)
    ]

    segments = []
    for idx, (start_pos, seg_num, input_name, seg_id) in enumerate(seg_starts):
        end_pos = seg_starts[idx + 1][0] if idx + 1 < len(seg_starts) else len(step_text)
        seg_text = step_text[start_pos:end_pos]

        seg = {
            'index': seg_id,
            'number': seg_num,
            'input_name': input_name,
            'raw_frames': None,
            'output_frames': None,
            'expected_frames': None,
            'frame_diff': None,
            'mult': None,              # ifrnet 插帧倍数（段内 ×2.0）
            'speed': None,             # ifrnet 原始帧/s
            'output_file': '',
            'output_size': '',
            'output_duration': '',
            'elapsed': '',
            'model_load_time': '',
            'perf_total_sec': None,
            'perf_avg_ms': None,
            'perf_fps': None,
            'gfpgan_mode': 'unknown',
            'elapsed_estimate': '',
            'remaining_estimate': '',
        }

        fm = frame_pat.search(seg_text)
        if fm:
            seg['raw_frames'] = int(fm.group(1))
            seg['output_frames'] = int(fm.group(2))
            if fm.group(3):
                seg['mult'] = float(fm.group(3))

        gm = gpu_done_pat.search(seg_text)
        if gm:
            if seg['raw_frames'] is None:
                seg['raw_frames'] = int(gm.group(1))
                seg['output_frames'] = int(gm.group(2))
            seg['speed'] = float(gm.group(3))

        ofm = out_file_pat.search(seg_text)
        if ofm:
            seg['output_file'] = ofm.group(1).strip()
            seg['output_size'] = ofm.group(2).strip()

        tm = time_pat.search(seg_text)
        if tm:
            seg['output_duration'] = tm.group(1)
            seg['elapsed'] = tm.group(2)

        mlm = model_load_pat.search(seg_text)
        if mlm:
            seg['model_load_time'] = mlm.group(1)

        pm = perf_pat.search(seg_text)
        if pm:
            seg['perf_total_sec'] = float(pm.group(1))
            seg['perf_avg_ms'] = float(pm.group(2))
            seg['perf_fps'] = float(pm.group(3))

        gm2 = gfpgan_pat.search(seg_text)
        if gm2:
            seg['gfpgan_mode'] = gm2.group(1).strip()

        em = estimate_pat.search(seg_text)
        if em:
            seg['elapsed_estimate'] = em.group(1)
            seg['remaining_estimate'] = em.group(2)

        # 期望输出帧数
        if seg['raw_frames'] is not None:
            if step_type == 'ifrnet':
                mult = seg['mult'] or default_mult
                seg['expected_frames'] = int((seg['raw_frames'] - 1) * mult + 1)
            else:
                # 超分/未知：帧数守恒
                seg['expected_frames'] = seg['raw_frames']
            if seg['output_frames'] is not None:
                seg['frame_diff'] = seg['output_frames'] - seg['expected_frames']

        segments.append(seg)

    return segments


def parse_pipeline_log(log_path: str) -> dict:
    """解析全流程流水线日志，返回步骤列表及全局信息"""
    with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    info = {
        'source_video': '',
        'output_video': '',
        'mode': '',
        'interpolation_factor': 2,
        'upscale_factor': 2,
        'segment_duration': 300,
        'num_segments': 0,
        'total_output_size': '',
        'total_elapsed': '',
        'steps': [],
    }

    # ── 全局配置摘要 ──
    m = re.search(r'处理模式\s*[:：]\s*(\S+)', content)
    if m:
        info['mode'] = m.group(1).strip()
    m = re.search(r'插帧倍数\s*[:：]\s*(\d+)', content)
    if m:
        info['interpolation_factor'] = int(m.group(1))
    m = re.search(r'超分倍数\s*[:：]\s*(\d+)', content)
    if m:
        info['upscale_factor'] = int(m.group(1))
    m = re.search(r'分段时长\s*[:：]\s*(\d+)', content)
    if m:
        info['segment_duration'] = int(m.group(1))

    # 源视频（配置摘要中的首个 "输入: xxx.mp4"）
    m = re.search(r'^\s*输入\s*[:：]\s*(.+\.mp4)\s*$', content, re.MULTILINE)
    if m:
        info['source_video'] = m.group(1).strip()

    # 最终输出视频：优先 📤 输出/输出文件，其次 🔗 合并行，再次配置摘要 输出
    out_matches = re.findall(r'📤\s*输出(?:文件)?\s*[:：]\s*(.+\.mp4)', content)
    if out_matches:
        info['output_video'] = out_matches[-1].strip()
    else:
        m = re.search(r'🔗\s*合并\s*\d+\s*个最终分段\s*→\s*(.+\.mp4)', content)
        if m:
            info['output_video'] = m.group(1).strip()
        else:
            m = re.search(r'^\s*输出\s*[:：]\s*(.+\.mp4)\s*$', content, re.MULTILINE)
            if m:
                info['output_video'] = m.group(1).strip()

    m = re.search(r'📦\s*文件大小\s*[:：]\s*([\d.]+\s*[KMGT]B)', content)
    if m:
        info['total_output_size'] = m.group(1)
    m = re.search(r'总用时\s*[:：]\s*([\d:.]+)', content)
    if m:
        info['total_elapsed'] = m.group(1)

    # ── 步骤区块划分 ──
    # 方式1: "阶段 X: Step N/M — 标题"（全流程流水线日志）
    step_head = re.compile(
        r'阶段\s*(\d+)\s*[:：]\s*Step\s*(\d+)/(\d+)\s*[—–\-]\s*(.+)'
    )
    step_marks = [
        (m.start(), int(m.group(2)), int(m.group(3)), m.group(4).strip())
        for m in step_head.finditer(content)
    ]

    # 方式2: 独立步骤日志回退标记
    if not step_marks:
        for m in re.finditer(r'🎬\s*IFRNet 插帧处理', content):
            step_marks.append((m.start(), 1, 1, 'IFRNet 插帧'))
        for m in re.finditer(r'🎨\s*Real-ESRGAN 超分处理', content):
            step_marks.append((m.start(), 1, 1, 'Real-ESRGAN 超分'))
        step_marks.sort(key=lambda x: x[0])
        # 重新编号
        total = len(step_marks)
        step_marks = [
            (pos, i + 1, total, title)
            for i, (pos, _, _, title) in enumerate(step_marks)
        ]

    # 方式3: 无任何标记 —— 整个文件按片段头 emoji 推断单一步骤
    if not step_marks:
        title = 'IFRNet 插帧' if re.search(r'🎬\s*片段', content) else 'Real-ESRGAN 超分'
        step_marks = [(0, 1, 1, title)]

    for i, (pos, step_no, total_steps, title) in enumerate(step_marks):
        end_pos = step_marks[i + 1][0] if i + 1 < len(step_marks) else len(content)
        step_text = content[pos:end_pos]

        step_type = _detect_step_type(title)
        step = {
            'type': step_type,
            'name': STEP_TYPE_NAMES.get(step_type, title),
            'title': title,
            'step_no': step_no,
            'total_steps': total_steps,
            'num_segments': 0,
            'input_dir': '',
            'output_dir': '',
            'segments': [],
        }

        step['segments'] = _parse_step_segments(
            step_text, step_type, info['interpolation_factor']
        )
        if step['segments']:
            step['num_segments'] = max(s['number'] for s in step['segments'])
        else:
            m2 = re.search(r'✅\s*共\s*(\d+)\s*个片段', step_text)
            if m2:
                step['num_segments'] = int(m2.group(1))

        # 步骤输出目录：取该步骤第一个带输出路径的分段
        for seg in step['segments']:
            if seg.get('output_file'):
                step['output_dir'] = str(Path(seg['output_file']).parent)
                break

        info['steps'].append(step)

    # 全局分段数（取各步骤最大值）
    if info['steps']:
        info['num_segments'] = max(s['num_segments'] for s in info['steps'])
    else:
        m = re.search(r'✅\s*共\s*(\d+)\s*个片段', content)
        if m:
            info['num_segments'] = int(m.group(1))

    # ── 自动推导目录 ──
    log_dir = Path(log_path).resolve().parent
    _derive_step_dirs(info, log_dir)

    return info


def _derive_step_dirs(info: dict, log_dir: Path):
    """
    自动推导：
      - 初始分段目录（第一个步骤的输入）：由第一个步骤输出分段路径反推 ../segments
      - 中间分段位置：前一步输出目录 = 后一步输入目录
    """
    steps = info['steps']
    if not steps:
        return

    # 各步骤输出目录兜底
    for step in steps:
        if not step['output_dir']:
            step['output_dir'] = str(log_dir / 'processed')

    # 初始分段目录：由第一步输出路径反推
    # 例: .../ifrnet_source_xxx/processed/interpolated_segment_000.mp4
    #   → .../ifrnet_source_xxx/segments
    first = steps[0]
    candidates = []
    for seg in first['segments']:
        if seg.get('output_file'):
            out_p = Path(seg['output_file'])
            candidates.append(out_p.parent.parent / 'segments')
            break
    candidates.append(Path(first['output_dir']).parent / 'segments')
    candidates.append(log_dir / 'segments')
    candidates.append(log_dir.parent / 'segments')

    initial_dir = ''
    for cand in candidates:
        if cand.exists() and list(cand.glob('*segment_000.mp4')):
            initial_dir = str(cand)
            break
    if not initial_dir:
        for cand in candidates:
            if cand.exists():
                initial_dir = str(cand)
                break
    if not initial_dir:
        initial_dir = str(candidates[0]) if candidates else str(log_dir / 'segments')

    # 串联：步骤 k 的输入 = 步骤 k-1 的输出
    steps[0]['input_dir'] = initial_dir
    for k in range(1, len(steps)):
        steps[k]['input_dir'] = steps[k - 1]['output_dir']

    info['initial_segments_dir'] = initial_dir


# ═══════════════════════════════════════════════
# 2. ffmpeg 解码 + 错误检测
# ═══════════════════════════════════════════════

def _ffmpeg_decode(video_path: str, use_image2pipe: bool = False, timeout: int = 300) -> dict:
    """ffmpeg 解码（image2pipe 完整管线 或 null 轻量），流式捕获帧数与错误"""
    result = {'frame_count': None, 'errors': [], 'dup_count': None, 'drop_count': None}
    if use_image2pipe:
        cmd = ["ffmpeg", "-hide_banner", "-i", str(video_path),
               "-f", "image2pipe", "-vcodec", "png", "-"]
    else:
        cmd = ["ffmpeg", "-hide_banner", "-i", str(video_path), "-f", "null", "-"]
    try:
        proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL,
                                stderr=subprocess.PIPE, text=True, bufsize=1)
        try:
            for line in proc.stderr:
                stripped = line.strip()
                m = re.search(r'frame=\s*(\d+)', stripped)
                if m:
                    result['frame_count'] = int(m.group(1))
                m_dup = re.search(r'dup=(\d+)', stripped)
                m_drop = re.search(r'drop=(\d+)', stripped)
                if m_dup and m_drop and 'Lsize=' in stripped:
                    result['dup_count'] = int(m_dup.group(1))
                    result['drop_count'] = int(m_drop.group(1))
                lower = stripped.lower()
                if any(kw in lower for kw in [
                    'illegal', 'error', 'missing', 'corrupt',
                    'duplicated', 'duplicate', 'decoder',
                    'non-existing pps', 'no frame',
                    'more than 1000 frames duplicated',
                ]):
                    result['errors'].append(stripped)
            proc.wait(timeout=timeout)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg 解码超时")
    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")
    return result


def _find_segment_file(dir_path: Path, seg_idx: int) -> Path:
    """*segment_XXX.mp4 通配查找，适配 interpolated_/upscaled_ 等前缀"""
    matches = list(dir_path.glob(f"*segment_{seg_idx:03d}.mp4"))
    if matches:
        return matches[0]
    return dir_path / f"segment_{seg_idx:03d}.mp4"


def generate_ffmpeg_output(steps: list, output_path: str, use_image2pipe: bool = False) -> list:
    """
    对每个步骤的输出分段执行 ffmpeg 解码（中间分段 + 最终分段全覆盖）。
    返回 per-step 结果列表: [{'frames': [], 'errors': [], 'dup_drop': []}]
    """
    mode_label = "image2pipe" if use_image2pipe else "null"
    all_results = []

    with open(output_path, 'w', encoding='utf-8') as out:
        out.write("=" * 70 + "\n")
        out.write(" ffmpeg 逐段解码日志 — 全流程（各步骤输出分段）\n")
        out.write(f" 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        out.write(f" 解码模式: {mode_label}\n")
        out.write("=" * 70 + "\n\n")

        for step in steps:
            step_dir = Path(step['output_dir'])
            out.write(f"\n── Step {step['step_no']}/{step['total_steps']}: {step['name']} "
                      f"输出分段 ({step_dir}) ──\n\n")

            result = {'frames': [], 'errors': [], 'dup_drop': []}
            for i in range(step['num_segments']):
                seg_file = _find_segment_file(step_dir, i)
                out.write(f"$ ffmpeg -hide_banner -i {seg_file.name} "
                          f"-f {mode_label}{' -vcodec png' if use_image2pipe else ''} -\n")

                if not seg_file.exists():
                    print(f"  [WARN] 文件不存在: {seg_file}")
                    result['frames'].append(None)
                    result['errors'].append([f"[FILE NOT FOUND] {seg_file}"])
                    result['dup_drop'].append(None)
                    out.write("[FILE NOT FOUND]\n\n")
                    continue

                r = _ffmpeg_decode(str(seg_file), use_image2pipe=use_image2pipe)
                result['frames'].append(r['frame_count'])
                result['errors'].append(r['errors'])
                result['dup_drop'].append({'dup': r['dup_count'], 'drop': r['drop_count']})

                print(f"  [ffmpeg] step{step['step_no']} {seg_file.name}: "
                      f"frame={r['frame_count']}, dup={r['dup_count']}, "
                      f"drop={r['drop_count']}, errors={len(r['errors'])}")

                if r['frame_count'] is not None:
                    out.write(f"frame={r['frame_count']}\n")
                for err in r['errors']:
                    out.write(err + '\n')
                out.write("\n")

            all_results.append(result)

    print(f"✅ ffmpeg_output 已生成: {output_path}")
    return all_results


# ═══════════════════════════════════════════════
# 3. ffprobe 帧数统计
# ═══════════════════════════════════════════════

def _ffprobe_count(video_path: str, count_mode: str = 'packets', timeout: int = 60) -> str:
    if count_mode == 'packets':
        cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0",
               "-count_packets", "-show_entries", "stream=nb_read_packets",
               "-of", "csv=p=0", video_path]
    else:
        cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0",
               "-count_frames", "-show_entries", "stream=nb_read_frames",
               "-of", "csv=p=0", video_path]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        for text in (result.stdout.strip(), result.stderr.strip()):
            m = re.search(r'(\d+)', text)
            if m:
                return m.group(1)
        return result.stdout.strip() if result.stdout.strip() else "[ERROR]"
    except subprocess.TimeoutExpired:
        return "[TIMEOUT]"
    except FileNotFoundError:
        return "[SKIP: ffprobe not found]"


def _probe_dir(dir_path: Path, num_segments: int, label: str, out,
               packets_timeout=30, frames_timeout=300) -> dict:
    """对一个目录的所有分段执行 packets/frames 统计"""
    data = {'packets': [], 'frames': []}
    out.write(f"\n── {label}: nb_read_packets ──\n")
    for i in range(num_segments):
        seg_file = _find_segment_file(dir_path, i)
        out.write(f"ffprobe ... -count_packets {seg_file.name}\n")
        if seg_file.exists():
            count = _ffprobe_count(str(seg_file), 'packets', timeout=packets_timeout)
            print(f"  [{label}/packets] {seg_file.name}: {count}")
            data['packets'].append(count)
            out.write(count + '\n')
        else:
            data['packets'].append("[NOT FOUND]")
            out.write("[FILE NOT FOUND]\n")

    out.write(f"\n── {label}: nb_read_frames ──\n")
    for i in range(num_segments):
        seg_file = _find_segment_file(dir_path, i)
        out.write(f"ffprobe ... -count_frames {seg_file.name}\n")
        if seg_file.exists():
            count = _ffprobe_count(str(seg_file), 'frames', timeout=frames_timeout)
            print(f"  [{label}/frames]  {seg_file.name}: {count}")
            data['frames'].append(count)
            out.write(count + '\n')
        else:
            data['frames'].append("[NOT FOUND]")
            out.write("[FILE NOT FOUND]\n")
    return data


def generate_ffprob_output(info: dict, output_video: str, output_path: str) -> dict:
    """
    对 初始分段目录 + 各步骤输出目录 + 最终输出视频 执行 ffprobe 统计。
    步骤 k 的输入 = 步骤 k-1 的输出，各目录只统计一次后按步骤串联。
    返回: {
      'initial': {'packets': [], 'frames': []},
      'steps':   [{'packets': [], 'frames': []}, ...],
      'final':   {'packets': str|None, 'frames': str|None},
    }
    """
    steps = info['steps']
    data = {'initial': {'packets': [], 'frames': []}, 'steps': [],
            'final': {'packets': None, 'frames': None}}

    with open(output_path, 'w', encoding='utf-8') as out:
        out.write("=" * 70 + "\n")
        out.write(" ffprobe 帧数统计 — 全流程\n")
        out.write(f" 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        out.write(f" 初始分段目录: {info.get('initial_segments_dir', '')}\n")
        for step in steps:
            out.write(f" Step{step['step_no']} ({step['name']}): "
                      f"输入={step['input_dir']}  输出={step['output_dir']}\n")
        out.write(f" 最终输出视频: {output_video}\n")
        out.write("=" * 70 + "\n")

        probed = {}  # dir -> data，避免重复统计同一目录

        def probe(dir_str, nsegs, label):
            key = str(Path(dir_str))
            if key in probed:
                out.write(f"\n（{label} 与前面已统计目录相同，复用结果）\n")
                return probed[key]
            d = _probe_dir(Path(dir_str), nsegs, label, out)
            probed[key] = d
            return d

        # 初始分段目录
        data['initial'] = probe(info.get('initial_segments_dir', ''),
                                info['num_segments'], '初始分段 segments')

        # 各步骤输出目录
        for step in steps:
            label = f"Step{step['step_no']} {step['name']} 输出分段"
            data['steps'].append(probe(step['output_dir'], step['num_segments'], label))

        # 最终输出视频
        if output_video and Path(output_video).exists():
            out_name = Path(output_video).name
            out.write(f"\n── 最终输出视频: {out_name} ──\n")
            count = _ffprobe_count(str(output_video), 'packets', timeout=60)
            print(f"  [final/packets] {out_name}: {count}")
            data['final']['packets'] = count
            out.write(f"nb_read_packets: {count}\n")
            count = _ffprobe_count(str(output_video), 'frames', timeout=1800)
            print(f"  [final/frames]  {out_name}: {count}")
            data['final']['frames'] = count
            out.write(f"nb_read_frames: {count}\n")
        else:
            out.write(f"\n[WARN] 最终输出视频不存在: {output_video}\n")

        out.write("\n" + "=" * 70 + "\n")

    print(f"✅ ffprob_output 已生成: {output_path}")
    return data


# ═══════════════════════════════════════════════
# 4. xlsx 生成
# ═══════════════════════════════════════════════

def _to_int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


# 通用样式
def _styles():
    return {
        'header_font': Font(name='微软雅黑', size=10, bold=True, color='FFFFFF'),
        'header_fill': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),
        'title_font': Font(name='微软雅黑', size=14, bold=True, color='1F3864'),
        'normal_font': Font(name='微软雅黑', size=10),
        'warn_font': Font(name='微软雅黑', size=10, color='CC6600'),
        'bad_font': Font(name='微软雅黑', size=10, color='FF0000', bold=True),
        'ok_font': Font(name='微软雅黑', size=10, color='008000'),
        'thin_border': Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin')),
        'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'warn_fill': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'bad_fill': PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid'),
        'sum_fill': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
        'sum_font': Font(name='微软雅黑', size=10, bold=True),
    }


def _analyze_step_rows(step, in_packets, out_packets, out_frames,
                       ffmpeg_res, mult_default):
    """计算单个步骤的逐段行数据与汇总（供 xlsx 使用）"""
    segs = step['segments']
    rows = []
    totals = {
        'raw': 0, 'expected': 0, 'log_output': 0,
        'ffprob_packets': 0, 'ffprob_frames': 0, 'ffmpeg_frames': 0,
        'diff1_sum': 0, 'diff2_sum': 0, 'diff3_sum': 0, 'diff4_sum': 0,
        'dup_sum': 0, 'drop_sum': 0,
        'bad_segments': 0, 'warn_segments': 0,
        'perf_time_sum': 0.0, 'perf_fps_list': [],
    }

    ff_frames = ffmpeg_res['frames'] if ffmpeg_res else []
    ff_errors = ffmpeg_res['errors'] if ffmpeg_res else []
    ff_dd = ffmpeg_res['dup_drop'] if ffmpeg_res else []

    for i, seg in enumerate(segs):
        # 输入帧数：优先本步骤输入分段的 ffprobe 包数，回退日志值
        in_probe = _to_int(in_packets[i]) if i < len(in_packets) else None
        raw = in_probe if in_probe is not None else (seg.get('raw_frames') or 0)

        if step['type'] == 'ifrnet':
            mult = seg.get('mult') or mult_default
            expected = int((raw - 1) * mult + 1)
        else:
            mult = None
            expected = raw

        log_output = seg.get('output_frames') or 0
        diff1 = log_output - expected

        pkt = _to_int(out_packets[i]) if i < len(out_packets) else None
        diff2 = (pkt - expected) if pkt is not None else None
        frm = _to_int(out_frames[i]) if i < len(out_frames) else None
        diff3 = (frm - expected) if frm is not None else None
        ffv = ff_frames[i] if i < len(ff_frames) else None
        diff4 = (ffv - expected) if ffv is not None else None

        dup_val = drop_val = None
        if i < len(ff_dd) and ff_dd[i]:
            dup_val = ff_dd[i].get('dup')
            drop_val = ff_dd[i].get('drop')
            if dup_val is not None:
                totals['dup_sum'] += dup_val
            if drop_val is not None:
                totals['drop_sum'] += drop_val

        # 异常检测
        errors = ff_errors[i] if i < len(ff_errors) else []
        el = ' '.join(e.lower() for e in errors)
        anomaly_parts = []
        if 'illegal' in el:
            anomaly_parts.append('⚠ illegal buffer(花屏)')
        if 'non-existing pps' in el:
            anomaly_parts.append('⚠ PPS丢失')
        if 'decode_slice_header error' in el:
            anomaly_parts.append('⚠ 解码错误')
        if 'no frame' in el:
            anomaly_parts.append('❌ no frame!(丢帧)')
        if 'missing' in el:
            anomaly_parts.append('⚠ missing ref(丢帧)')
        if 'more than 1000 frames duplicated' in el:
            anomaly_parts.append('🔁 大量重复帧(>1000)')
        if dup_val is not None and dup_val > 100:
            anomaly_parts.append(f'🔁 dup={dup_val}')
        if abs(diff1) > (0 if step['type'] != 'ifrnet' else 2):
            anomaly_parts.append(f'📉 日志偏差={diff1}')
        for label, d in (('包偏差', diff2), ('帧偏差', diff3), ('拆帧偏差', diff4)):
            if d is not None and abs(d) > 2:
                anomaly_parts.append(f'📉 {label}={d}')

        joined = ''.join(anomaly_parts)
        is_bad = any(kw in joined for kw in ['❌', '⚠ illegal', '⚠ PPS'])
        is_warn = (not is_bad) and len(anomaly_parts) > 0
        anomaly_str = '\n'.join(anomaly_parts) if anomaly_parts else '✅ 正常'
        if is_bad:
            totals['bad_segments'] += 1
        elif is_warn:
            totals['warn_segments'] += 1

        if seg.get('perf_total_sec') is not None:
            totals['perf_time_sum'] += seg['perf_total_sec']
        if seg.get('perf_fps') is not None:
            totals['perf_fps_list'].append(seg['perf_fps'])

        rows.append({
            'seq': i + 1,
            'seg': seg,
            'raw': raw, 'expected': expected, 'log_output': log_output,
            'diff1': diff1, 'pkt': pkt, 'diff2': diff2,
            'frm': frm, 'diff3': diff3, 'ffv': ffv, 'diff4': diff4,
            'dup': dup_val, 'drop': drop_val,
            'anomaly': anomaly_str, 'is_bad': is_bad, 'is_warn': is_warn,
            'anomaly_count': len(anomaly_parts),
            'mult': mult,
        })

        totals['raw'] += raw
        totals['expected'] += expected
        totals['log_output'] += log_output
        if pkt is not None:
            totals['ffprob_packets'] += pkt
        if frm is not None:
            totals['ffprob_frames'] += frm
        if ffv is not None:
            totals['ffmpeg_frames'] += ffv
        totals['diff1_sum'] += diff1
        if diff2 is not None:
            totals['diff2_sum'] += diff2
        if diff3 is not None:
            totals['diff3_sum'] += diff3
        if diff4 is not None:
            totals['diff4_sum'] += diff4

    return rows, totals


def generate_xlsx(info, ffprobe_data, ffmpeg_results, output_video, output_path):
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl 未安装，跳过 xlsx 生成")
        return

    S = _styles()
    steps = info['steps']
    wb = openpyxl.Workbook()

    # ════ Sheet 1: 全流程概览 ════
    ws0 = wb.active
    ws0.title = "全流程概览"
    ws0.merge_cells('A1:H1')
    ws0['A1'] = '视频增强全流程 — 检测分析概览'
    ws0['A1'].font = S['title_font']
    ws0['A1'].alignment = S['center']
    ws0.row_dimensions[1].height = 30
    ws0.merge_cells('A2:H2')
    ws0['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws0['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws0['A2'].alignment = S['center']

    row = 4
    step_chain = ' → '.join(f"Step{s['step_no']}({s['name']})" for s in steps)
    overview_items = [
        ('源视频', info.get('source_video', 'N/A')),
        ('最终输出视频', output_video or info.get('output_video', 'N/A')),
        ('处理模式', info.get('mode', 'N/A')),
        ('步骤链', step_chain),
        ('插帧倍数', f"{info.get('interpolation_factor', 2)}×"),
        ('超分倍数', f"{info.get('upscale_factor', 2)}×"),
        ('分段时长', f"{info.get('segment_duration', 300)}秒"),
        ('分段数', str(info.get('num_segments', 0))),
        ('初始分段目录', info.get('initial_segments_dir', 'N/A')),
        ('输出大小', info.get('total_output_size', 'N/A')),
        ('总耗时', info.get('total_elapsed', 'N/A')),
        ('最终视频 ffprobe 包数', str(ffprobe_data['final'].get('packets') or 'N/A')),
        ('最终视频 ffprobe 帧数', str(ffprobe_data['final'].get('frames') or 'N/A')),
    ]
    for key, val in overview_items:
        ws0.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws0.cell(row=row, column=1, value=key).font = Font(name='微软雅黑', size=10, bold=True)
        ws0.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws0.cell(row=row, column=3, value=val).font = S['normal_font']
        row += 1

    # 步骤汇总表
    row += 1
    step_headers = ['步骤', '类型', '输入目录(=上一步输出)', '输出目录', '分段数',
                    '输入总帧数(ffprobe)', '日志输出总帧数', 'ffprobe输出总帧数', '异常段数']
    step_widths = [8, 18, 42, 42, 8, 18, 16, 18, 12]
    for ci, (h, w) in enumerate(zip(step_headers, step_widths), 1):
        cell = ws0.cell(row=row, column=ci, value=h)
        cell.font = S['header_font']
        cell.fill = S['header_fill']
        cell.alignment = S['center']
        cell.border = S['thin_border']
        ws0.column_dimensions[get_column_letter(ci)].width = w
    ws0.row_dimensions[row].height = 30
    row += 1

    step_rows_data = []  # 缓存供后续 sheet 使用
    for si, step in enumerate(steps):
        # 输入 packets：步骤0 用初始目录，其余用上一步输出
        if si == 0:
            in_packets = ffprobe_data['initial']['packets']
        else:
            in_packets = ffprobe_data['steps'][si - 1]['packets']
        out_packets = ffprobe_data['steps'][si]['packets']
        out_frames = ffprobe_data['steps'][si]['frames']
        ffmpeg_res = ffmpeg_results[si] if si < len(ffmpeg_results) else None

        rows, totals = _analyze_step_rows(
            step, in_packets, out_packets, out_frames,
            ffmpeg_res, info['interpolation_factor'],
        )
        step_rows_data.append((rows, totals))

        n_anom = totals['bad_segments'] + totals['warn_segments']
        vals = [
            f"Step {step['step_no']}",
            step['name'],
            step['input_dir'],
            step['output_dir'],
            step['num_segments'],
            totals['raw'],
            totals['log_output'],
            totals['ffprob_packets'] if totals['ffprob_packets'] else 'N/A',
            f"❌{totals['bad_segments']} ⚠{totals['warn_segments']}" if n_anom else '✅ 0',
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws0.cell(row=row, column=ci, value=val)
            cell.font = S['normal_font']
            cell.alignment = S['center'] if ci not in (3, 4) else S['left']
            cell.border = S['thin_border']
        if totals['bad_segments']:
            ws0.cell(row=row, column=9).font = S['bad_font']
            ws0.cell(row=row, column=9).fill = S['bad_fill']
        elif totals['warn_segments']:
            ws0.cell(row=row, column=9).font = S['warn_font']
            ws0.cell(row=row, column=9).fill = S['warn_fill']
        row += 1

    # ════ 每个步骤一个明细 Sheet ════
    for si, step in enumerate(steps):
        rows, totals = step_rows_data[si]
        sheet_name = f"Step{step['step_no']}-{step['type']}"
        ws = wb.create_sheet(sheet_name[:31])

        is_ifrnet = step['type'] == 'ifrnet'
        is_esrgan = step['type'] == 'esrgan'

        headers = [
            '分段序号', '分段编号', '输入帧数\n(本步输入)',
            '期望输出帧数\n' + ('(输入-1)×倍数+1' if is_ifrnet else '(=输入帧)'),
            '日志提示\n输出帧数', '差值1\n(日志-期望)',
            'ffprobe\n-count_packets', '差值2\n(包-期望)',
            'ffprobe\n-count_frames', '差值3\n(帧-期望)',
            'ffmpeg\n拆帧', '差值4\n(ffmpeg-期望)',
            'ffmpeg\ndup', 'ffmpeg\ndrop',
            '异常检测\n(花屏/丢帧/重复帧)',
            '输出时长', '处理耗时',
        ]
        widths = [8, 12, 12, 14, 12, 12, 14, 12, 14, 12, 12, 12, 9, 9, 28, 13, 12]
        if is_ifrnet:
            headers += ['插帧倍数', '速度\n(原始帧/s)']
            widths += [9, 12]
        if is_esrgan:
            headers += ['性能统计\n(总秒/平均ms/FPS)', 'GFPGAN\n模式', '输出大小', '模型加载\n耗时']
            widths += [22, 10, 12, 12]
        headers += ['输出文件名']
        widths += [36]
        num_cols = len(headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws['A1'] = f"Step {step['step_no']}/{step['total_steps']} — {step['name']} 帧数差值统计"
        ws['A1'].font = S['title_font']
        ws['A1'].alignment = S['center']
        ws.row_dimensions[1].height = 28

        r = 3
        meta_items = [
            ('输入目录', step['input_dir']),
            ('输出目录', step['output_dir']),
        ]
        for key, val in meta_items:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.cell(row=r, column=1, value=key).font = Font(name='微软雅黑', size=10, bold=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
            ws.cell(row=r, column=3, value=val).font = S['normal_font']
            r += 1
        r += 1

        header_row = r
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = S['header_font']
            cell.fill = S['header_fill']
            cell.alignment = S['center']
            cell.border = S['thin_border']
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[r].height = 56
        r += 1

        for rd in rows:
            seg = rd['seg']
            perf_str = ''
            if seg.get('perf_total_sec') is not None:
                perf_str = (f"{seg['perf_total_sec']:.1f}s / "
                            f"{seg.get('perf_avg_ms', 0):.1f}ms / "
                            f"FPS {seg.get('perf_fps', 0):.2f}")
            out_name = Path(seg['output_file']).name if seg.get('output_file') else ''

            row_data = [
                rd['seq'],
                f"segment_{seg['index']:03d}",
                rd['raw'], rd['expected'], rd['log_output'],
                f"{rd['diff1']}",
                rd['pkt'] if rd['pkt'] is not None else 'N/A',
                f"{rd['diff2']}" if rd['diff2'] is not None else 'N/A',
                rd['frm'] if rd['frm'] is not None else 'N/A',
                f"{rd['diff3']}" if rd['diff3'] is not None else 'N/A',
                rd['ffv'] if rd['ffv'] is not None else 'N/A',
                f"{rd['diff4']}" if rd['diff4'] is not None else 'N/A',
                rd['dup'] if rd['dup'] is not None else '',
                rd['drop'] if rd['drop'] is not None else '',
                rd['anomaly'],
                seg.get('output_duration', ''),
                seg.get('elapsed', ''),
            ]
            if is_ifrnet:
                row_data += [
                    f"{rd['mult']:g}×" if rd['mult'] else f"{info['interpolation_factor']}×",
                    f"{seg['speed']:.1f}" if seg.get('speed') else '',
                ]
            if is_esrgan:
                row_data += [
                    perf_str,
                    seg.get('gfpgan_mode', 'N/A'),
                    seg.get('output_size', ''),
                    seg.get('model_load_time', ''),
                ]
            row_data += [out_name]

            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.alignment = S['left'] if ci == 15 else S['center']
                cell.border = S['thin_border']
                if ci == 15:
                    if rd['is_bad']:
                        cell.font = S['bad_font']
                        cell.fill = S['bad_fill']
                    elif rd['is_warn']:
                        cell.font = S['warn_font']
                        cell.fill = S['warn_fill']
                    else:
                        cell.font = S['ok_font']
                elif ci in (6, 8, 10, 12):
                    try:
                        dv = float(str(val).split()[0]) if val not in ('', 'N/A', None) else 0
                    except (ValueError, TypeError):
                        dv = 0
                    if abs(dv) > 2:
                        cell.font = S['bad_font']
                        cell.fill = S['bad_fill']
                    elif abs(dv) > 0.5:
                        cell.font = S['warn_font']
                        cell.fill = S['warn_fill']
                    else:
                        cell.font = S['ok_font']
                else:
                    cell.font = S['normal_font']
            ws.row_dimensions[r].height = max(30, rd['anomaly_count'] * 16 + 14)
            r += 1

        # 汇总行
        anomaly_summary_parts = []
        if totals['bad_segments']:
            anomaly_summary_parts.append(f"❌ 花屏/丢帧: {totals['bad_segments']}/{len(rows)}")
        if totals['warn_segments']:
            anomaly_summary_parts.append(f"⚠ 警告: {totals['warn_segments']}/{len(rows)}")
        if totals['dup_sum']:
            anomaly_summary_parts.append(f"总dup={totals['dup_sum']}")
        if totals['drop_sum']:
            anomaly_summary_parts.append(f"总drop={totals['drop_sum']}")
        anomaly_summary = '\n'.join(anomaly_summary_parts) if anomaly_summary_parts else '✅ 全部正常'

        sum_data = [
            '汇总', '',
            totals['raw'], totals['expected'], totals['log_output'],
            f"{totals['diff1_sum']}",
            totals['ffprob_packets'] or 'N/A', f"{totals['diff2_sum']}",
            totals['ffprob_frames'] or 'N/A', f"{totals['diff3_sum']}",
            totals['ffmpeg_frames'] or 'N/A', f"{totals['diff4_sum']}",
            totals['dup_sum'], totals['drop_sum'],
            anomaly_summary, '', '',
        ]
        if is_ifrnet:
            sum_data += ['', '']
        if is_esrgan:
            avg_fps = (sum(totals['perf_fps_list']) / len(totals['perf_fps_list'])) if totals['perf_fps_list'] else 0
            sum_data += [
                f"总{totals['perf_time_sum']:.1f}s | FPS均值{avg_fps:.2f}" if totals['perf_time_sum'] else '',
                '', '', '',
            ]
        sum_data += ['']
        for ci, val in enumerate(sum_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = S['sum_font']
            cell.fill = S['sum_fill']
            cell.alignment = S['left'] if ci == 15 else S['center']
            cell.border = S['thin_border']
        ws.row_dimensions[r].height = max(30, len(anomaly_summary_parts) * 16 + 14)
        ws.freeze_panes = f'A{header_row + 1}'

    # ════ ffmpeg 错误详情 Sheet ════
    ws2 = wb.create_sheet("ffmpeg错误详情")
    ws2.merge_cells('A1:F1')
    ws2['A1'] = 'ffmpeg 解码 — 各步骤逐段错误详情'
    ws2['A1'].font = S['title_font']
    ws2['A1'].alignment = S['center']
    ws2.row_dimensions[1].height = 28
    headers2 = ['步骤', '分段', 'ffmpeg拆帧帧数', 'dup', 'drop', '错误/警告信息']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = S['header_font']
        cell.fill = S['header_fill']
        cell.alignment = S['center']
        cell.border = S['thin_border']
    for col, w in zip('ABCDEF', [10, 14, 16, 10, 10, 80]):
        ws2.column_dimensions[col].width = w
    r = 4
    for si, step in enumerate(steps):
        ff = ffmpeg_results[si] if si < len(ffmpeg_results) else None
        if not ff:
            continue
        for i in range(step['num_segments']):
            ffv = ff['frames'][i] if i < len(ff['frames']) else 'N/A'
            dd = ff['dup_drop'][i] if i < len(ff['dup_drop']) else None
            errors = ff['errors'][i] if i < len(ff['errors']) else []
            err_text = '\n'.join(errors[:30]) if errors else '✅ 无错误'
            vals = [f"Step{step['step_no']}", f"segment_{i:03d}", ffv,
                    dd.get('dup', '') if dd else '', dd.get('drop', '') if dd else '',
                    err_text]
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=ci, value=val)
                cell.font = S['bad_font'] if (errors and ci == 6) else S['normal_font']
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True) if ci == 6 else S['center']
                cell.border = S['thin_border']
            ws2.row_dimensions[r].height = max(30, min(300, len(errors) * 16))
            r += 1
    ws2.freeze_panes = 'A4'

    # ════ 性能统计明细 Sheet ════
    ws3 = wb.create_sheet("性能统计明细")
    ws3.merge_cells('A1:I1')
    ws3['A1'] = '全流程性能统计明细'
    ws3['A1'].font = S['title_font']
    ws3['A1'].alignment = S['center']
    ws3.row_dimensions[1].height = 28
    headers3 = ['步骤', '分段', '速度(原始帧/s)', '总时间(秒)', '平均(ms)', 'FPS',
                '输出时长', '处理耗时', '模型加载耗时']
    widths3 = [10, 14, 14, 12, 12, 10, 12, 12, 14]
    for ci, (h, w) in enumerate(zip(headers3, widths3), 1):
        cell = ws3.cell(row=3, column=ci, value=h)
        cell.font = S['header_font']
        cell.fill = S['header_fill']
        cell.alignment = S['center']
        cell.border = S['thin_border']
        ws3.column_dimensions[get_column_letter(ci)].width = w
    r = 4
    for si, step in enumerate(steps):
        rows, totals = step_rows_data[si]
        for rd in rows:
            seg = rd['seg']
            vals = [
                f"Step{step['step_no']}", f"segment_{seg['index']:03d}",
                f"{seg['speed']:.1f}" if seg.get('speed') else '',
                seg.get('perf_total_sec', '') or '',
                seg.get('perf_avg_ms', '') or '',
                seg.get('perf_fps', '') or '',
                seg.get('output_duration', ''),
                seg.get('elapsed', ''),
                seg.get('model_load_time', ''),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=ci, value=val)
                cell.font = S['normal_font']
                cell.alignment = S['center']
                cell.border = S['thin_border']
            r += 1
        avg_fps = (sum(totals['perf_fps_list']) / len(totals['perf_fps_list'])) if totals['perf_fps_list'] else 0
        sum3 = [f"Step{step['step_no']}汇总", '', '',
                f"{totals['perf_time_sum']:.1f}" if totals['perf_time_sum'] else '',
                '', f"{avg_fps:.2f}" if avg_fps else '', '', '', '']
        for ci, val in enumerate(sum3, 1):
            cell = ws3.cell(row=r, column=ci, value=val)
            cell.font = S['sum_font']
            cell.fill = S['sum_fill']
            cell.alignment = S['center']
            cell.border = S['thin_border']
        r += 1
    ws3.freeze_panes = 'A4'

    wb.save(output_path)
    print(f"✅ xlsx 已生成: {output_path}")


# ═══════════════════════════════════════════════
# 5. 主函数
# ═══════════════════════════════════════════════

def _auto_suffix(log_name: str) -> str:
    """output28-9.txt → 28-9；output9.txt → 9；否则用文件 stem"""
    m = re.search(r'output(.+)\.txt$', log_name, re.IGNORECASE)
    if m:
        return m.group(1)
    return Path(log_name).stem


def main():
    parser = argparse.ArgumentParser(
        description='视频增强全流程检测分析脚本（IFRNet + Real-ESRGAN 合并版）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('log_file', help='全流程流水线输出日志文件 (如 output28-9.txt)')
    parser.add_argument('--suffix', default=None,
                        help='输出文件后缀 (默认: 从日志文件名自动提取)')
    parser.add_argument('--segments-dir', default=None,
                        help='初始分段目录 (默认: 由第一步输出分段路径反推 ../segments)')
    parser.add_argument('--output-video', default=None,
                        help='最终输出视频路径 (默认: 从日志 📤 输出/合并行解析)')
    parser.add_argument('--output-dir', default=None,
                        help='分析文件输出目录 (默认: 日志同目录)')
    parser.add_argument('--image2pipe', action='store_true',
                        help='ffmpeg 使用 -f image2pipe 完整解码模式（慢，可捕获 dup/drop）')
    parser.add_argument('--skip-ffmpeg', action='store_true', help='跳过 ffmpeg 解码')
    parser.add_argument('--skip-ffprobe', action='store_true', help='跳过 ffprobe 帧数统计')
    parser.add_argument('--skip-xlsx', action='store_true', help='跳过 xlsx 生成')

    args = parser.parse_args()

    log_path = Path(args.log_file).resolve()
    if not log_path.exists():
        print(f"[ERROR] 日志文件不存在: {log_path}")
        sys.exit(1)

    log_dir = log_path.parent
    output_dir = Path(args.output_dir) if args.output_dir else log_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = args.suffix if args.suffix is not None else _auto_suffix(log_path.name)

    # ── 解析日志（自动识别步骤顺序 + 推导目录） ──
    print(f"📖 解析日志: {log_path}")
    info = parse_pipeline_log(str(log_path))

    print(f"   处理模式: {info['mode'] or 'N/A'}")
    print(f"   步骤数: {len(info['steps'])}")
    for step in info['steps']:
        print(f"   Step {step['step_no']}/{step['total_steps']}: {step['name']} "
              f"({step['num_segments']} 段)")
        print(f"      输入目录: {step['input_dir']}")
        print(f"      输出目录: {step['output_dir']}")
    print(f"   初始分段目录: {info.get('initial_segments_dir', 'N/A')}")

    # 命令行覆盖初始分段目录
    if args.segments_dir:
        info['initial_segments_dir'] = args.segments_dir
        if info['steps']:
            info['steps'][0]['input_dir'] = args.segments_dir

    # 最终输出视频
    output_video = args.output_video or info.get('output_video', '')
    if not output_video or not Path(output_video).exists():
        for cand in [log_dir, log_dir.parent]:
            if not cand.exists():
                continue
            for f in cand.iterdir():
                if (f.suffix.lower() in ('.mp4', '.avi', '.mkv') and f.is_file()
                        and 'segment' not in f.name.lower()):
                    output_video = str(f)
                    break
            if output_video and Path(output_video).exists():
                break
    print(f"   最终输出视频: {output_video or 'N/A'}")

    steps = info['steps']
    ffmpeg_results = []
    ffprobe_data = {
        'initial': {'packets': [], 'frames': []},
        'steps': [{'packets': [], 'frames': []} for _ in steps],
        'final': {'packets': None, 'frames': None},
    }

    # ── ffmpeg 解码（中间分段 + 最终分段） ──
    if not args.skip_ffmpeg:
        ffmpeg_path = output_dir / f"ffmpeg_output{suffix}.txt"
        mode_label = "image2pipe" if args.image2pipe else "null"
        print(f"\n🎬 ffmpeg 逐段解码 ({mode_label} 模式): {ffmpeg_path}")
        try:
            ffmpeg_results = generate_ffmpeg_output(steps, str(ffmpeg_path),
                                                    use_image2pipe=args.image2pipe)
        except Exception as e:
            print(f"[WARN] ffmpeg 生成异常: {e}，继续后续步骤")

    # ── ffprobe 帧数统计（初始 + 各步骤 + 最终视频） ──
    if not args.skip_ffprobe:
        ffprob_path = output_dir / f"ffprob_output{suffix}.txt"
        print(f"\n🔍 ffprobe 帧数统计: {ffprob_path}")
        try:
            ffprobe_data = generate_ffprob_output(info, output_video, str(ffprob_path))
        except Exception as e:
            print(f"[WARN] ffprobe 生成异常: {e}，继续后续步骤")

    # ── xlsx ──
    if not args.skip_xlsx:
        xlsx_path = output_dir / f"帧数差值统计汇总_output{suffix}.xlsx"
        print(f"\n📊 生成 xlsx: {xlsx_path}")
        try:
            generate_xlsx(info, ffprobe_data, ffmpeg_results, output_video, str(xlsx_path))
        except Exception as e:
            print(f"[ERROR] xlsx 生成失败: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "=" * 60)
    print("  ✅ 全流程检测分析文件生成完毕！")
    print("=" * 60)


if __name__ == '__main__':
    main()
