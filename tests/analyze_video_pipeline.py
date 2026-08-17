#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频增强流水线 全流程检测分析脚本
=================================
合并 analyze_video_ifrnet.py / analyze_video_realesrgan.py 为全自动全流程分析：

  根据 log_file 自动解析：
    1. 插帧/超分步骤顺序（"阶段 N: Step x/y — ..." 标题，兼容单步骤旧日志）
    2. 初始分段目录（由第一个步骤的临时输出分段位置反推，如 ../segments）
    3. 中间分段位置（前面步骤的分段输出即为后面步骤的分段输入）
    4. 最终输出视频文件（合并行 / 📤 输出行）

生成以下检测分析文件：
  1. ffmpeg_output{suffix}.txt           — ffmpeg 逐段解码日志（每个步骤的输出分段）
  2. ffprob_output{suffix}.txt           — ffprobe 逐段帧数统计（初始分段 + 各步骤输出 + 最终视频）
  3. 帧数差值统计汇总_output{suffix}.xlsx — 帧数差值汇总 Excel
       Sheet1 全流程汇总  — 每段跨步骤帧数链（输入→插帧→超分→最终）
       Sheet2..N 各步骤明细 — 帧数差值/异常检测/性能统计
       末 Sheet ffmpeg错误详情

用法:
    python analyze_video_pipeline.py <log_file> [--suffix X] [--skip-ffmpeg] ...

示例:
    python analyze_video_pipeline.py temp/output28-9.txt
    python analyze_video_pipeline.py temp/output28-9.txt --segments-dir ./segments --skip-ffmpeg
"""

import argparse
import re
import subprocess
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl 未安装，将不会生成 xlsx 文件。安装: pip install openpyxl")


# ═══════════════════════════════════════════════
# 1. 日志解析（自动识别步骤顺序与分段位置）
# ═══════════════════════════════════════════════

def _split_prefix(basename: str):
    """'interpolated_segment_003.mp4' -> ('interpolated_segment', 3)"""
    m = re.match(r'(.+?)_(\d{3})\.mp4$', basename)
    if m:
        return m.group(1), int(m.group(2))
    return None, None


def _detect_step_type(name: str, text: str) -> str:
    """根据步骤名称/内容判定步骤类型: interpolate | upscale"""
    probe = (name or '') + '\n' + (text[:2000] if text else '')
    if re.search(r'超分|ESRGAN|esrgan|Real-ESRGAN', probe):
        return 'upscale'
    if re.search(r'插帧|IFRNet|ifrnet', probe):
        return 'interpolate'
    return 'unknown'


def parse_pipeline_log(log_path: str) -> dict:
    """解析全流程流水线日志，自动识别步骤顺序、分段位置与最终输出"""
    with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    info = {
        'mode': '',                 # 处理模式 interpolate_then_upscale 等
        'num_segments': 0,
        'segment_duration': 300,
        'source_video': '',
        'output_video': '',
        'total_output_size': '',
        'total_elapsed': '',
        'steps': [],                # 按顺序的步骤列表
    }

    # ── 全局配置摘要 ──
    m = re.search(r'处理模式\s*[:：]\s*(\S+)', content)
    if m:
        info['mode'] = m.group(1).strip()
    m = re.search(r'分段时长\s*[:：]\s*(\d+)', content)
    if m:
        info['segment_duration'] = int(m.group(1))
    m = re.search(r'^\s*输入\s*[:：]\s*(.+\.(?:mp4|avi|mkv))\s*$', content, re.M)
    if m:
        info['source_video'] = m.group(1).strip()

    # 最终输出视频：优先合并行，其次 📤 输出行，最后"最终输出"
    m = re.search(r'🔗\s*合并\s*\d+\s*个最终分段\s*→\s*(.+\.(?:mp4|avi|mkv))', content)
    if not m:
        m = re.search(r'(?:📤\s*输出(?:文件)?|最终输出)\s*[:：]\s*(.+\.(?:mp4|avi|mkv))', content)
    if m:
        info['output_video'] = m.group(1).strip()

    m = re.search(r'📦\s*文件大小\s*[:：]\s*([\d.]+\s*[KMGT]B)', content)
    if m:
        info['total_output_size'] = m.group(1)
    m = re.search(r'总用时\s*[:：]\s*([\d:.]+)', content)
    if m:
        info['total_elapsed'] = m.group(1)

    # ── 切分步骤（阶段标题），兼容无阶段标题的单步骤旧日志 ──
    stage_pattern = re.compile(r'阶段\s*(\d+)\s*[:：]\s*Step\s*(\d+)\s*/\s*(\d+)\s*—\s*(.+)')
    stage_matches = list(stage_pattern.finditer(content))

    step_blocks = []  # (name, text)
    if stage_matches:
        for si, sm in enumerate(stage_matches):
            end_pos = stage_matches[si + 1].start() if si + 1 < len(stage_matches) else len(content)
            # 阶段 3（后处理）等非处理步骤不计
            name = sm.group(4).strip()
            if not re.search(r'插帧|超分|IFRNet|ESRGAN', name):
                continue
            step_blocks.append((name, content[sm.start():end_pos]))
    else:
        # 单步骤日志：整体作为一个步骤
        whole_type = _detect_step_type('', content)
        step_blocks.append(('IFRNet 插帧' if whole_type == 'interpolate' else 'Real-ESRGAN 超分', content))

    # ── 逐步骤解析分段 ──
    segment_pattern = re.compile(
        r'(?:🎬|🎨)\s*片段\s*(\d+)/(\d+):\s*(\S*segment_(\d+)\.mp4)'
    )
    frame_pattern = re.compile(r'原始帧:\s*(\d+)\s*→\s*输出帧:\s*(\d+)')
    time_pattern = re.compile(r'输出时长\s*([\d:.]+)\s*\|\s*耗时\s*([\d:.]+)')
    # 输出文件：兼容 "   输出: path (xx MB)" 与 "✅ 输出: path (xx MB)"
    output_file_pattern = re.compile(
        r'(?:✅\s*)?输出:\s*(.+?segment_\d+\.mp4)\s*\(([\d.]+\s*[KMGT]B)\)'
    )
    perf_pattern = re.compile(
        r'\[性能统计\]\s*总时间\s*[:：]\s*([\d.]+)秒\s*\|\s*平均\s*[:：]\s*([\d.]+)ms\s*\|\s*FPS\s*[:：]\s*([\d.]+)'
    )
    gfpgan_pattern = re.compile(r'\[性能统计\]\s*GFPGAN\s*模式\s*[:：]\s*(\S+)')
    model_load_pattern = re.compile(r'总耗时（含模型加载）\s*[:：]\s*([\d分秒]+)')
    estimate_pattern = re.compile(r'已用时\s*[:：]\s*([\d:.]+)\s*,\s*预计剩余\s*[:：]\s*([\d:.]+)')

    for order, (step_name, step_text) in enumerate(step_blocks, 1):
        step_type = _detect_step_type(step_name, step_text)

        step = {
            'order': order,
            'name': step_name,
            'type': step_type,               # interpolate | upscale
            'factor': None,                  # 插帧倍数 或 超分倍数
            'gfpgan_mode': 'unknown',
            'input_prefix': 'segment',
            'output_prefix': None,
            'input_dir': '',                 # 待 main() 推断
            'output_dir': '',                # 由首个输出文件反推
            'segments': [],
        }

        if step_type == 'interpolate':
            m = re.search(r'插帧倍数\s*[:：]\s*(\d+)', step_text)
        else:
            m = re.search(r'超分倍数\s*[:：]\s*(\d+)', step_text)
        if m:
            step['factor'] = int(m.group(1))

        gm = gfpgan_pattern.search(step_text)
        if gm:
            step['gfpgan_mode'] = gm.group(1).strip()

        seg_starts = [(m.start(), int(m.group(1)), m.group(3), int(m.group(4)))
                      for m in segment_pattern.finditer(step_text)]

        for idx, (start_pos, seg_num, in_name, seg_id) in enumerate(seg_starts):
            end_pos = seg_starts[idx + 1][0] if idx + 1 < len(seg_starts) else len(step_text)
            seg_text = step_text[start_pos:end_pos]

            seg = {
                'index': seg_id,
                'number': seg_num,
                'input_name': in_name,
                'raw_frames': None,          # 日志原始帧（= 本步骤输入帧）
                'output_frames': None,       # 日志提示输出帧
                'output_duration': '',
                'elapsed': '',
                'output_file': '',
                'output_size': '',
                'model_load_time': '',
                'perf_total_sec': None,
                'perf_avg_ms': None,
                'perf_fps': None,
                'elapsed_estimate': '',
                'remaining_estimate': '',
            }

            # 输入前缀（第一段确定一次）
            if idx == 0:
                pfx, _ = _split_prefix(in_name)
                if pfx:
                    step['input_prefix'] = pfx

            fm = frame_pattern.search(seg_text)
            if fm:
                seg['raw_frames'] = int(fm.group(1))
                seg['output_frames'] = int(fm.group(2))

            tm = time_pattern.search(seg_text)
            if tm:
                seg['output_duration'] = tm.group(1)
                seg['elapsed'] = tm.group(2)

            ofm = output_file_pattern.search(seg_text)
            if ofm:
                seg['output_file'] = ofm.group(1).strip()
                seg['output_size'] = ofm.group(2).strip()
                if not step['output_dir']:
                    step['output_dir'] = str(Path(seg['output_file']).parent)
                    pfx, _ = _split_prefix(Path(seg['output_file']).name)
                    if pfx:
                        step['output_prefix'] = pfx

            pm = perf_pattern.search(seg_text)
            if pm:
                seg['perf_total_sec'] = float(pm.group(1))
                seg['perf_avg_ms'] = float(pm.group(2))
                seg['perf_fps'] = float(pm.group(3))

            mlm = model_load_pattern.search(seg_text)
            if mlm:
                seg['model_load_time'] = mlm.group(1)

            em = estimate_pattern.search(seg_text)
            if em:
                seg['elapsed_estimate'] = em.group(1)
                seg['remaining_estimate'] = em.group(2)

            step['segments'].append(seg)

        info['steps'].append(step)

    # 分段总数：取各步骤最大分段数
    if info['steps']:
        info['num_segments'] = max(len(s['segments']) for s in info['steps'])
    if info['num_segments'] == 0:
        m = re.search(r'✅\s*共\s*(\d+)\s*个片段', content)
        if m:
            info['num_segments'] = int(m.group(1))

    return info


def resolve_step_dirs(info: dict, log_dir: Path, segments_dir_arg: str = None):
    """
    推断每个步骤的输入/输出目录：
      - 第一个步骤的输入 = 初始分段目录（由其输出目录反推 ../segments，或 CLI 指定）
      - 后续步骤的输入 = 前一步骤的输出目录
    """
    prev_output_dir = None
    for si, step in enumerate(info['steps']):
        if si == 0:
            cand = []
            if segments_dir_arg:
                cand.append(Path(segments_dir_arg))
            if step['output_dir']:
                cand.append(Path(step['output_dir']).parent / 'segments')  # ../segments 反推
            cand.append(log_dir / 'segments')
            step['input_dir'] = ''
            for c in cand:
                if c.exists() and list(c.glob(f"{step['input_prefix']}_*.mp4")):
                    step['input_dir'] = str(c)
                    break
            if not step['input_dir']:
                step['input_dir'] = str(cand[0])
        else:
            step['input_dir'] = prev_output_dir
        prev_output_dir = step['output_dir']


# ═══════════════════════════════════════════════
# 2. ffmpeg 解码 + 错误检测
# ═══════════════════════════════════════════════

def _ffmpeg_decode(video_path: str, use_image2pipe: bool = False, timeout: int = 300) -> dict:
    """ffmpeg 解码（image2pipe 完整管线 或 null 轻量），流式捕获帧数/dup/drop/错误"""
    result = {'frame_count': None, 'errors': [], 'dup_count': None, 'drop_count': None}
    if use_image2pipe:
        cmd = ["ffmpeg", "-hide_banner", "-i", str(video_path), "-f", "image2pipe", "-vcodec", "png", "-"]
    else:
        cmd = ["ffmpeg", "-hide_banner", "-i", str(video_path), "-f", "null", "-"]
    try:
        proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE,
                                text=True, bufsize=1)
        try:
            for line in proc.stderr:
                stripped = line.strip()
                m = re.search(r'frame=\s*(\d+)', stripped)
                if m:
                    result['frame_count'] = int(m.group(1))
                m_dup = re.search(r'dup=(\d+)', stripped)
                m_drop = re.search(r'drop=(\d+)', stripped)
                if m_dup and m_drop and 'Lsize=' in stripped:
                    result['dup_count'] = int(m_dup.group(1))
                    result['drop_count'] = int(m_drop.group(1))
                lower = stripped.lower()
                if any(kw in lower for kw in [
                    'illegal', 'error', 'missing', 'corrupt',
                    'duplicated', 'duplicate', 'decoder',
                    'non-existing pps', 'no frame',
                    'more than 1000 frames duplicated',
                ]):
                    result['errors'].append(stripped)
            proc.wait(timeout=timeout)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg 解码超时")
    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")
    return result


def _find_segment_file(directory: Path, prefix: str, seg_idx: int) -> Path:
    """优先精确前缀匹配，回退 *segment_XXX.mp4 通配"""
    exact = directory / f"{prefix}_{seg_idx:03d}.mp4"
    if exact.exists():
        return exact
    matches = list(directory.glob(f"*segment_{seg_idx:03d}.mp4"))
    if matches:
        return matches[0]
    return exact


def generate_ffmpeg_output(info: dict, output_path: str, use_image2pipe: bool = False):
    """对每个步骤的输出分段执行 ffmpeg 解码。返回 {step_order: (frames, errors, dup_drop)}"""
    results = {}
    mode = "image2pipe" if use_image2pipe else "null"
    with open(output_path, 'w', encoding='utf-8') as out:
        for step in info['steps']:
            order = step['order']
            out.write(f"\n{'=' * 60}\n")
            out.write(f" Step {order}/{len(info['steps'])} — {step['name']} ({step['type']}) [{mode}]\n")
            out.write(f" output_dir: {step['output_dir']}\n")
            out.write(f"{'=' * 60}\n\n")
            frames, errors_list, dup_drop = [], [], []
            out_dir = Path(step['output_dir']) if step['output_dir'] else Path('.')
            for seg in step['segments']:
                seg_file = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
                out.write(f"$ ffmpeg -hide_banner -i {seg_file.name} -f {mode} -\n")
                if not seg_file.exists():
                    print(f"  [WARN] 文件不存在: {seg_file}")
                    frames.append(None)
                    errors_list.append([f"[FILE NOT FOUND] {seg_file}"])
                    dup_drop.append(None)
                    out.write("[FILE NOT FOUND]\n\n")
                    continue
                r = _ffmpeg_decode(str(seg_file), use_image2pipe=use_image2pipe)
                frames.append(r['frame_count'])
                errors_list.append(r['errors'])
                dup_drop.append({'dup': r['dup_count'], 'drop': r['drop_count']})
                print(f"  [ffmpeg] Step{order} {seg_file.name}: frame={r['frame_count']}, "
                      f"dup={r['dup_count']}, drop={r['drop_count']}, errors={len(r['errors'])}")
                if r['frame_count'] is not None:
                    out.write(f"frame={r['frame_count']}\n")
                for err in r['errors']:
                    out.write(err + '\n')
                out.write('\n')
            results[order] = (frames, errors_list, dup_drop)
    print(f"✅ ffmpeg_output 已生成: {output_path}")
    return results


# ═══════════════════════════════════════════════
# 3. ffprobe 帧数统计
# ═══════════════════════════════════════════════

def _ffprobe_count(video_path: str, count_mode: str = 'packets', timeout: int = 60) -> str:
    """ffprobe 获取帧/包计数。count_mode: 'packets' | 'frames'"""
    flag = '-count_packets' if count_mode == 'packets' else '-count_frames'
    entry = 'stream=nb_read_packets' if count_mode == 'packets' else 'stream=nb_read_frames'
    cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0", flag,
           "-show_entries", entry, "-of", "csv=p=0", str(video_path)]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        for text in (result.stdout.strip(), result.stderr.strip()):
            m = re.search(r'(\d+)', text)
            if m:
                return m.group(1)
        return result.stdout.strip() if result.stdout.strip() else "[ERROR]"
    except subprocess.TimeoutExpired:
        return "[TIMEOUT]"
    except FileNotFoundError:
        return "[SKIP: ffprobe not found]"


def _to_int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def generate_ffprob_output(info: dict, output_path: str):
    """
    对初始分段 + 各步骤输出分段 + 最终输出视频执行 ffprobe 统计。
    中间分段（前步骤输出=后步骤输入）只统计一次。
    返回 {step_order: {'in_packets': [], 'in_frames': [], 'out_packets': [], 'out_frames': []},
           'final': {'packets': X, 'frames': Y}}
    """
    data = {}
    cache = {}  # path -> {'packets': X, 'frames': Y}，避免重复统计中间分段

    def probe(video_path: Path, mode: str, timeout: int) -> str:
        key = str(video_path)
        if key not in cache:
            cache[key] = {}
        if mode not in cache[key]:
            cache[key][mode] = _ffprobe_count(key, mode, timeout)
        return cache[key][mode]

    with open(output_path, 'w', encoding='utf-8') as out:
        out.write("=" * 70 + "\n")
        out.write(" ffprobe 帧数统计 — 视频增强全流程\n")
        out.write(f" 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        out.write("=" * 70 + "\n")

        for step in info['steps']:
            order = step['order']
            sdata = {'in_packets': [], 'in_frames': [], 'out_packets': [], 'out_frames': []}
            out.write(f"\n── Step {order}: {step['name']} ({step['type']}) ──\n")
            out.write(f"   input_dir : {step['input_dir']}  (prefix: {step['input_prefix']})\n")
            out.write(f"   output_dir: {step['output_dir']}  (prefix: {step['output_prefix']})\n")

            in_dir = Path(step['input_dir']) if step['input_dir'] else None
            out_dir = Path(step['output_dir']) if step['output_dir'] else None

            for mode, key, timeout in (('packets', 'in_packets', 60), ('frames', 'in_frames', 600)):
                out.write(f"\n  [input]  nb_read_{mode}:\n")
                for seg in step['segments']:
                    f = _find_segment_file(in_dir, step['input_prefix'], seg['index']) if in_dir else None
                    if f and f.exists():
                        c = probe(f, mode, timeout)
                        print(f"  [ffprobe] Step{order} in/{mode} {f.name}: {c}")
                        sdata[key].append(c)
                        out.write(f"    {f.name}: {c}\n")
                    else:
                        sdata[key].append("[NOT FOUND]")
                        out.write(f"    {step['input_prefix']}_{seg['index']:03d}.mp4: [FILE NOT FOUND]\n")

            for mode, key, timeout in (('packets', 'out_packets', 60), ('frames', 'out_frames', 600)):
                out.write(f"\n  [output] nb_read_{mode}:\n")
                for seg in step['segments']:
                    f = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index']) if out_dir else None
                    if f and f.exists():
                        c = probe(f, mode, timeout)
                        print(f"  [ffprobe] Step{order} out/{mode} {f.name}: {c}")
                        sdata[key].append(c)
                        out.write(f"    {f.name}: {c}\n")
                    else:
                        sdata[key].append("[NOT FOUND]")
                        out.write(f"    {(step['output_prefix'] or 'segment')}_{seg['index']:03d}.mp4: [FILE NOT FOUND]\n")

            data[order] = sdata

        # ── 最终输出视频 ──
        final_video = info.get('output_video', '')
        data['final'] = {'packets': None, 'frames': None}
        if final_video and Path(final_video).exists():
            out.write(f"\n── 最终输出视频: {Path(final_video).name} ──\n")
            c = _ffprobe_count(final_video, 'packets', 120)
            print(f"  [ffprobe] final/packets {Path(final_video).name}: {c}")
            data['final']['packets'] = c
            out.write(f"  nb_read_packets: {c}\n")
            c = _ffprobe_count(final_video, 'frames', 3600)
            print(f"  [ffprobe] final/frames  {Path(final_video).name}: {c}")
            data['final']['frames'] = c
            out.write(f"  nb_read_frames : {c}\n")
        elif final_video:
            out.write(f"\n── 最终输出视频不存在: {final_video} ──\n")

        out.write("\n" + "=" * 70 + "\n")

    print(f"✅ ffprob_output 已生成: {output_path}")
    return data


# ═══════════════════════════════════════════════
# 4. 帧数链计算
# ═══════════════════════════════════════════════

def build_chain(info: dict, ffprobe_data: dict, ffmpeg_results: dict):
    """
    计算每段跨步骤帧数链。
    返回 rows: [{'index', 'chain': [{step_order, in_frames, expected, log_out, actual, diff,
                                     log_diff, ffmpeg_val, ffmpeg_diff, dup, drop, errors}], }]
    expected 规则:
      interpolate: (输入帧 - 1) * 倍数 + 1
      upscale:     输入帧（= 前一步骤实际输出，ffprobe 优先）
    """
    rows = []
    n = info['num_segments']
    prev_actual = [None] * n  # 前一步骤实际输出帧数

    for step in info['steps']:
        order = step['order']
        sdata = ffprobe_data.get(order, {})
        in_pk = sdata.get('in_packets', [])
        in_fr = sdata.get('in_frames', [])
        out_pk = sdata.get('out_packets', [])
        out_fr = sdata.get('out_frames', [])
        ff_frames, ff_errors, ff_dd = ffmpeg_results.get(order, ([], [], []))

        for i, seg in enumerate(step['segments']):
            while len(rows) <= i:
                rows.append({'index': rows.__len__(), 'chain': []})

            # 输入帧数：ffprobe frames > packets > 前一步骤实际输出 > 日志 raw
            in_frames = (_to_int(in_fr[i]) if i < len(in_fr) else None) or \
                        (_to_int(in_pk[i]) if i < len(in_pk) else None)
            if in_frames is None:
                in_frames = prev_actual[i] if prev_actual[i] is not None else (seg['raw_frames'] or 0)

            if step['type'] == 'interpolate':
                mult = step['factor'] or 2
                expected = (in_frames - 1) * mult + 1
            else:
                expected = in_frames

            actual = (_to_int(out_fr[i]) if i < len(out_fr) else None) or \
                     (_to_int(out_pk[i]) if i < len(out_pk) else None)
            log_out = seg['output_frames'] or 0

            ff_val = ff_frames[i] if i < len(ff_frames) else None
            dd = ff_dd[i] if i < len(ff_dd) else None
            errs = ff_errors[i] if i < len(ff_errors) else []

            entry = {
                'step_order': order,
                'step_type': step['type'],
                'in_frames': in_frames,
                'expected': expected,
                'log_out': log_out,
                'log_diff': log_out - expected,
                'actual': actual,
                'diff': (actual - expected) if actual is not None else None,
                'ff_val': ff_val,
                'ff_diff': (ff_val - expected) if ff_val is not None else None,
                'dup': dd.get('dup') if dd else None,
                'drop': dd.get('drop') if dd else None,
                'errors': errs,
                'seg': seg,
            }
            rows[i]['chain'].append(entry)
            prev_actual[i] = actual if actual is not None else (log_out or prev_actual[i])

    return rows


def detect_anomalies(entry: dict):
    """返回 (anomaly_str, is_bad, is_warn)"""
    errors = entry['errors']
    has_illegal = any('illegal' in e.lower() for e in errors)
    has_missing = any('missing' in e.lower() for e in errors)
    has_duplicate = any('more than 1000 frames duplicated' in e.lower() for e in errors)
    has_no_frame = any('no frame' in e.lower() for e in errors)
    has_decode_error = any('decode_slice_header error' in e.lower() for e in errors)
    has_pps_error = any('non-existing pps' in e.lower() for e in errors)

    parts = []
    if has_illegal:
        parts.append('⚠ illegal buffer(花屏)')
    if has_pps_error:
        parts.append('⚠ PPS丢失')
    if has_decode_error:
        parts.append('⚠ 解码错误')
    if has_no_frame:
        parts.append('❌ no frame!(丢帧)')
    if has_missing:
        parts.append('⚠ missing ref(丢帧)')
    if has_duplicate:
        parts.append('🔁 大量重复帧(>1000)')
    if entry['dup'] is not None and entry['dup'] > 100:
        parts.append(f"🔁 dup={entry['dup']}")
    # 超分场景偏差应严格为 0，插帧容忍 ±2
    tol = 0 if entry['step_type'] == 'upscale' else 2
    if entry['log_diff'] is not None and abs(entry['log_diff']) > tol:
        parts.append(f"📉 日志偏差={entry['log_diff']}")
    if entry['diff'] is not None and abs(entry['diff']) > max(2, tol):
        parts.append(f"📉 ffprobe偏差={entry['diff']}")
    if entry['ff_diff'] is not None and abs(entry['ff_diff']) > max(2, tol):
        parts.append(f"📉 拆帧偏差={entry['ff_diff']}")

    joined = ''.join(parts)
    is_bad = any(kw in joined for kw in ['❌', '⚠ illegal', '⚠ PPS'])
    is_warn = (not is_bad) and len(parts) > 0
    return ('\n'.join(parts) if parts else '✅ 正常'), is_bad, is_warn


# ═══════════════════════════════════════════════
# 5. xlsx 生成
# ═══════════════════════════════════════════════

def generate_xlsx(info: dict, ffprobe_data: dict, ffmpeg_results: dict, chain_rows: list, output_path: str):
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl 未安装，跳过 xlsx 生成")
        return

    wb = openpyxl.Workbook()

    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='微软雅黑', size=14, bold=True, color='1F3864')
    normal_font = Font(name='微软雅黑', size=10)
    warn_font = Font(name='微软雅黑', size=10, color='CC6600')
    bad_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
    ok_font = Font(name='微软雅黑', size=10, color='008000')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bad_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
    sum_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    sum_font = Font(name='微软雅黑', size=10, bold=True)

    steps = info['steps']
    n_steps = len(steps)

    def style_cell(cell, font=normal_font, align=center_align, fill=None):
        cell.font = font
        cell.alignment = align
        cell.border = thin_border
        if fill:
            cell.fill = fill

    def diff_font(dv):
        if dv is None:
            return normal_font, None
        if abs(dv) > 2:
            return bad_font, bad_fill
        if abs(dv) > 0.5:
            return warn_font, warn_fill
        return ok_font, None

    # ══ Sheet 1: 全流程汇总 ══
    ws = wb.active
    ws.title = "全流程汇总"

    # 列: 分段序号 | 分段编号 | (每步: 输入/期望/日志输出/实际输出/差值) | 异常 | 最终比对
    headers = ['分段序号', '分段编号']
    for s in steps:
        headers += [f"Step{s['order']}\n输入帧", f"Step{s['order']}\n期望输出",
                    f"Step{s['order']}\n实际输出", f"Step{s['order']}\n差值"]
    headers += ['异常检测\n(全部步骤)']

    num_cols = len(headers)
    widths = [8, 16] + [12, 12, 12, 10] * n_steps + [40]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = '视频增强全流程 — 帧数链统计汇总'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # 概览
    row = 4
    step_desc = ' → '.join(f"Step{s['order']}:{s['name']}" for s in steps)
    info_items = [
        ('源视频', info.get('source_video', 'N/A')),
        ('最终输出', info.get('output_video', 'N/A')),
        ('处理模式', info.get('mode', 'N/A')),
        ('步骤链', step_desc),
        ('初始分段目录', steps[0]['input_dir'] if steps else 'N/A'),
        ('分段数', str(info.get('num_segments', 0))),
        ('分段时长', f"{info.get('segment_duration', 300)}秒"),
        ('输出大小', info.get('total_output_size', 'N/A')),
        ('总用时', info.get('total_elapsed', 'N/A')),
    ]
    final_probe = ffprobe_data.get('final', {})
    if final_probe.get('frames'):
        info_items.append(('最终视频帧数', f"packets={final_probe.get('packets')} frames={final_probe.get('frames')}"))
    for key, val in info_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws[f'A{row}'] = key
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=min(8, num_cols))
        ws[f'C{row}'] = val
        ws[f'C{row}'].font = normal_font
        row += 1
    row += 1

    header_row = row
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=ci, value=h)
        style_cell(cell, header_font, center_align, header_fill)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 48
    row += 1

    total_bad = 0
    total_warn = 0
    sum_expected = [0] * n_steps
    sum_actual = [0] * n_steps

    for i, r in enumerate(chain_rows):
        anomaly_all = []
        row_vals = [i + 1, f"segment_{r['index']:03d}"]
        for si, entry in enumerate(r['chain']):
            a_str, is_bad, is_warn = detect_anomalies(entry)
            if is_bad:
                total_bad += 1
                anomaly_all.append(f"[Step{entry['step_order']}] " + a_str.replace('\n', ' / '))
            elif is_warn:
                total_warn += 1
                anomaly_all.append(f"[Step{entry['step_order']}] " + a_str.replace('\n', ' / '))
            row_vals += [entry['in_frames'], entry['expected'],
                         entry['actual'] if entry['actual'] is not None else 'N/A',
                         entry['diff'] if entry['diff'] is not None else 'N/A']
            sum_expected[si] += entry['expected'] or 0
            if entry['actual'] is not None:
                sum_actual[si] += entry['actual']
        anomaly_str = '\n'.join(anomaly_all) if anomaly_all else '✅ 正常'
        row_vals.append(anomaly_str)

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            if ci == num_cols:
                font = bad_font if anomaly_all and any('❌' in a or 'illegal' in a or 'PPS' in a for a in anomaly_all) else (warn_font if anomaly_all else ok_font)
                fill = bad_fill if '❌' in anomaly_str or 'illegal' in anomaly_str or 'PPS' in anomaly_str else (warn_fill if anomaly_all else None)
                style_cell(cell, font, left_align, fill)
            elif ci >= 3 and (ci - 3) % 4 == 3:  # 差值列
                dv = val if isinstance(val, (int, float)) else None
                f, fill = diff_font(dv)
                style_cell(cell, f, center_align, fill)
            else:
                style_cell(cell)
        ws.row_dimensions[row].height = max(24, len(anomaly_all) * 16 + 10)
        row += 1

    # 汇总行
    sum_vals = ['汇总', '']
    for si in range(n_steps):
        d = sum_actual[si] - sum_expected[si]
        sum_vals += ['', sum_expected[si], sum_actual[si], d]
    summary_txt = f"❌ 异常段 {total_bad} | ⚠ 警告 {total_warn}" if (total_bad or total_warn) else '✅ 全部正常'
    sum_vals.append(summary_txt)
    for ci, val in enumerate(sum_vals, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        style_cell(cell, sum_font, center_align if ci != num_cols else left_align, sum_fill)
    row += 1

    # 最终视频比对行
    if final_probe.get('frames') and sum_actual:
        expected_final = sum_actual[-1]
        final_frames = _to_int(final_probe['frames']) or 0
        final_diff = final_frames - expected_final
        fvals = ['最终视频', Path(info.get('output_video', '')).name] + [''] * (num_cols - 4) + [final_diff, '']
        fvals[2 + (n_steps - 1) * 4 + 2] = final_frames  # 放到最后一步"实际输出"列
        for ci, val in enumerate(fvals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            if val == final_diff and ci > 2:
                f, fill = diff_font(final_diff)
                style_cell(cell, f, center_align, fill)
            else:
                style_cell(cell, sum_font, center_align, sum_fill)

    ws.freeze_panes = f'A{header_row + 1}'

    # ══ Sheet 2..N: 各步骤明细 ══
    for step in steps:
        order = step['order']
        ws_s = wb.create_sheet(f"Step{order}-{step['type']}")
        sdata = ffprobe_data.get(order, {})
        in_pk = sdata.get('in_packets', [])
        in_fr = sdata.get('in_frames', [])
        out_pk = sdata.get('out_packets', [])
        out_fr = sdata.get('out_frames', [])
        ff_frames, ff_errors, ff_dd = ffmpeg_results.get(order, ([], [], []))

        headers_s = ['分段序号', '分段编号', '输入帧数', '期望输出', '日志输出', '差值1\n(日志-期望)',
                     'ffprobe\npackets', '差值2', 'ffprobe\nframes', '差值3',
                     'ffmpeg\n拆帧', '差值4', 'dup', 'drop',
                     '异常检测', '输出时长', '处理耗时', '性能统计\n(总s/平均ms/FPS)',
                     'GFPGAN', '输出大小', '输出文件名']
        widths_s = [8, 22, 11, 11, 11, 12, 12, 9, 12, 9, 11, 9, 8, 8, 30, 12, 12, 22, 10, 11, 40]
        ncols_s = len(headers_s)

        ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_s)
        ws_s['A1'] = f"Step {order}/{n_steps} — {step['name']} ({step['type']})"
        ws_s['A1'].font = title_font
        ws_s['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_s.row_dimensions[1].height = 28

        ws_s['A2'] = f"输入目录: {step['input_dir']}"
        ws_s['A3'] = f"输出目录: {step['output_dir']}"
        factor_desc = f"插帧倍数: {step['factor']}×" if step['type'] == 'interpolate' else f"超分倍数: {step['factor']}×"
        ws_s['A4'] = f"{factor_desc} | 期望规则: " + \
                     ("(输入-1)×倍数+1" if step['type'] == 'interpolate' else "输入帧（超分不改变帧数）")
        for rr in (2, 3, 4):
            ws_s[f'A{rr}'].font = Font(name='微软雅黑', size=9, color='666666')

        hr = 6
        for ci, (h, w) in enumerate(zip(headers_s, widths_s), 1):
            cell = ws_s.cell(row=hr, column=ci, value=h)
            style_cell(cell, header_font, center_align, header_fill)
            ws_s.column_dimensions[get_column_letter(ci)].width = w
        ws_s.row_dimensions[hr].height = 44

        rr = hr + 1
        step_bad = 0
        step_warn = 0
        for i, seg in enumerate(step['segments']):
            chain_entry = chain_rows[i]['chain'][order - 1] if i < len(chain_rows) and len(chain_rows[i]['chain']) >= order else None
            if chain_entry is None:
                continue
            a_str, is_bad, is_warn = detect_anomalies(chain_entry)
            if is_bad:
                step_bad += 1
            elif is_warn:
                step_warn += 1

            perf_str = ''
            if seg.get('perf_total_sec') is not None:
                perf_str = (f"{seg['perf_total_sec']:.1f}s / {seg.get('perf_avg_ms', 0):.1f}ms / "
                            f"FPS {seg.get('perf_fps', 0):.2f}")
            out_name = Path(seg['output_file']).name if seg.get('output_file') else ''
            vals = [
                i + 1, f"{step['output_prefix']}_{seg['index']:03d}",
                chain_entry['in_frames'], chain_entry['expected'],
                chain_entry['log_out'], chain_entry['log_diff'],
                _to_int(out_pk[i]) if i < len(out_pk) and _to_int(out_pk[i]) is not None else 'N/A',
                chain_entry['diff'] if chain_entry['actual'] is not None and _to_int(out_pk[i]) is None else
                ((_to_int(out_pk[i]) - chain_entry['expected']) if i < len(out_pk) and _to_int(out_pk[i]) is not None else 'N/A'),
                _to_int(out_fr[i]) if i < len(out_fr) and _to_int(out_fr[i]) is not None else 'N/A',
                chain_entry['diff'] if chain_entry['diff'] is not None else 'N/A',
                chain_entry['ff_val'] if chain_entry['ff_val'] is not None else 'N/A',
                chain_entry['ff_diff'] if chain_entry['ff_diff'] is not None else 'N/A',
                chain_entry['dup'] if chain_entry['dup'] is not None else '',
                chain_entry['drop'] if chain_entry['drop'] is not None else '',
                a_str, seg.get('output_duration', ''), seg.get('elapsed', ''),
                perf_str, step.get('gfpgan_mode', '') if step['type'] == 'upscale' else '',
                seg.get('output_size', ''), out_name,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws_s.cell(row=rr, column=ci, value=val)
                if ci == 15:
                    font = bad_font if is_bad else (warn_font if is_warn else ok_font)
                    fill = bad_fill if is_bad else (warn_fill if is_warn else None)
                    style_cell(cell, font, left_align, fill)
                elif ci in (6, 8, 10, 12):
                    dv = val if isinstance(val, (int, float)) else None
                    f, fill = diff_font(dv)
                    style_cell(cell, f, center_align, fill)
                else:
                    style_cell(cell)
            ws_s.row_dimensions[rr].height = max(26, a_str.count('\n') * 16 + 12)
            rr += 1

        # 步骤汇总行
        sum_line = f"异常段 {step_bad} | 警告 {step_warn}" if (step_bad or step_warn) else '✅ 全部正常'
        svals = ['汇总', ''] + [''] * 12 + [sum_line] + [''] * 6
        for ci, val in enumerate(svals, 1):
            cell = ws_s.cell(row=rr, column=ci, value=val)
            style_cell(cell, sum_font, center_align if ci != 15 else left_align, sum_fill)
        ws_s.freeze_panes = f'A{hr + 1}'

    # ══ 末 Sheet: ffmpeg 错误详情 ══
    ws2 = wb.create_sheet("ffmpeg错误详情")
    ws2.merge_cells('A1:F1')
    ws2['A1'] = 'ffmpeg 解码 — 逐段错误详情（全部步骤）'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28
    headers2 = ['步骤', '分段', 'ffmpeg拆帧帧数', 'dup', 'drop', '错误/警告信息']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        style_cell(cell, header_font, center_align, header_fill)
    for col, w in zip('ABCDEF', (10, 34, 16, 10, 10, 80)):
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[3].height = 26

    rr = 4
    for step in steps:
        order = step['order']
        ff_frames, ff_errors, ff_dd = ffmpeg_results.get(order, ([], [], []))
        for i, seg in enumerate(step['segments']):
            errs = ff_errors[i] if i < len(ff_errors) else []
            dd = ff_dd[i] if i < len(ff_dd) else None
            vals = [f"Step{order}",
                    f"{step['output_prefix']}_{seg['index']:03d}",
                    ff_frames[i] if i < len(ff_frames) else 'N/A',
                    dd.get('dup', '') if dd else '', dd.get('drop', '') if dd else '',
                    '\n'.join(errs[:30]) if errs else '✅ 无错误']
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=rr, column=ci, value=val)
                style_cell(cell, bad_font if (errs and ci == 6) else normal_font,
                           left_align if ci == 6 else center_align)
            ws2.row_dimensions[rr].height = max(26, min(300, len(errs) * 16))
            rr += 1
    ws2.freeze_panes = 'A4'

    wb.save(output_path)
    print(f"✅ xlsx 已生成: {output_path}")


# ═══════════════════════════════════════════════
# 6. 主函数
# ═══════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='视频增强流水线 全流程检测分析脚本（自动解析插帧/超分步骤顺序与分段位置）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('log_file', help='流水线输出日志文件 (如 output28-9.txt)')
    parser.add_argument('--suffix', default=None,
                        help='输出文件后缀 (默认: 从日志文件名推导, 如 output28-9.txt -> 28-9)')
    parser.add_argument('--segments-dir', default=None,
                        help='初始分段目录 (默认: 由第一个步骤输出目录反推 ../segments)')
    parser.add_argument('--output-video', default=None,
                        help='最终输出视频路径 (默认: 从日志自动解析)')
    parser.add_argument('--output-dir', default=None,
                        help='分析文件输出目录 (默认: 日志同目录)')
    parser.add_argument('--image2pipe', action='store_true',
                        help='ffmpeg 使用 -f image2pipe 完整管线（慢，可捕获 dup/drop；默认 -f null 快速模式）')
    parser.add_argument('--skip-ffmpeg', action='store_true', help='跳过 ffmpeg 解码')
    parser.add_argument('--skip-ffprobe', action='store_true', help='跳过 ffprobe 帧数统计')
    parser.add_argument('--skip-xlsx', action='store_true', help='跳过 xlsx 生成')

    args = parser.parse_args()

    log_path = Path(args.log_file).resolve()
    if not log_path.exists():
        print(f"[ERROR] 日志文件不存在: {log_path}")
        sys.exit(1)

    log_dir = log_path.parent
    output_dir = Path(args.output_dir) if args.output_dir else log_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    suffix = args.suffix
    if suffix is None:
        suffix = re.sub(r'^output', '', log_path.stem) or 'pipeline'

    # ── 解析日志：步骤顺序 / 分段位置 / 最终输出 ──
    print(f"📖 解析日志: {log_path}")
    info = parse_pipeline_log(str(log_path))
    if args.output_video:
        info['output_video'] = args.output_video

    resolve_step_dirs(info, log_dir, args.segments_dir)

    print(f"   处理模式: {info['mode'] or 'N/A'} | 分段数: {info['num_segments']}")
    for step in info['steps']:
        print(f"   Step {step['order']}: {step['name']} [{step['type']}] "
              f"factor={step['factor']} 段数={len(step['segments'])}")
        print(f"      输入: {step['input_dir']}")
        print(f"      输出: {step['output_dir']}")
    print(f"   最终输出: {info.get('output_video', 'N/A')}")

    if not info['steps']:
        print("[ERROR] 未能从日志中解析出任何处理步骤")
        sys.exit(1)

    # ── ffprobe 帧数统计 ──
    ffprobe_data = {}
    if not args.skip_ffprobe:
        ffprob_path = output_dir / f"ffprob_output{suffix}.txt"
        print(f"\n🔍 生成 ffprobe 输出: {ffprob_path}")
        try:
            ffprobe_data = generate_ffprob_output(info, str(ffprob_path))
        except Exception as e:
            print(f"[WARN] ffprobe 生成异常: {e}，继续后续步骤")

    # ── ffmpeg 解码 ──
    ffmpeg_results = {}
    if not args.skip_ffmpeg:
        ffmpeg_path = output_dir / f"ffmpeg_output{suffix}.txt"
        mode_label = "image2pipe" if args.image2pipe else "null"
        print(f"\n🎬 生成 ffmpeg 输出 ({mode_label} 解码模式): {ffmpeg_path}")
        try:
            ffmpeg_results = generate_ffmpeg_output(info, str(ffmpeg_path), use_image2pipe=args.image2pipe)
        except Exception as e:
            print(f"[WARN] ffmpeg 生成异常: {e}，继续后续步骤")

    # ── 帧数链计算 ──
    chain_rows = build_chain(info, ffprobe_data, ffmpeg_results)

    # ── xlsx ──
    if not args.skip_xlsx:
        xlsx_path = output_dir / f"帧数差值统计汇总_output{suffix}.xlsx"
        print(f"\n📊 生成 xlsx: {xlsx_path}")
        try:
            generate_xlsx(info, ffprobe_data, ffmpeg_results, chain_rows, str(xlsx_path))
        except Exception as e:
            print(f"[ERROR] xlsx 生成失败: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "=" * 60)
    print("  ✅ 全流程检测分析文件生成完毕！")
    print("=" * 60)


if __name__ == '__main__':
    main()
