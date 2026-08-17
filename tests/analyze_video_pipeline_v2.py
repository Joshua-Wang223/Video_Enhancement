#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频增强流水线 全流程检测分析脚本 v2（CPU 并行计算版）
====================================================
基于 analyze_video_pipeline.py 升级：

  v2 新增能力：
    1. 自动探测系统资源（容器感知）：CPU 逻辑核数 / RAM 总量与可用量。
       CPU 优先读 cgroup v1/v2 quota；内存用 cgroup limit 修正 psutil/sysconf 结果；
       Windows 回退 GlobalMemoryStatusEx；Linux 回退 /proc/meminfo/sysconf
    2. 并行执行引擎：ffprobe 帧数统计 + ffmpeg 解码全部任务化，
       支持 多线程(thread) / 多进程(process) 两种并行模式
       - 多个小文件（多日志）：跨日志汇总任务到统一并行池，打满系统资源
       - 单个大文件（单日志多分段）：分段级任务并行
    3. workers 自动计算: min(CPU核数, RAM总计×90%/单任务内存估计)，可用 --workers 覆盖
    4. 输出文件格式与 v1 完全一致（结果按原始分段顺序落盘，与并行完成顺序无关）

  v1 原有功能（保持不变）：
    根据 log_file 自动解析：
      1. 插帧/超分步骤顺序（"阶段 N: Step x/y — ..." 标题，兼容单步骤旧日志）
      2. 初始分段目录（由第一个步骤的临时输出分段位置反推，如 ../segments）
      3. 中间分段位置（前面步骤的分段输出即为后面步骤的分段输入）
      4. 最终输出视频文件（合并行 / 📤 输出行）

生成以下检测分析文件：
  1. ffmpeg_output{suffix}.txt           — ffmpeg 逐段解码日志（每个步骤的输出分段）
  2. ffprob_output{suffix}.txt           — ffprobe 逐段帧数统计（初始分段 + 各步骤输出 + 最终视频）
  3. 帧数差值统计汇总_output{suffix}.xlsx — 帧数差值汇总 Excel
       Sheet1 全流程汇总  — 每段跨步骤帧数链（输入→插帧→超分→最终）
       Sheet2..N 各步骤明细 — 帧数差值/异常检测/性能统计
       末 Sheet ffmpeg错误详情

用法:
    python analyze_video_pipeline_v2.py <log_file> [log_file2 ...] [选项]

示例:
    python analyze_video_pipeline_v2.py temp/output28-9.txt
    python analyze_video_pipeline_v2.py temp/output28-9.txt temp/output28-10.txt --workers 8
    python analyze_video_pipeline_v2.py temp/output28-9.txt --parallel process --task-ram-mb 512
    python analyze_video_pipeline_v2.py temp/output28-9.txt --segments-dir ./segments --skip-ffmpeg
"""

import argparse
import math
import os
import re
import shutil
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl 未安装，将不会生成 xlsx 文件。安装: pip install openpyxl")


# ═══════════════════════════════════════════════
# 1. 日志解析（自动识别步骤顺序与分段位置）
# ═══════════════════════════════════════════════

def _split_prefix(basename: str):
    """'interpolated_segment_003.mp4' -> ('interpolated_segment', 3)"""
    m = re.match(r'(.+?)_(\d{3})\.mp4$', basename)
    if m:
        return m.group(1), int(m.group(2))
    return None, None


def _detect_step_type(name: str, text: str) -> str:
    """根据步骤名称/内容判定步骤类型: interpolate | upscale"""
    probe = (name or '') + '\n' + (text[:2000] if text else '')
    if re.search(r'超分|ESRGAN|esrgan|Real-ESRGAN', probe):
        return 'upscale'
    if re.search(r'插帧|IFRNet|ifrnet', probe):
        return 'interpolate'
    return 'unknown'


def parse_pipeline_log(log_path: str) -> dict:
    """解析全流程流水线日志，自动识别步骤顺序、分段位置与最终输出"""
    with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    info = {
        'mode': '',                 # 处理模式 interpolate_then_upscale 等
        'num_segments': 0,
        'segment_duration': 300,
        'source_video': '',
        'output_video': '',
        'total_output_size': '',
        'total_elapsed': '',
        'steps': [],                # 按顺序的步骤列表
    }

    # ── 全局配置摘要 ──
    m = re.search(r'处理模式\s*[:：]\s*(\S+)', content)
    if m:
        info['mode'] = m.group(1).strip()
    m = re.search(r'分段时长\s*[:：]\s*(\d+)', content)
    if m:
        info['segment_duration'] = int(m.group(1))
    m = re.search(r'^\s*输入\s*[:：]\s*(.+\.(?:mp4|avi|mkv))\s*$', content, re.M)
    if m:
        info['source_video'] = m.group(1).strip()

    # 最终输出视频：优先合并行，其次 📤 输出行，最后"最终输出"
    m = re.search(r'🔗\s*合并\s*\d+\s*个最终分段\s*→\s*(.+\.(?:mp4|avi|mkv))', content)
    if not m:
        m = re.search(r'(?:📤\s*输出(?:文件)?|最终输出)\s*[:：]\s*(.+\.(?:mp4|avi|mkv))', content)
    if m:
        info['output_video'] = m.group(1).strip()

    m = re.search(r'📦\s*文件大小\s*[:：]\s*([\d.]+\s*[KMGT]B)', content)
    if m:
        info['total_output_size'] = m.group(1)
    m = re.search(r'总用时\s*[:：]\s*([\d:.]+)', content)
    if m:
        info['total_elapsed'] = m.group(1)

    # ── 切分步骤（阶段标题），兼容无阶段标题的单步骤旧日志 ──
    stage_pattern = re.compile(r'阶段\s*(\d+)\s*[:：]\s*Step\s*(\d+)\s*/\s*(\d+)\s*—\s*(.+)')
    stage_matches = list(stage_pattern.finditer(content))

    step_blocks = []  # (name, text)
    if stage_matches:
        for si, sm in enumerate(stage_matches):
            end_pos = stage_matches[si + 1].start() if si + 1 < len(stage_matches) else len(content)
            # 阶段 3（后处理）等非处理步骤不计
            name = sm.group(4).strip()
            if not re.search(r'插帧|超分|IFRNet|ESRGAN', name):
                continue
            step_blocks.append((name, content[sm.start():end_pos]))
    else:
        # 单步骤日志：整体作为一个步骤
        whole_type = _detect_step_type('', content)
        step_blocks.append(('IFRNet 插帧' if whole_type == 'interpolate' else 'Real-ESRGAN 超分', content))

    # ── 逐步骤解析分段 ──
    segment_pattern = re.compile(
        r'(?:🎬|🎨)\s*片段\s*(\d+)/(\d+):\s*(\S*segment_(\d+)\.mp4)'
    )
    frame_pattern = re.compile(r'原始帧:\s*(\d+)\s*→\s*输出帧:\s*(\d+)')
    time_pattern = re.compile(r'输出时长\s*([\d:.]+)\s*\|\s*耗时\s*([\d:.]+)')
    # 输出文件：兼容 "   输出: path (xx MB)" 与 "✅ 输出: path (xx MB)"
    output_file_pattern = re.compile(
        r'(?:✅\s*)?输出:\s*(.+?segment_\d+\.mp4)\s*\(([\d.]+\s*[KMGT]B)\)'
    )
    perf_pattern = re.compile(
        r'\[性能统计\]\s*总时间\s*[:：]\s*([\d.]+)秒\s*\|\s*平均\s*[:：]\s*([\d.]+)ms\s*\|\s*FPS\s*[:：]\s*([\d.]+)'
    )
    gfpgan_pattern = re.compile(r'\[性能统计\]\s*GFPGAN\s*模式\s*[:：]\s*(\S+)')
    model_load_pattern = re.compile(r'总耗时（含模型加载）\s*[:：]\s*([\d分秒]+)')
    estimate_pattern = re.compile(r'已用时\s*[:：]\s*([\d:.]+)\s*,\s*预计剩余\s*[:：]\s*([\d:.]+)')
    # IFRNet GPU 性能行: [GPU0] 完成 | 原始帧=8974 → 输出帧=17947 | 12.6 原始帧/s
    ifrnet_perf_pattern = re.compile(
        r'\[GPU\d+\]\s*完成\s*\|\s*原始帧=(\d+)\s*→\s*输出帧=(\d+)\s*\|\s*([\d.]+)\s*原始帧/s'
    )

    for order, (step_name, step_text) in enumerate(step_blocks, 1):
        step_type = _detect_step_type(step_name, step_text)

        step = {
            'order': order,
            'name': step_name,
            'type': step_type,               # interpolate | upscale
            'factor': None,                  # 插帧倍数 或 超分倍数
            'gfpgan_mode': 'unknown',
            'input_prefix': 'segment',
            'output_prefix': None,
            'input_dir': '',                 # 待 main() 推断
            'output_dir': '',                # 由首个输出文件反推
            'segments': [],
        }

        if step_type == 'interpolate':
            m = re.search(r'插帧倍数\s*[:：]\s*(\d+)', step_text)
        else:
            m = re.search(r'超分倍数\s*[:：]\s*(\d+)', step_text)
        if m:
            step['factor'] = int(m.group(1))

        gm = gfpgan_pattern.search(step_text)
        if gm:
            step['gfpgan_mode'] = gm.group(1).strip()

        seg_starts = [(m.start(), int(m.group(1)), m.group(3), int(m.group(4)))
                      for m in segment_pattern.finditer(step_text)]

        for idx, (start_pos, seg_num, in_name, seg_id) in enumerate(seg_starts):
            end_pos = seg_starts[idx + 1][0] if idx + 1 < len(seg_starts) else len(step_text)
            seg_text = step_text[start_pos:end_pos]

            seg = {
                'index': seg_id,
                'number': seg_num,
                'input_name': in_name,
                'raw_frames': None,          # 日志原始帧（= 本步骤输入帧）
                'output_frames': None,       # 日志提示输出帧
                'output_duration': '',
                'elapsed': '',
                'output_file': '',
                'output_size': '',
                'model_load_time': '',
                'perf_total_sec': None,
                'perf_avg_ms': None,
                'perf_fps': None,
                'ifrnet_raw_fps': None,       # IFRNet GPU 报告的原始帧/s
                'elapsed_estimate': '',
                'remaining_estimate': '',
            }

            # 输入前缀（第一段确定一次）
            if idx == 0:
                pfx, _ = _split_prefix(in_name)
                if pfx:
                    step['input_prefix'] = pfx

            fm = frame_pattern.search(seg_text)
            if fm:
                seg['raw_frames'] = int(fm.group(1))
                seg['output_frames'] = int(fm.group(2))

            tm = time_pattern.search(seg_text)
            if tm:
                seg['output_duration'] = tm.group(1)
                seg['elapsed'] = tm.group(2)

            ofm = output_file_pattern.search(seg_text)
            if ofm:
                seg['output_file'] = ofm.group(1).strip()
                seg['output_size'] = ofm.group(2).strip()
                if not step['output_dir']:
                    step['output_dir'] = str(Path(seg['output_file']).parent)
                    pfx, _ = _split_prefix(Path(seg['output_file']).name)
                    if pfx:
                        step['output_prefix'] = pfx

            pm = perf_pattern.search(seg_text)
            if pm:
                seg['perf_total_sec'] = float(pm.group(1))
                seg['perf_avg_ms'] = float(pm.group(2))
                seg['perf_fps'] = float(pm.group(3))

            # IFRNet GPU 性能行: [GPU0] 完成 | 原始帧=8974 → 输出帧=17947 | 12.6 原始帧/s
            ipm = ifrnet_perf_pattern.search(seg_text)
            if ipm:
                seg['ifrnet_raw_fps'] = float(ipm.group(3))
                # GPU 行报告的帧数与 ✅ 插帧完成！行可能重复，兜底补齐
                if seg['raw_frames'] is None:
                    seg['raw_frames'] = int(ipm.group(1))
                if seg['output_frames'] is None:
                    seg['output_frames'] = int(ipm.group(2))

            mlm = model_load_pattern.search(seg_text)
            if mlm:
                seg['model_load_time'] = mlm.group(1)

            em = estimate_pattern.search(seg_text)
            if em:
                seg['elapsed_estimate'] = em.group(1)
                seg['remaining_estimate'] = em.group(2)

            step['segments'].append(seg)

        info['steps'].append(step)

    # 分段总数：取各步骤最大分段数
    if info['steps']:
        info['num_segments'] = max(len(s['segments']) for s in info['steps'])
    if info['num_segments'] == 0:
        m = re.search(r'✅\s*共\s*(\d+)\s*个片段', content)
        if m:
            info['num_segments'] = int(m.group(1))

    return info


def resolve_step_dirs(info: dict, log_dir: Path, segments_dir_arg: str = None):
    """
    推断每个步骤的输入/输出目录：
      - 第一个步骤的输入 = 初始分段目录（由其输出目录反推 ../segments，或 CLI 指定）
      - 后续步骤的输入 = 前一步骤的输出目录
    """
    prev_output_dir = None
    for si, step in enumerate(info['steps']):
        if si == 0:
            cand = []
            if segments_dir_arg:
                cand.append(Path(segments_dir_arg))
            if step['output_dir']:
                cand.append(Path(step['output_dir']).parent / 'segments')  # ../segments 反推
            cand.append(log_dir / 'segments')
            step['input_dir'] = ''
            for c in cand:
                if c.exists() and list(c.glob(f"{step['input_prefix']}_*.mp4")):
                    step['input_dir'] = str(c)
                    break
            if not step['input_dir']:
                step['input_dir'] = str(cand[0])
        else:
            step['input_dir'] = prev_output_dir
        prev_output_dir = step['output_dir']


# ═══════════════════════════════════════════════
# 2. ffmpeg 解码 + 错误检测
# ═══════════════════════════════════════════════

def _ffmpeg_decode(video_path: str, use_image2pipe: bool = False, timeout: int = 300) -> dict:
    """ffmpeg 解码（image2pipe 完整管线 或 null 轻量），流式捕获帧数/dup/drop/错误。
    dup/drop 从每条进度行独立捕获（不再要求同含 Lsize=），最后一条即为总计。"""
    result = {'frame_count': None, 'errors': [], 'dup_count': None, 'drop_count': None}
    if use_image2pipe:
        cmd = ["ffmpeg", "-hide_banner", "-i", str(video_path), "-f", "image2pipe", "-vcodec", "png", "-"]
    else:
        cmd = ["ffmpeg", "-hide_banner", "-i", str(video_path), "-f", "null", "-"]
    try:
        proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE,
                                text=True, bufsize=1)
        try:
            for line in proc.stderr:
                stripped = line.strip()
                m = re.search(r'frame=\s*(\d+)', stripped)
                if m:
                    result['frame_count'] = int(m.group(1))
                m_dup = re.search(r'dup=(\d+)', stripped)
                if m_dup:
                    result['dup_count'] = int(m_dup.group(1))
                m_drop = re.search(r'drop=(\d+)', stripped)
                if m_drop:
                    result['drop_count'] = int(m_drop.group(1))
                lower = stripped.lower()
                if any(kw in lower for kw in [
                    'illegal', 'error', 'missing', 'corrupt',
                    'duplicated', 'duplicate', 'decoder',
                    'non-existing pps', 'no frame',
                    'more than 1000 frames duplicated',
                ]):
                    result['errors'].append(stripped)
            proc.wait(timeout=timeout)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg 解码超时")
    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")
    return result


def _find_segment_file(directory: Path, prefix: str, seg_idx: int) -> Path:
    """优先精确前缀匹配，回退 *segment_XXX.mp4 通配"""
    exact = directory / f"{prefix}_{seg_idx:03d}.mp4"
    if exact.exists():
        return exact
    matches = list(directory.glob(f"*segment_{seg_idx:03d}.mp4"))
    if matches:
        return matches[0]
    return exact


def collect_ffmpeg_tasks(info: dict, use_image2pipe: bool = False, timeout: int = 300) -> list:
    """
    收集单个日志全部 ffmpeg 解码任务（v2 并行用）。
    任务 key: ('ffmpeg', path, use_image2pipe)，跨步骤/跨日志去重由 run 阶段统一处理。
    """
    tasks = []
    seen = set()
    for step in info['steps']:
        out_dir = Path(step['output_dir']) if step['output_dir'] else Path('.')
        for seg in step['segments']:
            seg_file = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
            key = ('ffmpeg', str(seg_file), use_image2pipe)
            # 不存在的文件不入任务池（生成阶段仍按 v1 逻辑写 [FILE NOT FOUND]）
            if key in seen or not seg_file.exists():
                continue
            seen.add(key)
            tasks.append({'kind': 'ffmpeg', 'key': key, 'path': str(seg_file),
                          'image2pipe': use_image2pipe, 'timeout': timeout})
    return tasks


def generate_ffmpeg_output(info: dict, output_path: str, use_image2pipe: bool = False,
                           prefetched: dict = None):
    """对每个步骤的输出分段执行 ffmpeg 解码。返回 {step_order: (frames, errors, dup_drop)}

    v2: prefetched 为并行预取结果 {('ffmpeg', path, use_image2pipe): result}，
        命中时直接读取（保持 v1 输出文件格式与分段顺序），缺项回退串行解码。
    """
    results = {}
    mode = "image2pipe" if use_image2pipe else "null"
    inline = prefetched is None  # 无预取结果时为 v1 串行模式（逐条打印进度）
    with open(output_path, 'w', encoding='utf-8') as out:
        for step in info['steps']:
            order = step['order']
            out.write(f"\n{'=' * 60}\n")
            out.write(f" Step {order}/{len(info['steps'])} — {step['name']} ({step['type']}) [{mode}]\n")
            out.write(f" output_dir: {step['output_dir']}\n")
            if not use_image2pipe:
                out.write(f" ⚠ dup/drop 列为空：-f null 模式不输出 dup/drop，请使用 --image2pipe 获取\n")
            out.write(f"{'=' * 60}\n\n")
            frames, errors_list, dup_drop = [], [], []
            out_dir = Path(step['output_dir']) if step['output_dir'] else Path('.')
            for seg in step['segments']:
                seg_file = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
                out.write(f"$ ffmpeg -hide_banner -i {seg_file.name} -f {mode} -\n")
                if not seg_file.exists():
                    print(f"  [WARN] 文件不存在: {seg_file}")
                    frames.append(None)
                    errors_list.append([f"[FILE NOT FOUND] {seg_file}"])
                    dup_drop.append(None)
                    out.write("[FILE NOT FOUND]\n\n")
                    continue
                r = None
                if prefetched is not None:
                    r = prefetched.get(('ffmpeg', str(seg_file), use_image2pipe))
                if r is None:
                    r = _ffmpeg_decode(str(seg_file), use_image2pipe=use_image2pipe)
                frames.append(r['frame_count'])
                errors_list.append(r['errors'])
                dup_drop.append({'dup': r['dup_count'], 'drop': r['drop_count']})
                if inline:
                    print(f"  [ffmpeg] Step{order} {seg_file.name}: frame={r['frame_count']}, "
                          f"dup={r['dup_count']}, drop={r['drop_count']}, errors={len(r['errors'])}")
                if r['frame_count'] is not None:
                    out.write(f"frame={r['frame_count']}\n")
                for err in r['errors']:
                    out.write(err + '\n')
                out.write('\n')
            results[order] = (frames, errors_list, dup_drop)
    print(f"✅ ffmpeg_output 已生成: {output_path}")
    return results


# ═══════════════════════════════════════════════
# 3. ffprobe 帧数统计
# ═══════════════════════════════════════════════

def _ffprobe_count(video_path: str, count_mode: str = 'packets', timeout: int = 60) -> str:
    """ffprobe 获取帧/包计数。count_mode: 'packets' | 'frames'"""
    flag = '-count_packets' if count_mode == 'packets' else '-count_frames'
    entry = 'stream=nb_read_packets' if count_mode == 'packets' else 'stream=nb_read_frames'
    cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0", flag,
           "-show_entries", entry, "-of", "csv=p=0", str(video_path)]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        for text in (result.stdout.strip(), result.stderr.strip()):
            m = re.search(r'(\d+)', text)
            if m:
                return m.group(1)
        return result.stdout.strip() if result.stdout.strip() else "[ERROR]"
    except subprocess.TimeoutExpired:
        return "[TIMEOUT]"
    except FileNotFoundError:
        return "[SKIP: ffprobe not found]"


def _to_int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _parse_elapsed_to_sec(elapsed_str: str) -> float:
    """
    解析耗时字符串为总秒数，用于没有 [性能统计] 的步骤推算 FPS/平均耗时。
    支持格式: 'HH:MM:SS.ms' / 'MM:SS.ms' / 'SS.ms'
    """
    if not elapsed_str:
        return None
    try:
        parts = elapsed_str.strip().split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + float(parts[1])
        else:
            return float(parts[0])
    except (ValueError, TypeError):
        return None


def collect_ffprobe_packets_tasks(info: dict) -> list:
    """
    收集 packets 模式（快速）ffprobe 统计任务 — 第一阶段扫描。
    仅读容器 header，不逐帧解码，O(1) 成本获取各分段总帧数。
    任务 key: ('ffprobe', path, 'packets')
    """
    tasks = []
    seen = set()

    def add(video_path, mode, timeout):
        key = ('ffprobe', str(video_path), mode)
        if key in seen or not Path(video_path).exists():
            return
        seen.add(key)
        tasks.append({'kind': 'ffprobe', 'key': key, 'path': str(video_path),
                      'mode': mode, 'timeout': timeout})

    for step in info['steps']:
        in_dir = Path(step['input_dir']) if step['input_dir'] else None
        out_dir = Path(step['output_dir']) if step['output_dir'] else None
        for seg in step['segments']:
            if in_dir:
                f = _find_segment_file(in_dir, step['input_prefix'], seg['index'])
                add(f, 'packets', 60)
            if out_dir:
                f = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
                add(f, 'packets', 60)

    final_video = info.get('output_video', '')
    if final_video and Path(final_video).exists():
        add(final_video, 'packets', 120)

    return tasks


def collect_ffprobe_frames_tasks(info: dict) -> list:
    """
    收集 frames 模式（全量解码）ffprobe 统计任务 — 第二阶段重任务。
    需逐帧解码全部画面，成本 O(帧数)，由第一阶段 packets 结果估算后调度。
    任务 key: ('ffprobe', path, 'frames')
    """
    tasks = []
    seen = set()

    def add(video_path, mode, timeout):
        key = ('ffprobe', str(video_path), mode)
        if key in seen or not Path(video_path).exists():
            return
        seen.add(key)
        tasks.append({'kind': 'ffprobe', 'key': key, 'path': str(video_path),
                      'mode': mode, 'timeout': timeout})

    for step in info['steps']:
        in_dir = Path(step['input_dir']) if step['input_dir'] else None
        out_dir = Path(step['output_dir']) if step['output_dir'] else None
        for seg in step['segments']:
            if in_dir:
                f = _find_segment_file(in_dir, step['input_prefix'], seg['index'])
                add(f, 'frames', 900)
            if out_dir:
                f = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
                add(f, 'frames', 900)

    # 最终合并输出视频：无条件加入（文件可能不存在于本地，但 cost 由分段之和估算）
    final_video = info.get('output_video', '')
    if final_video:
        fv_str = str(final_video)
        fv_key = ('ffprobe', fv_str, 'frames')
        if fv_key not in seen:
            seen.add(fv_key)
            tasks.append({'kind': 'ffprobe', 'key': fv_key, 'path': fv_str,
                          'mode': 'frames', 'timeout': 3600})

    return tasks


def generate_ffprob_output(info: dict, output_path: str, prefetched: dict = None,
                           skip_frames: bool = False):
    """
    对初始分段 + 各步骤输出分段 + 最终输出视频执行 ffprobe 统计。
    中间分段（前步骤输出=后步骤输入）只统计一次。
    返回 {step_order: {'in_packets': [], 'in_frames': [], 'out_packets': [], 'out_frames': []},
           'final': {'packets': X, 'frames': Y}}

    v2: prefetched 为并行预取结果 {('ffprobe', path, mode): count_str}，
        命中时直接填充缓存（保持 v1 输出文件格式与分段顺序），缺项回退串行统计。
    skip_frames: 跳过 -count_frames（全量解码），-count_packets 快速扫描不受影响。
    """
    data = {}
    cache = {}  # path -> {'packets': X, 'frames': Y}，避免重复统计中间分段
    inline = prefetched is None  # 无预取结果时为 v1 串行模式（逐条打印进度）

    def probe(video_path: Path, mode: str, timeout: int) -> str:
        key = str(video_path)
        if key not in cache:
            cache[key] = {}
        if mode not in cache[key]:
            hit = prefetched.get(('ffprobe', key, mode)) if prefetched is not None else None
            cache[key][mode] = hit if hit is not None else _ffprobe_count(key, mode, timeout)
        return cache[key][mode]

    with open(output_path, 'w', encoding='utf-8') as out:
        out.write("=" * 70 + "\n")
        out.write(" ffprobe 帧数统计 — 视频增强全流程\n")
        out.write(f" 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        out.write("=" * 70 + "\n")

        for step in info['steps']:
            order = step['order']
            sdata = {'in_packets': [], 'in_frames': [], 'out_packets': [], 'out_frames': []}
            out.write(f"\n── Step {order}: {step['name']} ({step['type']}) ──\n")
            out.write(f"   input_dir : {step['input_dir']}  (prefix: {step['input_prefix']})\n")
            out.write(f"   output_dir: {step['output_dir']}  (prefix: {step['output_prefix']})\n")

            in_dir = Path(step['input_dir']) if step['input_dir'] else None
            out_dir = Path(step['output_dir']) if step['output_dir'] else None

            for mode, key, timeout in (('packets', 'in_packets', 60), ('frames', 'in_frames', 900)):
                if skip_frames and mode == 'frames':
                    continue
                out.write(f"\n  [input]  nb_read_{mode}:\n")
                for seg in step['segments']:
                    f = _find_segment_file(in_dir, step['input_prefix'], seg['index']) if in_dir else None
                    if f and f.exists():
                        c = probe(f, mode, timeout)
                        if inline:
                            print(f"  [ffprobe] Step{order} in/{mode} {f.name}: {c}")
                        sdata[key].append(c)
                        out.write(f"    {f.name}: {c}\n")
                    else:
                        sdata[key].append("[NOT FOUND]")
                        out.write(f"    {step['input_prefix']}_{seg['index']:03d}.mp4: [FILE NOT FOUND]\n")

            for mode, key, timeout in (('packets', 'out_packets', 60), ('frames', 'out_frames', 900)):
                if skip_frames and mode == 'frames':
                    continue
                out.write(f"\n  [output] nb_read_{mode}:\n")
                for seg in step['segments']:
                    f = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index']) if out_dir else None
                    if f and f.exists():
                        c = probe(f, mode, timeout)
                        if inline:
                            print(f"  [ffprobe] Step{order} out/{mode} {f.name}: {c}")
                        sdata[key].append(c)
                        out.write(f"    {f.name}: {c}\n")
                    else:
                        sdata[key].append("[NOT FOUND]")
                        out.write(f"    {(step['output_prefix'] or 'segment')}_{seg['index']:03d}.mp4: [FILE NOT FOUND]\n")

            data[order] = sdata

        # ── 最终输出视频 ──
        final_video = info.get('output_video', '')
        data['final'] = {'packets': None, 'frames': None}
        if final_video and Path(final_video).exists():
            out.write(f"\n── 最终输出视频: {Path(final_video).name} ──\n")
            c = probe(Path(final_video), 'packets', 120)
            if inline:
                print(f"  [ffprobe] final/packets {Path(final_video).name}: {c}")
            data['final']['packets'] = c
            out.write(f"  nb_read_packets: {c}\n")
            if not skip_frames:
                c = probe(Path(final_video), 'frames', 3600)
                if inline:
                    print(f"  [ffprobe] final/frames  {Path(final_video).name}: {c}")
                data['final']['frames'] = c
                out.write(f"  nb_read_frames : {c}\n")
        elif final_video:
            out.write(f"\n── 最终输出视频不存在: {final_video} ──\n")

        out.write("\n" + "=" * 70 + "\n")

    print(f"✅ ffprob_output 已生成: {output_path}")
    return data


# ═══════════════════════════════════════════════
# 4. 系统资源探测与并行执行引擎（v2 新增）
# ═══════════════════════════════════════════════

def _read_text_strip(path: str) -> str:
    """读取文件并 strip，失败返回 None。"""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception:
        return None


def _read_int(path: str) -> int:
    text = _read_text_strip(path)
    if text is None:
        return None
    try:
        return int(text)
    except Exception:
        return None


def _parse_cpuset_count(cpuset_text: str) -> int:
    """解析 cpuset.cpus 如 '0-3,5,7-9'，返回核心数。"""
    if not cpuset_text:
        return 0
    count = 0
    for part in cpuset_text.split(','):
        part = part.strip()
        if '-' in part:
            try:
                a, b = part.split('-', 1)
                count += int(b) - int(a) + 1
            except Exception:
                continue
        elif part:
            try:
                count += 1
            except Exception:
                continue
    return count


def detect_cpu_count() -> int:
    """
    容器感知的 CPU 核数探测。
      - cgroup v2: /sys/fs/cgroup/cpu.max
      - cgroup v1: /sys/fs/cgroup/cpu/cpu.cfs_quota_us / cpu.cfs_period_us
      - 无 quota 时回退 cpuset（v1/v2 路径），最后再回退 os.cpu_count()
    os.cpu_count() 在容器里通常返回宿主机核心数，因此优先读 cgroup。
    """
    # cgroup v2
    cpu_max = _read_text_strip('/sys/fs/cgroup/cpu.max')
    if cpu_max:
        parts = cpu_max.split()
        if len(parts) == 2 and parts[0].lower() != 'max':
            try:
                quota = int(parts[0])
                period = int(parts[1])
                if period > 0 and quota > 0:
                    return max(1, math.ceil(quota / period))
            except Exception:
                pass
    # cgroup v1
    quota = _read_int('/sys/fs/cgroup/cpu/cpu.cfs_quota_us')
    period = _read_int('/sys/fs/cgroup/cpu/cpu.cfs_period_us')
    if quota is not None and period and period > 0 and quota > 0:
        return max(1, math.ceil(quota / period))
    # cpuset 回退
    for cpuset_path in (
        '/sys/fs/cgroup/cpuset/cpuset.cpus',
        '/sys/fs/cgroup/cpuset.cpus.effective',
        '/sys/fs/cgroup/cpuset.cpus',
    ):
        cpuset = _read_text_strip(cpuset_path)
        if cpuset:
            cnt = _parse_cpuset_count(cpuset)
            if cnt > 0:
                return cnt
    return os.cpu_count() or 1


def _detect_cgroup_memory_gb():
    """
    读取 cgroup 内存限制与已用量（GB），返回 (limit_gb, usage_gb)。
    如果未设置限制，limit_gb 为 None；'max' 或极大值也表示无限制。
    """
    limit_gb = None
    usage_gb = None
    # cgroup v2
    limit_text = _read_text_strip('/sys/fs/cgroup/memory.max')
    if limit_text is not None:
        if limit_text.lower() != 'max':
            try:
                limit_gb = int(limit_text) / 1e9
            except Exception:
                pass
        usage_text = _read_text_strip('/sys/fs/cgroup/memory.current')
        if usage_text:
            try:
                usage_gb = int(usage_text) / 1e9
            except Exception:
                pass
        return limit_gb, usage_gb
    # cgroup v1
    limit_text = _read_text_strip('/sys/fs/cgroup/memory/memory.limit_in_bytes')
    if limit_text is not None:
        try:
            limit_gb = int(limit_text) / 1e9
        except Exception:
            pass
        usage_text = _read_text_strip('/sys/fs/cgroup/memory/memory.usage_in_bytes')
        if usage_text:
            try:
                usage_gb = int(usage_text) / 1e9
            except Exception:
                pass
    return limit_gb, usage_gb


def _detect_host_ram_gb():
    """
    跨平台探测物理内存，返回 (total_gb, avail_gb)；失败返回 (0.0, 0.0)。
    优先 psutil；Windows 回退 GlobalMemoryStatusEx；Linux 回退 /proc/meminfo → sysconf。
    """
    try:
        import psutil
        vm = psutil.virtual_memory()
        return vm.total / 1e9, vm.available / 1e9
    except ImportError:
        pass
    except Exception:
        pass
    if sys.platform == 'win32':
        try:
            import ctypes

            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [("dwLength", ctypes.c_ulong), ("dwMemoryLoad", ctypes.c_ulong),
                            ("ullTotalPhys", ctypes.c_ulonglong), ("ullAvailPhys", ctypes.c_ulonglong),
                            ("ullTotalPageFile", ctypes.c_ulonglong), ("ullAvailPageFile", ctypes.c_ulonglong),
                            ("ullTotalVirtual", ctypes.c_ulonglong), ("ullAvailVirtual", ctypes.c_ulonglong),
                            ("sullAvailExtendedVirtual", ctypes.c_ulonglong)]

            stat = MEMORYSTATUSEX()
            stat.dwLength = ctypes.sizeof(stat)
            ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(stat))
            return stat.ullTotalPhys / 1e9, stat.ullAvailPhys / 1e9
        except Exception:
            return 0.0, 0.0
    else:
        try:
            with open('/proc/meminfo', 'r', encoding='utf-8') as f:
                meminfo = f.read()
            total = int(re.search(r'MemTotal:\s+(\d+)\s*kB', meminfo).group(1)) / 1e6
            avail = int(re.search(r'MemAvailable:\s+(\d+)\s*kB', meminfo).group(1)) / 1e6
            return total, avail
        except Exception:
            try:
                page_size = os.sysconf('SC_PAGE_SIZE')
                return (os.sysconf('SC_PHYS_PAGES') * page_size / 1e9,
                        os.sysconf('SC_AVPHYS_PAGES') * page_size / 1e9)
            except Exception:
                return 0.0, 0.0


def detect_ram_gb():
    """
    跨平台 + 容器感知探测内存，返回 (total_gb, avail_gb)；失败返回 (0.0, 0.0)。
      - 宿主机：psutil 优先；Windows 回退 GlobalMemoryStatusEx；Linux 回退 /proc/meminfo/sysconf
      - 容器：再用 cgroup v1/v2 的 memory.limit/usage 修正，避免读到宿主机全部内存
    """
    total, avail = _detect_host_ram_gb()
    cg_limit, cg_usage = _detect_cgroup_memory_gb()
    if cg_limit is not None:
        # 如果 limit 极大（如 2^63-1），说明未限制，仍用宿主机值
        total = min(total, cg_limit) if total > 0 else cg_limit
        if cg_usage is not None:
            avail = max(0.0, total - cg_usage)
        else:
            avail = min(avail, total) if avail > 0 else total
    return total, avail


def compute_auto_workers(task_ram_mb: int = 300, reserve_ratio: float = 0.10) -> int:
    """
    按容器感知后的系统资源自动计算并行 workers：
      workers = min(容器可用 CPU, RAM总计*(1-预留比例) / 单任务内存估计)
    ffprobe/ffmpeg 解码子进程本身吃 CPU，故不超过 CPU 核数；RAM 不足时自动缩减。
    默认预留 10% 给系统/其他进程（参考 subtitle_generator_whisper_v2_process.py 思路，
    但按容器 cgroup 修正 total/avail）。
    """
    cpu = detect_cpu_count()
    total_gb, avail_gb = detect_ram_gb()
    by_ram = cpu
    if total_gb > 0:
        usable_gb = total_gb * (1 - reserve_ratio)
    elif avail_gb > 0:
        usable_gb = avail_gb * 0.9
    else:
        usable_gb = 0.0
    usable_mb = max(0.0, usable_gb * 1024)
    if usable_mb > 0:
        by_ram = max(1, int(usable_mb // max(1, task_ram_mb)))
    return max(1, min(cpu, by_ram))


def _extract_segment_frame_counts(prefetched: dict) -> dict:
    """
    从第一阶段 packets 结果中提取各分段文件的帧数。
    返回 {path_str: frame_count_int}。
    无法解析的忽略，缺项后续排序时 cost=0 放在队尾。
    """
    costs = {}
    for (kind, path, mode), val in prefetched.items():
        if kind == 'ffprobe' and mode == 'packets':
            try:
                costs[path] = int(val)
            except (ValueError, TypeError):
                pass
    return costs


def _path_step_priority(path_str: str, step_dir_priority: dict):
    """根据文件路径返回 (step_order, type_priority)。type: output=1(高), input=0(低)。"""
    ps = str(Path(path_str).resolve())
    best = (-1, 0)
    for sd, (so, tp) in step_dir_priority.items():
        if ps.startswith(sd) or sd in ps:
            if (so, tp) > best:
                best = (so, tp)
    return best


def _enrich_and_sort_heavy_tasks(tasks: list, frame_costs: dict,
                                   step_dir_priority: dict = None,
                                   image2pipe: bool = False,
                                   final_video_paths: set = None) -> list:
    """
    Phase 2 重任务排序（与阶段 1 一致）：
      1. 最终视频 (cost 最大)
      2. Step N 输出 → Step N 输入 → Step N-1 输出 → ...
      3. 同组内：-f null 时 ffprobe 先于 ffmpeg；-f image2pipe 时 ffmpeg 先于 ffprobe
      4. 同组同类型内：cost 降序（大文件优先）
    """
    step_dir_priority = step_dir_priority or {}
    final_video_paths = final_video_paths or set()
    for t in tasks:
        t['cost'] = frame_costs.get(t['path'], 0)

    def sort_key(t):
        path = t['path']
        # 最终视频：最高优先级
        if path in final_video_paths:
            return (999, 0, 0, -t['cost'])
        # Step 优先级
        sp = _path_step_priority(path, step_dir_priority)
        order = sp[0]  # 越大越靠后 = 越优先
        typ = sp[1]    # output=1(高), input=0(低)
        # 任务类型优先级（reverse=True 时 kind_p=1 排前）：
        #   -f null: ffprobe(1) 先, ffmpeg(0) 后  — ffprobe 慢，先跑
        #   -f image2pipe: ffmpeg(1) 先, ffprobe(0) 后 — ffmpeg 慢，先跑
        if image2pipe:
            kind_p = 1 if t['kind'] == 'ffmpeg' else 0
        else:
            kind_p = 1 if t['kind'] == 'ffprobe' else 0
        return (order, typ, kind_p, -t['cost'])

    return sorted(tasks, key=sort_key, reverse=True)


def _execute_task(task: dict):
    """执行单个探测/解码任务（模块级函数，ProcessPool 下可 pickle）。
    返回 (result, actual_start_ts) —— 开始时间在 worker 内记录，
    随返回值跨进/线程回传，保证 thread/process 双模式计时准确。"""
    start = time.time()
    if task['kind'] == 'ffprobe':
        return _ffprobe_count(task['path'], task['mode'], task['timeout']), start
    return _ffmpeg_decode(task['path'], use_image2pipe=task.get('image2pipe', False),
                          timeout=task.get('timeout', 300)), start


def _task_error_result(task: dict, exc: Exception):
    """任务异常时生成与正常结果同结构的兜底值，保证后续生成流程不中断"""
    if task['kind'] == 'ffprobe':
        return f"[ERROR: {exc}]"
    return {'frame_count': None, 'errors': [f"[ERROR] 并行执行异常: {exc}"],
            'dup_count': None, 'drop_count': None}


def _print_task_progress(done: int, total: int, task: dict, result,
                         task_duration: float, work_done: int = 0,
                         total_work: int = 0, total_elapsed: float = 0):
    """逐任务进度打印。task_duration=本任务用时，total_elapsed=批次累计（助 ETA 估算）。"""
    name = Path(task['path']).name
    work_str = ""
    if total_work and total_work > 0:
        pct = min(100, work_done * 100 // total_work)
        eta_s = (total_elapsed / max(1, work_done)) * max(0, total_work - work_done) if work_done > 0 else 0
        eta = f" ETA {eta_s/60:.0f}分{eta_s%60:.0f}秒" if eta_s > 0 else ""
        work_str = f" [工作量 {pct}%{eta}]"
    if task['kind'] == 'ffprobe':
        print(f"  [{done}/{total}] ffprobe/{task['mode']} {name}: {result}{work_str}  (用时 {task_duration:.1f}s)")
    elif isinstance(result, dict):
        print(f"  [{done}/{total}] ffmpeg {name}: frame={result.get('frame_count')}, "
              f"dup={result.get('dup_count')}, drop={result.get('drop_count')}, "
              f"errors={len(result.get('errors', []))}{work_str}  (用时 {task_duration:.1f}s)")
    else:
        print(f"  [{done}/{total}] ffmpeg {name}: {result}{work_str}  (用时 {task_duration:.1f}s)")


def run_tasks_parallel(tasks: list, workers: int, parallel_mode: str = 'thread',
                       total_work: int = 0, phase_label: str = '') -> dict:
    """
    并行执行任务列表，返回 {task['key']: result}。
      - workers<=1 时退化为 v1 串行语义
      - parallel_mode: 'thread' / 'process'
      - total_work: 任务总帧数（用于工作量百分比 + ETA），0 则仅显示任务计数
      - phase_label: 阶段标识字符串，如 '阶段 1/2: 快速扫描'
    """
    results = {}
    total = len(tasks)
    if total == 0:
        return results
    workers = max(1, min(workers, total))
    t_batch = time.time()

    if workers == 1:
        work_acc = 0
        for i, t in enumerate(tasks, 1):
            t0 = time.time()
            real_start = t0
            try:
                res, real_start = _execute_task(t)
                results[t['key']] = res
            except Exception as e:
                results[t['key']] = _task_error_result(t, e)
            task_dur = time.time() - real_start
            total_elapsed = time.time() - t_batch
            work_acc += t.get('cost', 0)
            _print_task_progress(i, total, t, results[t['key']], task_dur,
                                 work_done=work_acc, total_work=total_work,
                                 total_elapsed=total_elapsed)
        return results

    executor_cls = ProcessPoolExecutor if parallel_mode == 'process' else ThreadPoolExecutor
    done = 0
    work_acc = 0
    with executor_cls(max_workers=workers) as ex:
        fut2task = {}
        for t in tasks:
            fut2task[ex.submit(_execute_task, t)] = t
        for fut in as_completed(fut2task):
            t = fut2task[fut]
            total_elapsed = time.time() - t_batch
            try:
                res, real_start = fut.result()
                results[t['key']] = res
                task_dur = time.time() - real_start
            except Exception as e:
                results[t['key']] = _task_error_result(t, e)
                task_dur = time.time() - t_batch  # 异常回落：批次已用时间
            done += 1
            work_acc += t.get('cost', 0)
            _print_task_progress(done, total, t, results[t['key']], task_dur,
                                 work_done=work_acc, total_work=total_work,
                                 total_elapsed=total_elapsed)
    return results


# ═══════════════════════════════════════════════
# 5. 帧数链计算
# ═══════════════════════════════════════════════

def build_chain(info: dict, ffprobe_data: dict, ffmpeg_results: dict):
    """
    计算每段跨步骤帧数链。
    返回 rows: [{'index', 'chain': [{step_order, in_frames, expected, log_out, actual, diff,
                                     log_diff, ffmpeg_val, ffmpeg_diff, dup, drop, errors}], }]
    expected 规则:
      interpolate: (输入帧 - 1) * 倍数 + 1
      upscale:     输入帧（= 前一步骤实际输出，ffprobe 优先）
    """
    rows = []
    n = info['num_segments']
    prev_actual = [None] * n  # 前一步骤实际输出帧数

    for step in info['steps']:
        order = step['order']
        sdata = ffprobe_data.get(order, {})
        in_pk = sdata.get('in_packets', [])
        in_fr = sdata.get('in_frames', [])
        out_pk = sdata.get('out_packets', [])
        out_fr = sdata.get('out_frames', [])
        ff_frames, ff_errors, ff_dd = ffmpeg_results.get(order, ([], [], []))

        for i, seg in enumerate(step['segments']):
            while len(rows) <= i:
                rows.append({'index': rows.__len__(), 'chain': []})

            # 输入帧数：ffprobe frames > packets > 前一步骤实际输出 > 日志 raw
            in_frames = (_to_int(in_fr[i]) if i < len(in_fr) else None) or \
                        (_to_int(in_pk[i]) if i < len(in_pk) else None)
            if in_frames is None:
                in_frames = prev_actual[i] if prev_actual[i] is not None else (seg['raw_frames'] or 0)

            if step['type'] == 'interpolate':
                mult = step['factor'] or 2
                expected = (in_frames - 1) * mult + 1
            else:
                expected = in_frames

            actual = (_to_int(out_fr[i]) if i < len(out_fr) else None) or \
                     (_to_int(out_pk[i]) if i < len(out_pk) else None)
            log_out = seg['output_frames'] or 0

            ff_val = ff_frames[i] if i < len(ff_frames) else None
            dd = ff_dd[i] if i < len(ff_dd) else None
            errs = ff_errors[i] if i < len(ff_errors) else []

            entry = {
                'step_order': order,
                'step_type': step['type'],
                'in_frames': in_frames,
                'expected': expected,
                'log_out': log_out,
                'log_diff': log_out - expected,
                'actual': actual,
                'diff': (actual - expected) if actual is not None else None,
                'ff_val': ff_val,
                'ff_diff': (ff_val - expected) if ff_val is not None else None,
                'dup': dd.get('dup') if dd else None,
                'drop': dd.get('drop') if dd else None,
                'errors': errs,
                'seg': seg,
            }
            rows[i]['chain'].append(entry)
            prev_actual[i] = actual if actual is not None else (log_out or prev_actual[i])

    return rows


def detect_anomalies(entry: dict):
    """返回 (anomaly_str, is_bad, is_warn)"""
    errors = entry['errors']
    has_illegal = any('illegal' in e.lower() for e in errors)
    has_missing = any('missing' in e.lower() for e in errors)
    has_duplicate = any('more than 1000 frames duplicated' in e.lower() for e in errors)
    has_no_frame = any('no frame' in e.lower() for e in errors)
    has_decode_error = any('decode_slice_header error' in e.lower() for e in errors)
    has_pps_error = any('non-existing pps' in e.lower() for e in errors)

    parts = []
    if has_illegal:
        parts.append('⚠ illegal buffer(花屏)')
    if has_pps_error:
        parts.append('⚠ PPS丢失')
    if has_decode_error:
        parts.append('⚠ 解码错误')
    if has_no_frame:
        parts.append('❌ no frame!(丢帧)')
    if has_missing:
        parts.append('⚠ missing ref(丢帧)')
    if has_duplicate:
        parts.append('🔁 大量重复帧(>1000)')
    if entry['dup'] is not None and entry['dup'] > 100:
        parts.append(f"🔁 dup={entry['dup']}")
    # 超分场景偏差应严格为 0，插帧容忍 ±2
    tol = 0 if entry['step_type'] == 'upscale' else 2
    if entry['log_diff'] is not None and abs(entry['log_diff']) > tol:
        parts.append(f"📉 日志偏差={entry['log_diff']}")
    if entry['diff'] is not None and abs(entry['diff']) > max(2, tol):
        parts.append(f"📉 ffprobe偏差={entry['diff']}")
    if entry['ff_diff'] is not None and abs(entry['ff_diff']) > max(2, tol):
        parts.append(f"📉 拆帧偏差={entry['ff_diff']}")

    joined = ''.join(parts)
    is_bad = any(kw in joined for kw in ['❌', '⚠ illegal', '⚠ PPS'])
    is_warn = (not is_bad) and len(parts) > 0
    return ('\n'.join(parts) if parts else '✅ 正常'), is_bad, is_warn


# ═══════════════════════════════════════════════
# 6. xlsx 生成
# ═══════════════════════════════════════════════

def generate_xlsx(info: dict, ffprobe_data: dict, ffmpeg_results: dict, chain_rows: list, output_path: str):
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl 未安装，跳过 xlsx 生成")
        return

    wb = openpyxl.Workbook()

    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='微软雅黑', size=14, bold=True, color='1F3864')
    normal_font = Font(name='微软雅黑', size=10)
    warn_font = Font(name='微软雅黑', size=10, color='CC6600')
    bad_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
    ok_font = Font(name='微软雅黑', size=10, color='008000')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bad_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
    sum_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    sum_font = Font(name='微软雅黑', size=10, bold=True)

    steps = info['steps']
    n_steps = len(steps)

    def style_cell(cell, font=normal_font, align=center_align, fill=None):
        cell.font = font
        cell.alignment = align
        cell.border = thin_border
        if fill:
            cell.fill = fill

    def diff_font(dv):
        if dv is None:
            return normal_font, None
        if abs(dv) > 2:
            return bad_font, bad_fill
        if abs(dv) > 0.5:
            return warn_font, warn_fill
        return ok_font, None

    # ══ Sheet 1: 全流程汇总 ══
    ws = wb.active
    ws.title = "全流程汇总"

    # 列: 分段序号 | 分段编号 | (每步: 输入/期望/日志输出/实际输出/差值) | 异常 | 最终比对
    headers = ['分段序号', '分段编号']
    for s in steps:
        headers += [f"Step{s['order']}\n输入帧", f"Step{s['order']}\n期望输出",
                    f"Step{s['order']}\n实际输出", f"Step{s['order']}\n差值"]
    headers += ['异常检测\n(全部步骤)']

    num_cols = len(headers)
    widths = [8, 16] + [12, 12, 12, 10] * n_steps + [40]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = '视频增强全流程 — 帧数链统计汇总'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # 概览
    row = 4
    step_desc = ' → '.join(f"Step{s['order']}:{s['name']}" for s in steps)
    info_items = [
        ('源视频', info.get('source_video', 'N/A')),
        ('最终输出', info.get('output_video', 'N/A')),
        ('处理模式', info.get('mode', 'N/A')),
        ('步骤链', step_desc),
        ('初始分段目录', steps[0]['input_dir'] if steps else 'N/A'),
        ('分段数', str(info.get('num_segments', 0))),
        ('分段时长', f"{info.get('segment_duration', 300)}秒"),
        ('输出大小', info.get('total_output_size', 'N/A')),
        ('总用时', info.get('total_elapsed', 'N/A')),
    ]
    final_probe = ffprobe_data.get('final', {})
    if final_probe.get('frames'):
        info_items.append(('最终视频帧数', f"packets={final_probe.get('packets')} frames={final_probe.get('frames')}"))
    for key, val in info_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws[f'A{row}'] = key
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=min(8, num_cols))
        ws[f'C{row}'] = val
        ws[f'C{row}'].font = normal_font
        row += 1
    row += 1

    header_row = row
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=ci, value=h)
        style_cell(cell, header_font, center_align, header_fill)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 48
    row += 1

    total_bad = 0
    total_warn = 0
    sum_expected = [0] * n_steps
    sum_actual = [0] * n_steps

    for i, r in enumerate(chain_rows):
        anomaly_all = []
        row_vals = [i + 1, f"segment_{r['index']:03d}"]
        for si, entry in enumerate(r['chain']):
            a_str, is_bad, is_warn = detect_anomalies(entry)
            if is_bad:
                total_bad += 1
                anomaly_all.append(f"[Step{entry['step_order']}] " + a_str.replace('\n', ' / '))
            elif is_warn:
                total_warn += 1
                anomaly_all.append(f"[Step{entry['step_order']}] " + a_str.replace('\n', ' / '))
            row_vals += [entry['in_frames'], entry['expected'],
                         entry['actual'] if entry['actual'] is not None else 'N/A',
                         entry['diff'] if entry['diff'] is not None else 'N/A']
            sum_expected[si] += entry['expected'] or 0
            if entry['actual'] is not None:
                sum_actual[si] += entry['actual']
        anomaly_str = '\n'.join(anomaly_all) if anomaly_all else '✅ 正常'
        row_vals.append(anomaly_str)

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            if ci == num_cols:
                font = bad_font if anomaly_all and any('❌' in a or 'illegal' in a or 'PPS' in a for a in anomaly_all) else (warn_font if anomaly_all else ok_font)
                fill = bad_fill if '❌' in anomaly_str or 'illegal' in anomaly_str or 'PPS' in anomaly_str else (warn_fill if anomaly_all else None)
                style_cell(cell, font, left_align, fill)
            elif ci >= 3 and (ci - 3) % 4 == 3:  # 差值列
                dv = val if isinstance(val, (int, float)) else None
                f, fill = diff_font(dv)
                style_cell(cell, f, center_align, fill)
            else:
                style_cell(cell)
        ws.row_dimensions[row].height = max(24, len(anomaly_all) * 16 + 10)
        row += 1

    # 汇总行
    sum_vals = ['汇总', '']
    for si in range(n_steps):
        d = sum_actual[si] - sum_expected[si]
        sum_vals += ['', sum_expected[si], sum_actual[si], d]
    summary_txt = f"❌ 异常段 {total_bad} | ⚠ 警告 {total_warn}" if (total_bad or total_warn) else '✅ 全部正常'
    sum_vals.append(summary_txt)
    for ci, val in enumerate(sum_vals, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        style_cell(cell, sum_font, center_align if ci != num_cols else left_align, sum_fill)
    row += 1

    # 最终视频比对行
    if final_probe.get('frames') and sum_actual:
        expected_final = sum_actual[-1]
        final_frames = _to_int(final_probe['frames']) or 0
        final_diff = final_frames - expected_final
        fvals = ['最终视频', Path(info.get('output_video', '')).name] + [''] * (num_cols - 4) + [final_diff, '']
        fvals[2 + (n_steps - 1) * 4 + 2] = final_frames  # 放到最后一步"实际输出"列
        for ci, val in enumerate(fvals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            if val == final_diff and ci > 2:
                f, fill = diff_font(final_diff)
                style_cell(cell, f, center_align, fill)
            else:
                style_cell(cell, sum_font, center_align, sum_fill)

    ws.freeze_panes = f'A{header_row + 1}'

    # ══ Sheet 2..N: 各步骤明细 ══
    for step in steps:
        order = step['order']
        ws_s = wb.create_sheet(f"Step{order}-{step['type']}")
        sdata = ffprobe_data.get(order, {})
        in_pk = sdata.get('in_packets', [])
        in_fr = sdata.get('in_frames', [])
        out_pk = sdata.get('out_packets', [])
        out_fr = sdata.get('out_frames', [])
        ff_frames, ff_errors, ff_dd = ffmpeg_results.get(order, ([], [], []))

        headers_s = ['分段序号', '分段编号', '输入帧数', '期望输出', '日志输出', '差值1\n(日志-期望)',
                     'ffprobe\npackets', '差值2', 'ffprobe\nframes', '差值3',
                     'ffmpeg\n拆帧', '差值4', 'dup', 'drop',
                     '异常检测', '输出时长', '处理耗时', '性能统计\n(总s/平均ms/FPS)',
                     'GFPGAN', '输出大小', '输出文件名']
        widths_s = [8, 22, 11, 11, 11, 12, 12, 9, 12, 9, 11, 9, 8, 8, 30, 12, 12, 22, 10, 11, 40]
        ncols_s = len(headers_s)

        ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_s)
        ws_s['A1'] = f"Step {order}/{n_steps} — {step['name']} ({step['type']})"
        ws_s['A1'].font = title_font
        ws_s['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_s.row_dimensions[1].height = 28

        ws_s['A2'] = f"输入目录: {step['input_dir']}"
        ws_s['A3'] = f"输出目录: {step['output_dir']}"
        factor_desc = f"插帧倍数: {step['factor']}×" if step['type'] == 'interpolate' else f"超分倍数: {step['factor']}×"
        ws_s['A4'] = f"{factor_desc} | 期望规则: " + \
                     ("(输入-1)×倍数+1" if step['type'] == 'interpolate' else "输入帧（超分不改变帧数）")
        for rr in (2, 3, 4):
            ws_s[f'A{rr}'].font = Font(name='微软雅黑', size=9, color='666666')

        hr = 6
        for ci, (h, w) in enumerate(zip(headers_s, widths_s), 1):
            cell = ws_s.cell(row=hr, column=ci, value=h)
            style_cell(cell, header_font, center_align, header_fill)
            ws_s.column_dimensions[get_column_letter(ci)].width = w
        ws_s.row_dimensions[hr].height = 44

        rr = hr + 1
        step_bad = 0
        step_warn = 0
        for i, seg in enumerate(step['segments']):
            chain_entry = chain_rows[i]['chain'][order - 1] if i < len(chain_rows) and len(chain_rows[i]['chain']) >= order else None
            if chain_entry is None:
                continue
            a_str, is_bad, is_warn = detect_anomalies(chain_entry)
            if is_bad:
                step_bad += 1
            elif is_warn:
                step_warn += 1

            perf_str = ''
            if seg.get('perf_total_sec') is not None:
                perf_str = (f"{seg['perf_total_sec']:.1f}s / {seg.get('perf_avg_ms', 0):.1f}ms / "
                            f"FPS {seg.get('perf_fps', 0):.2f}")
            else:
                # IFRNet: 无 [性能统计] 行，基于 GPU 原始帧/s + 耗时 + 输出帧推算
                out_frames = seg.get('output_frames')
                total_sec = _parse_elapsed_to_sec(seg.get('elapsed', ''))
                ifrnet_gpu_fps = seg.get('ifrnet_raw_fps')
                if total_sec and total_sec > 0 and out_frames:
                    out_fps = out_frames / total_sec
                    avg_ms = total_sec * 1000 / out_frames
                    if ifrnet_gpu_fps:
                        perf_str = (f"{total_sec:.1f}s / {avg_ms:.1f}ms / "
                                    f"出{out_fps:.1f}fps GPU{seg['raw_frames']}帧@{ifrnet_gpu_fps:.1f}raw/s")
                    else:
                        perf_str = (f"(推算) {total_sec:.1f}s / {avg_ms:.1f}ms / "
                                    f"FPS {out_fps:.2f}")
                elif total_sec:
                    perf_str = f"(耗时) {total_sec:.1f}s"
            out_name = Path(seg['output_file']).name if seg.get('output_file') else ''
            vals = [
                i + 1, f"{step['output_prefix']}_{seg['index']:03d}",
                chain_entry['in_frames'], chain_entry['expected'],
                chain_entry['log_out'], chain_entry['log_diff'],
                _to_int(out_pk[i]) if i < len(out_pk) and _to_int(out_pk[i]) is not None else 'N/A',
                chain_entry['diff'] if chain_entry['actual'] is not None and _to_int(out_pk[i]) is None else
                ((_to_int(out_pk[i]) - chain_entry['expected']) if i < len(out_pk) and _to_int(out_pk[i]) is not None else 'N/A'),
                _to_int(out_fr[i]) if i < len(out_fr) and _to_int(out_fr[i]) is not None else 'N/A',
                chain_entry['diff'] if chain_entry['diff'] is not None else 'N/A',
                chain_entry['ff_val'] if chain_entry['ff_val'] is not None else 'N/A',
                chain_entry['ff_diff'] if chain_entry['ff_diff'] is not None else 'N/A',
                chain_entry['dup'] if chain_entry['dup'] is not None else '',
                chain_entry['drop'] if chain_entry['drop'] is not None else '',
                a_str, seg.get('output_duration', ''), seg.get('elapsed', ''),
                perf_str, step.get('gfpgan_mode', '') if step['type'] == 'upscale' else '',
                seg.get('output_size', ''), out_name,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws_s.cell(row=rr, column=ci, value=val)
                if ci == 15:
                    font = bad_font if is_bad else (warn_font if is_warn else ok_font)
                    fill = bad_fill if is_bad else (warn_fill if is_warn else None)
                    style_cell(cell, font, left_align, fill)
                elif ci in (6, 8, 10, 12):
                    dv = val if isinstance(val, (int, float)) else None
                    f, fill = diff_font(dv)
                    style_cell(cell, f, center_align, fill)
                else:
                    style_cell(cell)
            ws_s.row_dimensions[rr].height = max(26, a_str.count('\n') * 16 + 12)
            rr += 1

        # 步骤汇总行
        sum_line = f"异常段 {step_bad} | 警告 {step_warn}" if (step_bad or step_warn) else '✅ 全部正常'
        svals = ['汇总', ''] + [''] * 12 + [sum_line] + [''] * 6
        for ci, val in enumerate(svals, 1):
            cell = ws_s.cell(row=rr, column=ci, value=val)
            style_cell(cell, sum_font, center_align if ci != 15 else left_align, sum_fill)
        ws_s.freeze_panes = f'A{hr + 1}'

    # ══ 末 Sheet: ffmpeg 错误详情 ══
    ws2 = wb.create_sheet("ffmpeg错误详情")
    ws2.merge_cells('A1:F1')
    ws2['A1'] = 'ffmpeg 解码 — 逐段错误详情（全部步骤）| dup/drop 列仅 --image2pipe 模式有值'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28
    headers2 = ['步骤', '分段', 'ffmpeg拆帧帧数', 'dup', 'drop', '错误/警告信息']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        style_cell(cell, header_font, center_align, header_fill)
    for col, w in zip('ABCDEF', (10, 34, 16, 10, 10, 80)):
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[3].height = 26

    rr = 4
    for step in steps:
        order = step['order']
        ff_frames, ff_errors, ff_dd = ffmpeg_results.get(order, ([], [], []))
        for i, seg in enumerate(step['segments']):
            errs = ff_errors[i] if i < len(ff_errors) else []
            dd = ff_dd[i] if i < len(ff_dd) else None
            vals = [f"Step{order}",
                    f"{step['output_prefix']}_{seg['index']:03d}",
                    ff_frames[i] if i < len(ff_frames) else 'N/A',
                    dd.get('dup', '') if dd else '', dd.get('drop', '') if dd else '',
                    '\n'.join(errs[:30]) if errs else '✅ 无错误']
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=rr, column=ci, value=val)
                style_cell(cell, bad_font if (errs and ci == 6) else normal_font,
                           left_align if ci == 6 else center_align)
            ws2.row_dimensions[rr].height = max(26, min(300, len(errs) * 16))
            rr += 1
    ws2.freeze_panes = 'A4'

    wb.save(output_path)
    print(f"✅ xlsx 已生成: {output_path}")


# ═══════════════════════════════════════════════
# 7. 主函数
# ═══════════════════════════════════════════════

def main():
    t_start = time.time()
    parser = argparse.ArgumentParser(
        description='视频增强流水线 全流程检测分析脚本 v2（多日志 + 自动资源探测 + 多进程/多线程并行）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('log_files', nargs='+',
                        help='流水线输出日志文件，支持多个 (如 output28-9.txt output28-10.txt)')
    parser.add_argument('--suffix', default=None,
                        help='输出文件后缀 (仅单日志时有效; 默认: 从日志文件名推导, 如 output28-9.txt -> 28-9)')
    parser.add_argument('--segments-dir', default=None,
                        help='初始分段目录 (默认: 由第一个步骤输出目录反推 ../segments; 多日志时应用于全部日志)')
    parser.add_argument('--output-video', default=None,
                        help='最终输出视频路径 (仅单日志时有效; 默认: 从日志自动解析)')
    parser.add_argument('--output-dir', default=None,
                        help='分析文件输出目录 (默认: 各日志同目录)')
    parser.add_argument('--image2pipe', action='store_true',
                        help='ffmpeg 使用 -f image2pipe 完整管线（慢，可捕获 dup/drop；默认 -f null 快速模式）')
    parser.add_argument('--skip-ffmpeg', action='store_true', help='跳过 ffmpeg 解码')
    parser.add_argument('--skip-ffprobe-frames', action='store_true',
                        help='跳过 ffprobe -count_frames (全量解码，慢)；-count_packets 快速扫描不受影响')
    parser.add_argument('--skip-xlsx', action='store_true', help='跳过 xlsx 生成')
    parser.add_argument('--workers', type=int, default=None,
                        help='并行 worker 数 (默认: 按 CPU 核数/可用RAM 自动计算)')
    parser.add_argument('--parallel', choices=['auto', 'thread', 'process'], default='auto',
                        help='并行模式: auto=自动(子进程型任务默认 thread) | thread=多线程 | process=多进程')
    parser.add_argument('--task-ram-mb', type=int, default=300,
                        help='自动 workers 时每任务内存估计 MB (默认 300; image2pipe 模式自动 ×2)')

    args = parser.parse_args()

    # ── 环境检查（跨平台: shutil.which 定位 ffmpeg/ffprobe）──
    if not args.skip_ffmpeg and shutil.which('ffmpeg') is None:
        print("[WARN] PATH 中未找到 ffmpeg，ffmpeg 解码任务将全部标记 [SKIP]")
    if shutil.which('ffprobe') is None:
        print("[WARN] PATH 中未找到 ffprobe，ffprobe 统计任务将全部标记 [SKIP]")

    # ── 系统资源探测与并行参数 ──
    cpu_count = detect_cpu_count()
    cpu_host = os.cpu_count() or 1
    ram_total, ram_avail = detect_ram_gb()
    ram_desc = f"总计 {ram_total:.1f}GB / 可用 {ram_avail:.1f}GB" if ram_total > 0 else "未知"
    # auto: 任务为子进程型（ffprobe/ffmpeg 自身独立进程解码），Python 侧仅等待，thread 开销最小
    parallel_mode = 'thread' if args.parallel == 'auto' else args.parallel
    task_ram = args.task_ram_mb * (2 if args.image2pipe else 1)
    workers = args.workers if args.workers and args.workers > 0 else compute_auto_workers(task_ram)
    cpu_note = f" (宿主机 {cpu_host})" if cpu_count != cpu_host else ""
    print(f"🖥️  系统资源: CPU×{cpu_count}{cpu_note} | RAM {ram_desc}")
    print(f"⚙️  并行配置: workers={workers} | 模式={parallel_mode} | 单任务内存估计={task_ram}MB | 内存预留=10%")

    # ── 解析全部日志：步骤顺序 / 分段位置 / 最终输出（CPU 轻量，串行即可）──
    jobs = []
    single = len(args.log_files) == 1
    if args.suffix and not single:
        print("[WARN] 多日志模式下 --suffix 无效，各日志自动推导后缀")
    for log_arg in args.log_files:
        log_path = Path(log_arg).resolve()
        if not log_path.exists():
            print(f"[ERROR] 日志文件不存在: {log_path}，已跳过")
            continue

        log_dir = log_path.parent
        output_dir = Path(args.output_dir).resolve() if args.output_dir else log_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        suffix = args.suffix if (args.suffix and single) else (re.sub(r'^output', '', log_path.stem) or 'pipeline')

        print(f"\n📖 解析日志: {log_path}")
        info = parse_pipeline_log(str(log_path))
        if args.output_video and single:
            info['output_video'] = args.output_video

        resolve_step_dirs(info, log_dir, args.segments_dir)

        print(f"   处理模式: {info['mode'] or 'N/A'} | 分段数: {info['num_segments']}")
        for step in info['steps']:
            print(f"   Step {step['order']}: {step['name']} [{step['type']}] "
                  f"factor={step['factor']} 段数={len(step['segments'])}")
            print(f"      输入: {step['input_dir']}")
            print(f"      输出: {step['output_dir']}")
        print(f"   最终输出: {info.get('output_video', 'N/A')}")

        if not info['steps']:
            print("[WARN] 未能从日志中解析出任何处理步骤，已跳过")
            continue

        jobs.append({'log_path': log_path, 'output_dir': output_dir,
                     'suffix': suffix, 'info': info})

    if not jobs:
        print("[ERROR] 没有可分析的有效日志")
        sys.exit(1)

    # ── 两阶段智能调度 ──
    #    阶段 1: 快速扫描 ffprobe/packets（仅读包头），获取各分段帧数（cost）
    #    阶段 2: 根据帧数排序后执行 ffprobe/frames + ffmpeg 解码（重任务）
    prefetched = {}

    # 预先计算共享排序数据：step 目录优先级 + 最终视频路径
    # typ: output=1(高), input=0(低) → reverse=True 时 output 排前
    step_dir_priority = {}
    for job in jobs:
        for step in job['info']['steps']:
            order = step['order']
            for key, typ in [('output_dir', 1), ('input_dir', 0)]:
                d = step.get(key, '')
                if d:
                    step_dir_priority[str(Path(d).resolve())] = (order, typ)
    final_video_paths = set()
    for job in jobs:
        fv = job['info'].get('output_video', '')
        if fv:
            final_video_paths.add(str(fv))

    # === 阶段 1：快速扫描 (ffprobe packets，始终执行，--skip-ffprobe-frames 不影响) ===
    fast_tasks = []
    for job in jobs:
        fast_tasks.extend(collect_ffprobe_packets_tasks(job['info']))
    if fast_tasks:
        uniq = {}
        for t in fast_tasks:
            uniq.setdefault(t['key'], t)
        fast_list = list(uniq.values())
        # 排序：最终视频 → Step N 输出 → Step N 输入 → ...
        fv_tasks = [t for t in fast_list if t['path'] in final_video_paths]
        other_tasks = [t for t in fast_list if t['path'] not in final_video_paths]
        other_tasks.sort(key=lambda t: _path_step_priority(t['path'], step_dir_priority),
                         reverse=True)
        fast_list = fv_tasks + other_tasks
        print(f"\n📊 阶段 1/2: 快速扫描 ffprobe/packets ({len(fast_list)} 个文件, workers={workers}) ...")
        t_phase1 = time.time()
        fast_results = run_tasks_parallel(fast_list, workers, parallel_mode,
                                          phase_label='阶段 1/2: 快速扫描')
        prefetched.update(fast_results)
        frame_costs = _extract_segment_frame_counts(prefetched)
        # 最终输出视频 cost = 最后步骤输出分段帧数之和（保证排序时排第一位）
        for job in jobs:
            fv = job['info'].get('output_video', '')
            if fv and fv not in frame_costs:
                steps = job['info'].get('steps', [])
                if steps:
                    last_step = steps[-1]
                    out_dir = Path(last_step['output_dir']) if last_step['output_dir'] else None
                    fv_cost = 0
                    if out_dir:
                        for seg_data in last_step['segments']:
                            sf = _find_segment_file(out_dir,
                                                    last_step['output_prefix'] or 'segment',
                                                    seg_data['index'])
                            fv_cost += frame_costs.get(str(sf), 0)
                    if fv_cost > 0:
                        frame_costs[fv] = fv_cost
        if frame_costs:
            total_frames = sum(frame_costs.values())
            fv_has = sum(1 for j in jobs if j['info'].get('output_video', '') in frame_costs)
            fv_note = f" (含 {fv_has} 个最终视频估算)" if fv_has else ""
            print(f"   已获取 {len(frame_costs)} 个文件的帧数，总计 {total_frames} 帧{fv_note}")
        else:
            print("   未获取到帧数（packets 结果不可解析），第二阶段将按任务数调度")
        print(f"   ✅ 阶段 1 完成，用时 {time.time() - t_phase1:.1f}s")
    else:
        frame_costs = {}

    # === 阶段 2：重任务 (ffprobe frames + ffmpeg 解码) ===
    heavy_tasks = []
    if not args.skip_ffprobe_frames:
        for job in jobs:
            heavy_tasks.extend(collect_ffprobe_frames_tasks(job['info']))
    if not args.skip_ffmpeg:
        for job in jobs:
            heavy_tasks.extend(collect_ffmpeg_tasks(job['info'], use_image2pipe=args.image2pipe))
    if heavy_tasks:
        uniq = {}
        for t in heavy_tasks:
            uniq.setdefault(t['key'], t)
        heavy_list = list(uniq.values())
        # 排序：最终视频 → Step N 输出 → Step N 输入 → ...
        #   同组内 -f null: ffprobe 先, ffmpeg 后; -f image2pipe: ffmpeg 先, ffprobe 后
        heavy_list = _enrich_and_sort_heavy_tasks(heavy_list, frame_costs,
                                                   step_dir_priority=step_dir_priority,
                                                   image2pipe=args.image2pipe,
                                                   final_video_paths=final_video_paths)
        total_work = sum(t.get('cost', 0) for t in heavy_list)
        n_desc = f"{len(heavy_list)} 任务, 总 {total_work} 帧" if total_work else f"{len(heavy_list)} 任务"
        print(f"\n🔬 阶段 2/2: 深度检测 ({n_desc}, workers={workers}) ...")
        t_phase2 = time.time()
        heavy_results = run_tasks_parallel(heavy_list, workers, parallel_mode,
                                           total_work=total_work,
                                           phase_label='阶段 2/2: 深度检测')
        prefetched.update(heavy_results)
        print(f"   ✅ 阶段 2 完成，用时 {time.time() - t_phase2:.1f}s")
        print(f"✅ 两阶段并行完成: {len(prefetched)} 个结果，"
              f"用时 {time.time() - t_start:.1f}s")

    # ── 逐日志生成分析文件（I/O 轻量串行，输出格式与 v1 完全一致）──
    for job in jobs:
        info, output_dir, suffix = job['info'], job['output_dir'], job['suffix']
        print(f"\n📦 生成分析文件: {job['log_path'].name} (suffix={suffix})")

        # ── ffprobe 帧数统计 ──
        ffprobe_data = {}
        # ffprobe 输出：packets 始终生成，frames 受 --skip-ffprobe-frames 控制
        ffprob_path = output_dir / f"ffprob_output{suffix}.txt"
        print(f"🔍 生成 ffprobe 输出: {ffprob_path}")
        try:
            ffprobe_data = generate_ffprob_output(info, str(ffprob_path),
                                                  prefetched=prefetched,
                                                  skip_frames=args.skip_ffprobe_frames)
        except Exception as e:
            print(f"[WARN] ffprobe 生成异常: {e}，继续后续步骤")

        # ── ffmpeg 解码 ──
        ffmpeg_results = {}
        if not args.skip_ffmpeg:
            ffmpeg_path = output_dir / f"ffmpeg_output{suffix}.txt"
            mode_label = "image2pipe" if args.image2pipe else "null"
            print(f"🎬 生成 ffmpeg 输出 ({mode_label} 解码模式): {ffmpeg_path}")
            try:
                ffmpeg_results = generate_ffmpeg_output(info, str(ffmpeg_path),
                                                        use_image2pipe=args.image2pipe,
                                                        prefetched=prefetched)
            except Exception as e:
                print(f"[WARN] ffmpeg 生成异常: {e}，继续后续步骤")

        # ── 帧数链计算 ──
        chain_rows = build_chain(info, ffprobe_data, ffmpeg_results)

        # ── xlsx ──
        if not args.skip_xlsx:
            xlsx_path = output_dir / f"帧数差值统计汇总_output{suffix}.xlsx"
            print(f"📊 生成 xlsx: {xlsx_path}")
            try:
                generate_xlsx(info, ffprobe_data, ffmpeg_results, chain_rows, str(xlsx_path))
            except Exception as e:
                print(f"[ERROR] xlsx 生成失败: {e}")
                import traceback
                traceback.print_exc()

    print("\n" + "=" * 60)
    print(f"  ✅ 全流程检测分析文件生成完毕！共 {len(jobs)} 个日志，总用时 {time.time() - t_start:.1f}s")
    print("=" * 60)


if __name__ == '__main__':
    main()
