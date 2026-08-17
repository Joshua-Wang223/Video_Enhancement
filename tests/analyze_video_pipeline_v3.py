#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
视频增强流水线 全流程检测分析脚本 v3（并行计算版 + GPU 硬件加速）
===============================================================
基于 analyze_video_pipeline.py 升级：

  v3 新增 (GPU 硬件加速):
    1. --hwaccel auto|cuda|off: 利用 NVDEC 加速 ffmpeg 解码任务
       - 统一使用 -vsync 0 -vf showinfo -f null - 检测原始时间戳异常
       - GPU 模式: NVDEC 硬解 + showinfo (自动 hwdownload, 2-5× 提速)
       - ffprobe/frames 统计: ffprobe 不支持 -hwaccel，改用等效 ffmpeg NVDEC null 管线 (2-5× 提速)
    2. --gpu-workers N: GPU 任务并发闸门 (默认 4，防 NVDEC 会话耗尽)
       - 兼顾 GPU 速度与软件解码器诊断精确度 (PPS/SPS/花屏错误串)
       - xlsx "ffmpeg错误详情" 新增"解码器"列，标记 GPU/CPU 来源
    3. 启动自检: ffmpeg -hwaccels + 1 帧试编解码，失败自动回退纯 CPU

  v2 新增能力：
    1. 自动探测系统资源（容器感知）：CPU 逻辑核数 / RAM 总量与可用量。
       CPU 优先读 cgroup v1/v2 quota；内存用 cgroup limit 修正 psutil/sysconf 结果；
       Windows 回退 GlobalMemoryStatusEx；Linux 回退 /proc/meminfo/sysconf
    2. 并行执行引擎：ffprobe 帧数统计 + ffmpeg 解码全部任务化，
       支持 多线程(thread) / 多进程(process) 两种并行模式
       - 多个小文件（多日志）：跨日志汇总任务到统一并行池，打满系统资源
       - 单个大文件（单日志多分段）：分段级任务并行
    3. workers 自动计算: min(CPU核数, RAM总计×90%/单任务内存估计)，可用 --workers 覆盖
    4. 输出文件格式与 v1 完全一致（结果按原始分段顺序落盘，与并行完成顺序无关）

  v1 原有功能（保持不变）：
    根据 log_file 自动解析：
      1. 插帧/超分步骤顺序（"阶段 N: Step x/y — ..." 标题，兼容单步骤旧日志）
      2. 初始分段目录（由第一个步骤的临时输出分段位置反推，如 ../segments）
      3. 中间分段位置（前面步骤的分段输出即为后面步骤的分段输入）
      4. 最终输出视频文件（合并行 / 📤 输出行）

生成以下检测分析文件：
  1. ffmpeg_output{suffix}.txt           — ffmpeg 逐段解码日志（每个步骤的输出分段）
  2. ffprob_output{suffix}.txt           — ffprobe 逐段帧数统计（初始分段 + 各步骤输出 + 最终视频）
  3. 帧数差值统计汇总_output{suffix}.xlsx — 帧数差值汇总 Excel
       Sheet1 全流程汇总  — 每段跨步骤帧数链（输入→插帧→超分→最终）
       Sheet2..N 各步骤明细 — 帧数差值/异常检测/性能统计
       末 Sheet ffmpeg错误详情 (含 GPU/CPU 解码器标记)

用法:
    python analyze_video_pipeline_v3.py <log_file> [log_file2 ...] [选项]

示例:
    python analyze_video_pipeline_v3.py temp/output28-9.txt
    python analyze_video_pipeline_v3.py temp/output28-9.txt temp/output28-10.txt --workers 8
    python analyze_video_pipeline_v3.py temp/output28-9.txt --parallel process --task-ram-mb 512
    python analyze_video_pipeline_v3.py temp/output28-9.txt --segments-dir ./segments --skip-ffmpeg
    python analyze_video_pipeline_v3.py temp/output28-9.txt --hwaccel cuda --gpu-workers 2
    python analyze_video_pipeline_v3.py temp/output28-9.txt --hwaccel off  # 纯 CPU 取证模式
    python analyze_video_pipeline_v3.py temp/output28-9.txt --skip-ffprobe-frames  # 跳过慢速帧统计
"""

import argparse
import math
import os
import re
import shutil
import subprocess
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl 未安装，将不会生成 xlsx 文件。安装: pip install openpyxl")


# ═══════════════════════════════════════════════
# 0. 模块级 GPU 基础设施
# ═══════════════════════════════════════════════

# [FIX-NVDEC-THREAD-CAP] NVDEC 解码线程钳位：
# ffmpeg 6.1.1 的 NVDEC 解码 surface 计算公式：
#   ulNumDecodeSurfaces = ref_frame_count + num_reorder_frames
#                         + 2(deinterlace) + thread_count + 3(基础工作 surface)
# 32 是驱动硬上限（cudaVideoDecoder 拒绝），ffmpeg 6.1.1 无 FFMIN(pool,32) 钳位。
# 实测：-threads 8 → 32 surfaces 成功；-threads 9 → 33 surfaces 被驱动拒绝。
_MAX_DECODE_THREADS = 8
_DECODE_THREAD_HINT_SHOWN = False

def _clamp_decode_threads(requested=None) -> int:
    """钳位 NVDEC 解码线程数（驱动 32-surface 硬上限），超限时打印一次提示。"""
    global _DECODE_THREAD_HINT_SHOWN
    n = int(requested) if requested else (os.cpu_count() or 4)
    if n > _MAX_DECODE_THREADS:
        if not _DECODE_THREAD_HINT_SHOWN:
            print(f"[decode] threads={n} exceeds max {_MAX_DECODE_THREADS}, "
                  f"clamped (NVDEC 32-surface limit)", flush=True)
            _DECODE_THREAD_HINT_SHOWN = True
        n = _MAX_DECODE_THREADS
    return n

# GPU 解码器错误关键词（与软件解码器错误字符串不同，需显式加入）
# 注意: 禁用裸设备名 'nvenc'/'nvdec'/'cuvid' —— 会误匹配 ffmpeg 信息行中的
# 编码器/解码器名称 (如 "h264 (h264_nvenc)" stream mapping 行、
# "encoder : LavcXX h264_nvenc" 元数据行)，导致每个文件恒定 errors=2 假阳性。
# 必须使用只在真实错误中出现、不会出现在设备名提及中的具体短语。
_GPU_DECODE_ERROR_KW = [
    'failed to decode',              # NVDEC: "hardware accelerator failed to decode picture"
    'hardware accelerator',          # Cuvid/NVDEC 硬件解码器报错前缀
    'cuda device',                   # CUDA 设备错误
    'cannot init cuvid',             # NVDEC 初始化失败
    'cannot load libnvcuvid',        # NVDEC 运行时库加载失败
    'no nvenc capable devices',      # NVENC 设备不可用 ("No NVENC capable devices found")
    'openencodesessionex failed',    # NVENC 会话创建失败 (OOM/会话耗尽)
    'encodepicture failed',          # NVENC 编码单帧失败
    'nvenc api version',             # NVENC API/驱动版本不兼容
    'vdpau',                         # VDPAU backend error
]

# ── ffmpeg 常规信息行排除（FIX-ERR-KW-FALSE-POSITIVE）──────────────────────
# 根因: 裸词 'error'/'decoder'/'missing' 等会误匹配 ffmpeg 正常信息行
# (如 "Stream mapping: ... h264 (h264_nvenc)"、"encoder : LavcXX h264_nvenc"
# 元数据行、NVENC/CUDA 正常初始化 context 行)，导致每个文件恒定误报 errors。
# 修复: 命中信息行前缀且不含致命关键词 → 跳过错误计数。
_FFMPEG_INFO_PREFIXES = (
    'stream mapping', 'metadata', 'input #', 'output #', 'duration:',
    'stream #', 'press [q]', 'encoder :', 'handler_name', 'major_brand',
    'minor_version', 'compatible_brands', 'configuration:', 'libav',
    'libsw', 'libpostproc', 'built with', 'side data',
)
# 正常初始化 context 前缀（NVENC/CUDA/滤镜图等正常日志行）
_FFMPEG_INFO_CTX_PREFIXES = (
    '[h264_nvenc @', '[hevc_nvenc @', '[av1_nvenc @', '[avhwdevicecontext @',
    '[graph', '[in#', '[out#', '[auto_', '[aformat', '[vist#', '[dec:', '[enc:',
    '[buffer @', '[buffersink @', '[null @', '[swscaler @',
)
# 致命关键词: 即使出现在信息行也必须计为真实错误（信息行排除不屏蔽）
_FATAL_ERR_KW = (
    'illegal', 'non-existing pps', 'no frame', 'decode_slice_header error',
    'missing reference picture', 'failed to decode', 'error while decoding',
    'concealing',
)


def _is_ffmpeg_info_line(lower: str) -> bool:
    """判定 stderr 行是否为 ffmpeg 常规信息行（无害，跳过错误计数）。

    命中信息行前缀（或正常初始化 context 前缀）且不含任一致命关键词时
    返回 True。lower 必须为已小写化的整行文本。
    """
    if any(kw in lower for kw in _FATAL_ERR_KW):
        return False
    return (lower.startswith(_FFMPEG_INFO_PREFIXES) or
            lower.startswith(_FFMPEG_INFO_CTX_PREFIXES))

# GPU 并发信号量 (thread 模式下生效，限制同时跑 GPU 的任务数)
_GPU_SEMAPHORE = None         # 在 main() 中按 --gpu-workers 初始化


# ═══════════════════════════════════════════════
# 1. 日志解析（自动识别步骤顺序与分段位置）
# ═══════════════════════════════════════════════

def _split_prefix(basename: str):
    """'interpolated_segment_003.mp4' -> ('interpolated_segment', 3)"""
    m = re.match(r'(.+?)_(\d{3})\.mp4$', basename)
    if m:
        return m.group(1), int(m.group(2))
    return None, None


_INTERPOLATE_KW_RE = re.compile(r'插帧|IFRNet|ifrnet')
_UPSCALE_KW_RE = re.compile(r'超分|ESRGAN|esrgan|Real-ESRGAN')


def _detect_step_type(name: str, text: str, skip_upscale: bool = False,
                      skip_interpolate: bool = False) -> str:
    """Return interpolate or upscale, honoring explicit skip flags."""
    name = name or ''
    text = text or ''
    name_interp = bool(_INTERPOLATE_KW_RE.search(name))
    name_upscale = bool(_UPSCALE_KW_RE.search(name))
    text_interp = bool(_INTERPOLATE_KW_RE.search(text[:4000]))
    text_upscale = bool(_UPSCALE_KW_RE.search(text[:4000]))

    # Step name is authoritative when unambiguous; skip flags resolve ambiguity.
    if skip_upscale and name_interp:
        return 'interpolate'
    if skip_interpolate and name_upscale:
        return 'upscale'
    if name_interp and not name_upscale:
        return 'interpolate'
    if name_upscale and not name_interp:
        return 'upscale'

    has_interp = name_interp or text_interp
    has_upscale = name_upscale or text_upscale

    if skip_upscale and has_interp:
        return 'interpolate'
    if skip_interpolate and has_upscale:
        return 'upscale'
    if has_upscale and not has_interp:
        return 'upscale'
    if has_interp and not has_upscale:
        return 'interpolate'

    # Preserve old fallback: prefer upscale when both appear and no skip flag.
    if has_upscale:
        return 'upscale'
    if has_interp:
        return 'interpolate'
    return 'unknown'


def _compose_mode_display(mode: str, skip_upscale: bool = False,
                          skip_interpolate: bool = False) -> str:
    """组合基础处理模式与跳过标志为完整模式字符串。

    例如: base='upscale_then_interpolate', skip_upscale=True
        -> 'upscale_then_interpolate+skip_upscale'
    若基础模式缺失则仅返回跳过标志；若模式中已含对应标志则不重复追加。
    """
    mode = (mode or '').strip()
    flags = []
    if skip_upscale and 'skip_upscale' not in mode:
        flags.append('skip_upscale')
    if skip_interpolate and 'skip_interpolate' not in mode:
        flags.append('skip_interpolate')
    if not flags:
        return mode
    return f"{mode}+{'+'.join(flags)}" if mode else '+'.join(flags)


def parse_pipeline_log(log_path: str) -> dict:
    """解析全流程流水线日志，自动识别步骤顺序、分段位置与最终输出"""
    with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    info = {
        'mode': '',                 # 完整处理模式（基础模式+跳过标志，如 upscale_then_interpolate+skip_upscale）
        'num_segments': 0,
        'segment_duration': 300,
        'source_video': '',
        'output_video': '',
        'total_output_size': '',
        'total_elapsed': '',
        'steps': [],                # 按顺序的步骤列表
    }

    # Skip flags, derived from the command line and summary text.
    skip_upscale = bool(re.search('--skip-upscale', content)) or bool(re.search('跳过超分步骤|仅插帧', content))
    skip_interpolate = bool(re.search('--skip-interpolate', content)) or bool(re.search('跳过插帧步骤|仅超分', content))
    info['skip_upscale'] = skip_upscale
    info['skip_interpolate'] = skip_interpolate
    # ── 全局配置摘要 ──
    m = re.search(r'处理模式\s*[:：]\s*(\S+)', content)
    if m:
        info['mode'] = m.group(1).strip()
    # 将跳过标志（skip_upscale / skip_interpolate）并入处理模式，
    # 例如: upscale_then_interpolate+skip_upscale
    info['mode'] = _compose_mode_display(info['mode'],
                                         skip_upscale=skip_upscale,
                                         skip_interpolate=skip_interpolate)
    m = re.search(r'分段时长\s*[:：]\s*(\d+)', content)
    if m:
        info['segment_duration'] = int(m.group(1))
    m = re.search(r'^\s*输入\s*[:：]\s*(.+\.(?:mp4|avi|mkv))\s*$', content, re.M)
    if m:
        info['source_video'] = m.group(1).strip()

    # 最终输出视频：优先合并行，其次 📤 输出行，最后"最终输出"
    m = re.search(r'🔗\s*合并\s*\d+\s*个最终分段\s*→\s*(.+\.(?:mp4|avi|mkv))', content)
    if not m:
        m = re.search(r'(?:📤\s*输出(?:文件)?|最终输出)\s*[:：]\s*(.+\.(?:mp4|avi|mkv))', content)
    if m:
        info['output_video'] = m.group(1).strip()

    m = re.search(r'📦\s*文件大小\s*[:：]\s*([\d.]+\s*[KMGT]B)', content)
    if m:
        info['total_output_size'] = m.group(1)
    m = re.search(r'总用时\s*[:：]\s*([\d:.]+)', content)
    if m:
        info['total_elapsed'] = m.group(1)

    # Split processing stages. Supports both:
    #   * Stage 1: Step 1/2 - Real-ESRGAN ...
    #   * Stage 2: IFRNet interpolation (skip-upscale mode)
    stage_pattern = re.compile(r'阶段\s*(\d+)\s*[：:]\s*(.+)')
    stage_matches = list(stage_pattern.finditer(content))

    step_blocks = []  # (name, text)
    if stage_matches:
        for si, sm in enumerate(stage_matches):
            end_pos = stage_matches[si + 1].start() if si + 1 < len(stage_matches) else len(content)
            raw_name = sm.group(2).strip()
            # Skip non-processing stages such as stage 0 and stage 3.
            if not re.search(r'插帧|超分|IFRNet|ESRGAN', raw_name):
                continue
            # Strip the "Step X/Y - " prefix, if present.
            step_m = re.search(r'Step\s*\d+\s*/\s*\d+\s*—\s*(.+)', raw_name)
            name = step_m.group(1).strip() if step_m else raw_name
            step_blocks.append((name, content[sm.start():end_pos]))
    else:
        # Legacy single-step logs: classify the whole log as one step.
        whole_type = _detect_step_type('', content,
                                       skip_upscale=skip_upscale,
                                       skip_interpolate=skip_interpolate)
        step_blocks.append(('IFRNet 插帧' if whole_type == 'interpolate' else 'Real-ESRGAN 超分', content))

    # ── 逐步骤解析分段 ──
    segment_pattern = re.compile(
        r'片段\s*(\d+)/(\d+):\s*(\S*segment_(\d+)\.mp4)'
    )
    frame_pattern = re.compile(r'原始帧:\s*(\d+)\s*→\s*输出帧:\s*(\d+)')
    time_pattern = re.compile(r'输出时长\s*([\d:.]+)\s*\|\s*耗时\s*([\d:.]+)')
    # 输出文件：兼容 "   输出: path (xx MB)" 与 "✅ 输出: path (xx MB)"
    output_file_pattern = re.compile(
        r'(?:✅\s*)?输出:\s*(.+?segment_\d+\.mp4)\s*\(([\d.]+\s*[KMGT]B)\)'
    )
    perf_pattern = re.compile(
        r'\[性能统计\]\s*总时间\s*[:：]\s*([\d.]+)秒\s*\|\s*平均\s*[:：]\s*([\d.]+)ms\s*\|\s*FPS\s*[:：]\s*([\d.]+)'
    )
    gfpgan_pattern = re.compile(r'\[性能统计\]\s*GFPGAN\s*模式\s*[:：]\s*(\S+)')
    model_load_pattern = re.compile(r'总耗时（含模型加载）\s*[:：]\s*([\d分秒]+)')
    estimate_pattern = re.compile(r'已用时\s*[:：]\s*([\d:.]+)\s*,\s*预计剩余\s*[:：]\s*([\d:.]+)')
    # IFRNet GPU 性能行: [GPU0] 完成 | 原始帧=8974 → 输出帧=17947 | 12.6 原始帧/s
    ifrnet_perf_pattern = re.compile(
        r'\[GPU\d+\]\s*完成\s*\|\s*原始帧=(\d+)\s*→\s*输出帧=(\d+)\s*\|\s*([\d.]+)\s*原始帧/s'
    )

    for order, (step_name, step_text) in enumerate(step_blocks, 1):
        step_type = _detect_step_type(step_name, step_text,
                                      skip_upscale=skip_upscale,
                                      skip_interpolate=skip_interpolate)

        step = {
            'order': order,
            'name': step_name,
            'type': step_type,               # interpolate | upscale
            'factor': None,                  # 插帧倍数 或 超分倍数
            'gfpgan_mode': 'unknown',
            'input_prefix': 'segment',
            'output_prefix': None,
            'input_dir': '',                 # 待 main() 推断
            'output_dir': '',                # 由首个输出文件反推
            'segments': [],
        }

        if step_type == 'interpolate':
            m = re.search(r'插帧倍数\s*[:：]\s*(\d+)', step_text)
        else:
            m = re.search(r'超分倍数\s*[:：]\s*(\d+)', step_text)
        if m:
            step['factor'] = int(m.group(1))

        gm = gfpgan_pattern.search(step_text)
        if gm:
            step['gfpgan_mode'] = gm.group(1).strip()

        seg_starts = [(m.start(), int(m.group(1)), m.group(3), int(m.group(4)))
                      for m in segment_pattern.finditer(step_text)]

        for idx, (start_pos, seg_num, in_name, seg_id) in enumerate(seg_starts):
            end_pos = seg_starts[idx + 1][0] if idx + 1 < len(seg_starts) else len(step_text)
            seg_text = step_text[start_pos:end_pos]

            seg = {
                'index': seg_id,
                'number': seg_num,
                'input_name': in_name,
                'raw_frames': None,          # 日志原始帧（= 本步骤输入帧）
                'output_frames': None,       # 日志提示输出帧
                'output_duration': '',
                'elapsed': '',
                'output_file': '',
                'output_size': '',
                'model_load_time': '',
                'perf_total_sec': None,
                'perf_avg_ms': None,
                'perf_fps': None,
                'ifrnet_raw_fps': None,       # IFRNet GPU 报告的原始帧/s
                'elapsed_estimate': '',
                'remaining_estimate': '',
            }

            # 输入前缀（第一段确定一次）
            if idx == 0:
                pfx, _ = _split_prefix(in_name)
                if pfx:
                    step['input_prefix'] = pfx

            fm = frame_pattern.search(seg_text)
            if fm:
                seg['raw_frames'] = int(fm.group(1))
                seg['output_frames'] = int(fm.group(2))

            tm = time_pattern.search(seg_text)
            if tm:
                seg['output_duration'] = tm.group(1)
                seg['elapsed'] = tm.group(2)

            ofm = output_file_pattern.search(seg_text)
            if ofm:
                seg['output_file'] = ofm.group(1).strip()
                seg['output_size'] = ofm.group(2).strip()
                if not step['output_dir']:
                    step['output_dir'] = str(Path(seg['output_file']).parent)
                    pfx, _ = _split_prefix(Path(seg['output_file']).name)
                    if pfx:
                        step['output_prefix'] = pfx

            pm = perf_pattern.search(seg_text)
            if pm:
                seg['perf_total_sec'] = float(pm.group(1))
                seg['perf_avg_ms'] = float(pm.group(2))
                seg['perf_fps'] = float(pm.group(3))

            # IFRNet GPU 性能行: [GPU0] 完成 | 原始帧=8974 → 输出帧=17947 | 12.6 原始帧/s
            ipm = ifrnet_perf_pattern.search(seg_text)
            if ipm:
                seg['ifrnet_raw_fps'] = float(ipm.group(3))
                # GPU 行报告的帧数与 ✅ 插帧完成！行可能重复，兜底补齐
                if seg['raw_frames'] is None:
                    seg['raw_frames'] = int(ipm.group(1))
                if seg['output_frames'] is None:
                    seg['output_frames'] = int(ipm.group(2))

            mlm = model_load_pattern.search(seg_text)
            if mlm:
                seg['model_load_time'] = mlm.group(1)

            em = estimate_pattern.search(seg_text)
            if em:
                seg['elapsed_estimate'] = em.group(1)
                seg['remaining_estimate'] = em.group(2)

            step['segments'].append(seg)

        info['steps'].append(step)

    # 分段总数：取各步骤最大分段数
    if info['steps']:
        info['num_segments'] = max(len(s['segments']) for s in info['steps'])
    if info['num_segments'] == 0:
        m = re.search(r'✅\s*共\s*(\d+)\s*个片段', content)
        if m:
            info['num_segments'] = int(m.group(1))

    return info


def resolve_step_dirs(info: dict, log_dir: Path, segments_dir_arg: str = None):
    """
    推断每个步骤的输入/输出目录：
      - 第一个步骤的输入 = 初始分段目录（由其输出目录反推 ../segments，或 CLI 指定）
      - 后续步骤的输入 = 前一步骤的输出目录
    """
    prev_output_dir = None
    for si, step in enumerate(info['steps']):
        if si == 0:
            cand = []
            if segments_dir_arg:
                cand.append(Path(segments_dir_arg))
            if step['output_dir']:
                cand.append(Path(step['output_dir']).parent / 'segments')  # ../segments 反推
            cand.append(log_dir / 'segments')
            step['input_dir'] = ''
            for c in cand:
                if c.exists() and list(c.glob(f"{step['input_prefix']}_*.mp4")):
                    step['input_dir'] = str(c)
                    break
            if not step['input_dir']:
                step['input_dir'] = str(cand[0])
        else:
            step['input_dir'] = prev_output_dir
        prev_output_dir = step['output_dir']


# ═══════════════════════════════════════════════
# 2. ffmpeg 解码 + 错误检测
# ═══════════════════════════════════════════════

def _build_ffmpeg_decode_err_kw() -> list:
    """合并软件解码器 + GPU 解码器错误关键词列表（已细化去除裸宽泛词）。"""
    base = [
        'illegal', 'corrupt',
        'non-existing pps', 'no frame',
        'more than 1000 frames duplicated',
        # 带上下文的真实错误短语（替代裸 'error'/'decoder'/'missing' 等）
        'error while decoding', 'failed to decode', 'missing reference picture',
        'decode_slice_header error', 'concealing',
    ]
    return base + _GPU_DECODE_ERROR_KW


# NVDEC hwaccel 初始化失败特征（ffmpeg 已内部回退软件解码，帧数仍有效，非真实解码错误）
_HWACCEL_INIT_FAILURE_KW = [
    'failed setup for format cuda',
    'hwaccel initialisation returned error',
    'cuvidcreatedecoder',
    'decode surfaces',  # "Using more than 32 (N) decode surfaces" 超 surface 上限告警
]


def _is_hwaccel_init_failure(lower_line: str) -> bool:
    """判定某行是否为 NVDEC hwaccel 初始化失败告警（而非真实解码错误）。

    这类失败是运行时环境污染（显存被占/驱动会话冲突）导致的瞬态问题，
    ffmpeg 会静默回退到软件解码，因此 frame 计数仍然正确，不应当作解码错误
    触发 CPU 复检或污染 decoder 标记。
    """
    return any(kw in lower_line for kw in _HWACCEL_INIT_FAILURE_KW)


def _ffmpeg_decode(video_path: str, timeout: int = 900, hwaccel: str = 'off',
                    pts_drop_threshold: float = 2.0) -> dict:
    """ffmpeg 解码：-vsync 0 -vf showinfo -f null -，检测原始 dup/drop 异常与时间戳跳跃。

    核心原理：
      - -vsync 0 (passthrough): 保持原始时间戳，ffmpeg 不主动生成 dup/drop
      - -vf showinfo: 输出每帧 pts，用于检测源视频本身的时间戳异常
      - 最终状态行的 dup=/drop= 反映的是 vsync 策略产生的数量，
        在 -vsync 0 下应为 0；非 0 说明时间戳已被修改（需警惕）
      - pts_dups / pts_drops: 基于 showinfo 逐帧解析的原始异常（真正的源视频问题）
      - pts_drop_frames_est: 估计缺失帧总数（pts 跳跃 / 平均间隔）

    hwaccel='cuda': NVDEC 硬解 + showinfo (自动 hwdownload)
    hwaccel='off':  纯 CPU 软解 + showinfo

    返回结构:
        frame_count: 总解码帧数
        errors:      解码错误列表
        warnings:    硬件回退警告
        dup_count:   vsync 报告的 dup 数（-vsync 0 下应为 0）
        drop_count:  vsync 报告的 drop 数（-vsync 0 下应为 0）
        pts_dups:    检测到的时间戳重复帧数（原始异常：连续帧 pts 相同）
        pts_drops:   检测到的时间戳跳跃次数（原始异常：pts 差值异常大）
        pts_drop_frames_est: 估计缺失帧总数
        pts_anomalies: 前若干条异常明细 [(n, pts, pts_time, type, note)]
        decoder:     'gpu' | 'gpu(retry)' | 'gpu→cpu' | 'cpu'
        exit_code:   进程返回码
    """
    result = {
        'frame_count': None,
        'errors': [],
        'warnings': [],
        'dup_count': None,
        'drop_count': None,
        'pts_dups': 0,
        'pts_drops': 0,
        'pts_drop_frames_est': 0,
        'pts_backward': 0,
        'pts_anomalies': [],
        'pts_last': None,
        'pts_interval': None,
        'decoder': 'cpu',
        'exit_code': None,
    }
    gpu_attempted = False
    cmd = None

    if hwaccel == 'cuda':
        gpu_attempted = True
        result['decoder'] = 'gpu'
        cmd = [
            "ffmpeg", "-hide_banner",
            "-hwaccel", "cuda",
            "-threads", str(_clamp_decode_threads()),  # [FIX-NVDEC-THREAD-CAP] 解码线程钳位 ≤8
            "-i", str(video_path),
            "-an", "-vsync", "0",
            "-vf", "showinfo",
            "-f", "null", "-"
        ]
    else:
        cmd = [
            "ffmpeg", "-hide_banner",
            "-threads", str(_clamp_decode_threads()),  # [FIX-NVDEC-THREAD-CAP] 解码线程钳位 ≤8
            "-i", str(video_path),
            "-an", "-vsync", "0",
            "-vf", "showinfo",
            "-f", "null", "-"
        ]

    def _parse_showinfo_line(line: str, r: dict, threshold: float):
        """解析 showinfo 滤镜输出的 pts 信息，检测时间戳异常。

        修复点：
          - 支持负 pts
          - pts 相同只计 dup，不污染 interval
          - pts 回退单独计数，不参与 interval 滑动平均
          - 首个有效正间隔才初始化 interval
          - 跳变阈值可配置，并累计估计缺失帧数
        """
        m = re.search(r'n:\s*(\d+)\s+pts:\s*(-?\d+)\s+pts_time:\s*([\d.]+)', line)
        if not m:
            # 兼容旧版 showinfo（无 pts_time）
            m = re.search(r'n:\s*(\d+)\s+pts:\s*(-?\d+)', line)
            if not m:
                return
        n = int(m.group(1))
        pts = int(m.group(2))
        pts_time = float(m.group(3)) if len(m.groups()) >= 3 else None

        last = r['pts_last']
        if last is not None:
            diff = pts - last
            if diff == 0:
                r['pts_dups'] += 1
                _record_anomaly(r, n, pts, pts_time, 'dup',
                                f"pts 与上一帧相同 ({pts})")
            elif diff < 0:
                r['pts_backward'] += 1
                _record_anomaly(r, n, pts, pts_time, 'backward',
                                f"pts 回退 {diff} ({last}->{pts})")
            else:
                iv = r['pts_interval']
                if iv is None:
                    r['pts_interval'] = diff
                elif diff > threshold * iv:
                    r['pts_drops'] += 1
                    est = max(1, int(round(diff / iv)) - 1)
                    r['pts_drop_frames_est'] += est
                    _record_anomaly(r, n, pts, pts_time, 'drop',
                                    f"pts 跳变 diff={diff}, interval≈{iv}, est_missing={est}")
                else:
                    r['pts_interval'] = (iv * 3 + diff) // 4
        r['pts_last'] = pts

    def _record_anomaly(r: dict, n: int, pts: int, pts_time, typ: str, note: str):
        if len(r['pts_anomalies']) < 20:
            r['pts_anomalies'].append((n, pts, pts_time, typ, note))

    def _parse_progress_segments(line: str, out: dict):
        """ffmpeg 进度以 \r 刷新，PIPE 按 \n 读取可能把多条进度粘成一行。
        这里把 \r 拆成独立段，取每个指标的最新值，避免 frame=/dup=/drop= 只命中首个。"""
        for seg in line.replace('\r', '\n').split('\n'):
            s = seg.strip()
            if not s:
                continue
            m = re.search(r'frame=\s*(\d+)', s)
            if m:
                out['frame_count'] = int(m.group(1))
            m_dup = re.search(r'dup=(\d+)', s)
            if m_dup:
                out['dup_count'] = int(m_dup.group(1))
            m_drop = re.search(r'drop=(\d+)', s)
            if m_drop:
                out['drop_count'] = int(m_drop.group(1))

    def _consume_stderr(proc, out: dict, include_warnings: bool):
        """统一消费 ffmpeg stderr：进度行 + showinfo + 错误关键词。"""
        for raw_line in proc.stderr:
            # 进度行可能含 \r，先整行解析所有进度段
            _parse_progress_segments(raw_line, out)
            # 再把 \r 替换成 \n，逐段处理 showinfo/错误（showinfo 以 \n 结尾不受影响）
            for segment in raw_line.replace('\r', '\n').split('\n'):
                stripped = segment.strip()
                if not stripped:
                    continue
                if 'Parsed_showinfo' in stripped:
                    _parse_showinfo_line(stripped, out, pts_drop_threshold)
                lower = stripped.lower()
                # 信息行排除: ffmpeg 常规信息行不计入错误 (FIX-ERR-KW-FALSE-POSITIVE)
                if (not _is_ffmpeg_info_line(lower) and
                        any(kw in lower for kw in _build_ffmpeg_decode_err_kw())):
                    out['errors'].append(stripped)
                if include_warnings and _is_hwaccel_init_failure(lower):
                    out['warnings'].append(stripped)

    try:
        proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE,
                                text=True, bufsize=1)
        try:
            _consume_stderr(proc, result, include_warnings=gpu_attempted)
            proc.wait(timeout=timeout)
            result['exit_code'] = proc.returncode
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait()
            result['errors'].append("[TIMEOUT] ffmpeg 解码超时")
            result['exit_code'] = -1
    except FileNotFoundError:
        result['errors'].append("[SKIP] ffmpeg 不可用")
        result['exit_code'] = -2

    # GPU 瞬态失败 → 先重试一次 GPU
    if gpu_attempted and result['exit_code'] == 0 and result['warnings']:
        # 修复 FIX-HWACCEL-INIT-FALSE-POSITIVE:
        # warnings 仅来自 hwaccel 初始化失败（如 cuvidCreateDecoder 返回
        # CUDA_ERROR_INVALID_VALUE —— H.264 level 超出 NVDEC 上限/驱动会话冲突
        # 等环境性问题）。ffmpeg 已静默回退软件解码，frame 计数仍然有效；
        # 纯 hwaccel 初始化失败不 sleep 重试、不标记 gpu→cpu（避免误报），
        # decoder 保持 'gpu'（GPU 已尝试，帧数可信）。仅当同批存在真实
        # GPU 解码错误（errors 非空）时才走重试 + 回退流程。
        if not result['errors']:
            return result
        time.sleep(2)
        retry_result = {
            'frame_count': None, 'errors': [], 'warnings': [],
            'dup_count': None, 'drop_count': None,
            'pts_dups': 0, 'pts_drops': 0, 'pts_drop_frames_est': 0,
            'pts_backward': 0, 'pts_anomalies': [],
            'pts_last': None, 'pts_interval': None,
            'decoder': 'gpu(retry)', 'exit_code': None,
        }
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE,
                                    text=True, bufsize=1)
            _consume_stderr(proc, retry_result, include_warnings=True)
            proc.wait(timeout=timeout)
            retry_result['exit_code'] = proc.returncode
        except Exception:
            retry_result['exit_code'] = -1

        if retry_result['exit_code'] == 0 and not retry_result['warnings']:
            result.update(retry_result)
            return result
        result['decoder'] = 'gpu→cpu'
        return result

    # GPU 硬失败 → 回退 CPU 重跑
    if gpu_attempted and result['exit_code'] is not None and result['exit_code'] != 0:
        cmd_cpu = [
            "ffmpeg", "-hide_banner",
            "-threads", str(_clamp_decode_threads()),  # [FIX-NVDEC-THREAD-CAP] 解码线程钳位 ≤8
            "-i", str(video_path),
            "-an", "-vsync", "0",
            "-vf", "showinfo",
            "-f", "null", "-"
        ]
        cpu_result = {
            'frame_count': None, 'errors': [], 'warnings': [],
            'dup_count': None, 'drop_count': None,
            'pts_dups': 0, 'pts_drops': 0, 'pts_drop_frames_est': 0,
            'pts_backward': 0, 'pts_anomalies': [],
            'pts_last': None, 'pts_interval': None,
            'decoder': 'gpu→cpu', 'exit_code': None,
        }
        cpu_result['errors'].append("[GPU-FALLBACK] GPU 解码失败，回退 CPU 软件解码重跑")
        try:
            proc2 = subprocess.Popen(cmd_cpu, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE,
                                     text=True, bufsize=1)
            _consume_stderr(proc2, cpu_result, include_warnings=False)
            proc2.wait(timeout=timeout)
            cpu_result['exit_code'] = proc2.returncode
        except subprocess.TimeoutExpired:
            proc2.kill()
            proc2.wait()
            cpu_result['errors'].append("[TIMEOUT] ffmpeg 解码超时(CPU回退)")
        except Exception:
            cpu_result['errors'].append("[SKIP] CPU 回退也失败")

        gpu_errs = [e for e in result['errors'] if not e.startswith('[GPU-FALLBACK]')]
        result = cpu_result
        for e in reversed(gpu_errs):
            result['errors'].insert(1, f"[GPU] {e}")

    return result


def _find_segment_file(directory: Path, prefix: str, seg_idx: int) -> Path:
    """优先精确前缀匹配，回退 *segment_XXX.mp4 通配"""
    exact = directory / f"{prefix}_{seg_idx:03d}.mp4"
    if exact.exists():
        return exact
    matches = list(directory.glob(f"*segment_{seg_idx:03d}.mp4"))
    if matches:
        return matches[0]
    return exact



def collect_ffmpeg_tasks(info: dict, timeout: int = 900,
                         hwaccel: str = 'off',
                         pts_drop_threshold: float = 2.0) -> list:
    """
    收集单个日志全部 ffmpeg 解码任务（v3 并行用）。
    任务 key: ('ffmpeg', path)，跨步骤/跨日志去重由 run 阶段统一处理。
    hwaccel: 'cuda' GPU 加速, 'off' 纯 CPU。
    """
    tasks = []
    seen = set()
    for step in info['steps']:
        out_dir = Path(step['output_dir']) if step['output_dir'] else Path('.')
        for seg in step['segments']:
            seg_file = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
            key = ('ffmpeg', str(seg_file))
            if key in seen or not seg_file.exists():
                continue
            seen.add(key)
            tasks.append({'kind': 'ffmpeg', 'key': key, 'path': str(seg_file),
                          'timeout': timeout, 'hwaccel': hwaccel,
                          'pts_drop_threshold': pts_drop_threshold})
    return tasks



def generate_ffmpeg_output(info: dict, output_path: str,
                           prefetched: dict = None,
                           hwaccel: str = 'off',
                           timeout: int = 900,
                           pts_drop_threshold: float = 2.0):
    """对每个步骤的输出分段执行 ffmpeg 解码。返回 {step_order: (frames, errors, dup_drop, decoders)}

    v3 修复: 统一使用 -vsync 0 -vf showinfo -f null - 检测原始时间戳异常。
      - dup_count/drop_count: ffmpeg 最终状态行报告（-vsync 0 下应为 0）
      - pts_dups/pts_drops: 基于 showinfo 逐帧 pts 解析的源视频原始异常
      - pts_drop_frames_est: 基于 pts 间隔估计的缺失帧总数
      - pts_anomalies: 前若干条异常明细

    prefetched 为并行预取结果 {('ffmpeg', path): result}，
    命中时直接读取（保持 v1 输出文件格式与分段顺序），缺项回退串行解码。
    """
    results = {}
    inline = prefetched is None
    with open(output_path, 'w', encoding='utf-8') as out:
        for step in info['steps']:
            order = step['order']
            out.write(f"\n{'=' * 60}\n")
            out.write(f" Step {order}/{len(info['steps'])} — {step['name']} ({step['type']}) [showinfo]\n")
            out.write(f" output_dir: {step['output_dir']}\n")
            out.write(f" 说明: -vsync 0 保持原始时间戳; showinfo 解析 pts 检测原始 dup/drop 异常\n")
            out.write(f"{'=' * 60}\n\n")
            frames, errors_list, dup_drop, decoders = [], [], [], []
            out_dir = Path(step['output_dir']) if step['output_dir'] else Path('.')
            for seg in step['segments']:
                seg_file = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
                out.write(f"$ ffmpeg -hide_banner -i {seg_file.name} -vsync 0 -vf showinfo -f null -\n")
                if not seg_file.exists():
                    print(f"  [WARN] 文件不存在: {seg_file}")
                    frames.append(None)
                    errors_list.append([f"[FILE NOT FOUND] {seg_file}"])
                    dup_drop.append(None)
                    decoders.append('n/a')
                    out.write("[FILE NOT FOUND]\n\n")
                    continue
                r = None
                if prefetched is not None:
                    r = prefetched.get(('ffmpeg', str(seg_file)))
                if r is None:
                    r = _ffmpeg_decode(str(seg_file), timeout=timeout, hwaccel=hwaccel,
                                       pts_drop_threshold=pts_drop_threshold)
                frames.append(r['frame_count'])
                segment_errors = list(r.get('errors') or [])
                # [HARD-FAIL-FRAME-CONSERVATION] 帧数守恒硬失败：ffmpeg 解码输出帧数
                # 远小于容器 packets（如损坏段1 frames=957 << packets=4960）时，
                # 显式标记。对齐 verify_segment_bitstream_v2 的 frames/packets 检查口径。
                _pkt_count = None
                if prefetched is not None:
                    _pv = prefetched.get(('ffprobe', str(seg_file), 'packets'))
                    if _pv is not None:
                        try:
                            _pkt_count = int(_pv)
                        except (TypeError, ValueError):
                            _pkt_count = None
                if (_pkt_count is not None and r['frame_count'] is not None
                        and _pkt_count > 0 and r['frame_count'] < int(_pkt_count * 0.8)):
                    _hard_note = ('[HARD-FAIL] 帧数守恒: frames(%s) << packets(%s)，'
                                  '静默丢帧/花屏风险' % (r['frame_count'], _pkt_count))
                    segment_errors.append(_hard_note)
                errors_list.append(segment_errors)
                # 对 pts 异常段额外检测 telecine/重复场标志
                rp = {'repeat_pict_frames': 0, 'repeat_pict_sum': 0}
                if (r.get('pts_dups', 0) > 0 or r.get('pts_drops', 0) > 0 or
                        r.get('pts_backward', 0) > 0):
                    rp = _ffprobe_repeat_pict(str(seg_file), timeout=min(timeout, 120))
                dup_drop.append({
                    'dup': r['dup_count'],
                    'drop': r['drop_count'],
                    'pts_dups': r.get('pts_dups', 0),
                    'pts_drops': r.get('pts_drops', 0),
                    'pts_drop_frames_est': r.get('pts_drop_frames_est', 0),
                    'pts_backward': r.get('pts_backward', 0),
                    'pts_anomalies': r.get('pts_anomalies', []),
                    'repeat_pict_frames': rp.get('repeat_pict_frames', 0),
                    'repeat_pict_sum': rp.get('repeat_pict_sum', 0),
                })
                decoder = r.get('decoder', 'cpu')
                decoders.append(decoder)
                if inline:
                    print(f"  [ffmpeg] Step{order} {seg_file.name}: frame={r['frame_count']}, "
                          f"vsync_dup={r['dup_count']}, vsync_drop={r['drop_count']}, "
                          f"pts_dups={r.get('pts_dups',0)}, pts_drops={r.get('pts_drops',0)}, "
                          f"pts_drop_est={r.get('pts_drop_frames_est',0)}, "
                          f"repeat_pict={rp.get('repeat_pict_frames',0)}, "
                          f"errors={len(r['errors'])} [{decoder}]")
                if r['frame_count'] is not None:
                    out.write(f"frame={r['frame_count']} [{decoder}]\n")
                if r.get('dup_count') is not None or r.get('drop_count') is not None:
                    out.write(f"vsync: dup={r.get('dup_count','?')} drop={r.get('drop_count','?')} "
                              f"(-vsync 0 下应为 0)\n")
                if r.get('pts_dups', 0) > 0 or r.get('pts_drops', 0) > 0:
                    out.write(f"[异常] 原始时间戳: pts_dups={r['pts_dups']} "
                              f"pts_drops={r['pts_drops']} "
                              f"pts_drop_frames_est={r.get('pts_drop_frames_est', 0)} "
                              f"pts_backward={r.get('pts_backward', 0)}\n")
                    for n, pts, pts_time, typ, note in r.get('pts_anomalies', []):
                        ts = f" pts_time={pts_time:.6f}" if pts_time is not None else ""
                        out.write(f"  [pts_anomaly] n={n} pts={pts}{ts} type={typ} {note}\n")
                if rp.get('repeat_pict_frames', 0) > 0:
                    out.write(f"[telecine] repeat_pict 帧数={rp['repeat_pict_frames']} "
                              f"sum={rp['repeat_pict_sum']} "
                              f"(可能存在 3:2 pulldown/重复场)\n")
                for warn in r.get('warnings', []):
                    out.write(f"[WARN-NVDEC] {warn}\n")
                for err in r['errors']:
                    out.write(err + '\n')
                out.write('\n')
            expected_segments = len(step['segments'])
            if not (len(frames) == len(errors_list) == len(dup_drop) == len(decoders) == expected_segments):
                raise RuntimeError(
                    f"Step {order} 结果列表长度不一致: "
                    f"segments={expected_segments}, frames={len(frames)}, "
                    f"errors={len(errors_list)}, dup_drop={len(dup_drop)}, "
                    f"decoders={len(decoders)}"
                )
            results[order] = (frames, errors_list, dup_drop, decoders)

    print(f"✅ ffmpeg_output 已生成: {output_path}")
    return results



def _ffprobe_count(video_path: str, count_mode: str = 'packets', timeout: int = 60,
                   hwaccel: str = 'off') -> str:
    """获取帧/包计数。count_mode: 'packets' | 'frames'

    CPU 路径: ffprobe -count_packets / -count_frames (packets 读包头 O(1); frames 逐帧软解)
    GPU 路径 (frames 模式, hwaccel='cuda'):
        ffprobe 本身不支持 -hwaccel，改用等效的 ffmpeg -hwaccel cuda -f null - 管线，
        解析 stderr 终末行的 frame= N 得到精确解码帧数（等价于 nb_read_frames），
        真正走 NVDEC 解码。GPU 失败时自动回退 CPU ffprobe。

    修复 4: packets 模式本质为 O(1) 读包头，却可能因系统 I/O 负载慢而被误判超时。
    将 packets 超时提升到 180s，且 [TIMEOUT] 后串行重试 1 次再判定。
    """
    gpu_attempted = (count_mode == 'frames' and hwaccel == 'cuda')
    if count_mode == 'packets' and timeout < 180:
        timeout = 180

    if gpu_attempted:
        # ffprobe 无 -hwaccel 选项 → 用 ffmpeg NVDEC null 管线等效替代
        cmd = ["ffmpeg", "-hide_banner",
               "-hwaccel", "cuda",
               "-threads", str(_clamp_decode_threads()),  # [FIX-NVDEC-THREAD-CAP] 解码线程钳位 ≤8
               "-i", str(video_path),
               "-an", "-f", "null", "-"]
    else:
        flag = '-count_packets' if count_mode == 'packets' else '-count_frames'
        entry = 'stream=nb_read_packets' if count_mode == 'packets' else 'stream=nb_read_frames'
        cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0", flag,
               "-show_entries", entry, "-of", "csv=p=0", str(video_path)]

    def _parse_output(result) -> str:
        """从 ffprobe stdout / ffmpeg stderr 提取帧数。
        优先级: stdout 纯数字 > stderr 'frame=N' 尾值（ffmpeg null 管线）> stdout strip > [ERROR]
        """
        stdout = (result.stdout or '').strip()
        stderr = (result.stderr or '').strip()
        # ffprobe 正常: stdout 为纯数字
        if stdout and re.fullmatch(r'\d+', stdout):
            return stdout
        # ffmpeg null 管线: stderr 末行含 frame= N
        for line in reversed(stderr.splitlines()):
            m = re.search(r'frame=\s*(\d+)', line)
            if m:
                return m.group(1)
        # 兜底: 返回非空文本（含错误提示），不把 stderr 垃圾数字误当帧数
        return stdout or "[ERROR]"

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        val = _parse_output(result)

        # GPU 失败 → 回退 CPU ffprobe
        if gpu_attempted and (val == "[ERROR]" or result.returncode != 0):
            flag_fb = '-count_frames'
            entry_fb = 'stream=nb_read_frames'
            cmd2 = ["ffprobe", "-v", "error", "-select_streams", "v:0", flag_fb,
                    "-show_entries", entry_fb, "-of", "csv=p=0", str(video_path)]
            try:
                result2 = subprocess.run(cmd2, capture_output=True, text=True, timeout=timeout)
                val2 = _parse_output(result2)
                return val2 if val2 != "[ERROR]" else "[ERROR]"
            except Exception:
                return val  # CPU 回退也失败，返回 GPU 错误描述
        return val
    except subprocess.TimeoutExpired:
        # 修复 4: 串行重试 1 次，规避系统 I/O 负载导致的瞬态超时
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
            val = _parse_output(result)
            return val
        except (subprocess.TimeoutExpired, Exception):
            return "[TIMEOUT]"
    except FileNotFoundError:
        return "[SKIP: ffprobe/ffmpeg not found]"


def _ffprobe_repeat_pict(video_path: str, timeout: int = 60) -> dict:
    """检测 telecine/重复场标志：统计 repeat_pict > 0 的帧数。

    ffprobe -show_entries frame=repeat_pict 会输出每帧的 repeat_pict 值。
    repeat_pict=N 表示该帧应显示 N+1 次（典型于 3:2 pulldown）。
    返回 {"repeat_pict_frames": int, "repeat_pict_sum": int, "details": list}
    """
    cmd = ["ffprobe", "-v", "error", "-select_streams", "v:0",
           "-show_entries", "frame=repeat_pict", "-of", "csv=p=0", str(video_path)]
    out = {'repeat_pict_frames': 0, 'repeat_pict_sum': 0, 'details': []}
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        if result.returncode != 0:
            return out
        for line in (result.stdout or '').splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                val = int(line)
            except ValueError:
                continue
            if val > 0:
                out['repeat_pict_frames'] += 1
                out['repeat_pict_sum'] += val
                if len(out['details']) < 10:
                    out['details'].append(val)
        return out
    except Exception:
        return out


def detect_gpu_hwaccel(ffmpeg_bin: str = 'ffmpeg') -> dict:
    """GPU 硬件加速能力自检。

    返回 {'cuda': True/False, 'nvenc': True/False, 'hint': str}
      - cuda: ffmpeg 列出 cuda hwaccel + NVDEC 解码 1 帧成功
      - nvenc: h264_nvenc 编码器可用
      任一 False → 全局回退纯 CPU。

    自检策略（无临时文件）：
      - NVDEC: lavfi → libx264 → -f h264 pipe（Annex B 比特流自带帧边界，
        无 rawvideo pipe 截断问题）
      - NVENC: lavfi → h264_nvenc → null（直接测试编码器，单条命令）
    """
    result = {'cuda': False, 'nvenc': False, 'hint': ''}

    # 1. Check hwaccel support
    try:
        r = subprocess.run([ffmpeg_bin, '-hide_banner', '-hwaccels'],
                           capture_output=True, text=True, timeout=10)
        if 'cuda' not in r.stdout.lower():
            result['hint'] = 'ffmpeg 未编译 CUDA hwaccel 支持'
            return result
    except (FileNotFoundError, subprocess.TimeoutExpired):
        result['hint'] = 'ffmpeg 不可用'
        return result

    # 2. Micro self-test: NVDEC decode 1 black frame
    #    使用 -f h264 pipe（Annex B 比特流带 start code 帧边界），
    #    避免 -f rawvideo pipe 截断导致 packet corrupt / Invalid buffer size 假阴性。
    try:
        r = subprocess.run([
            ffmpeg_bin, '-hide_banner', '-y',
            '-f', 'lavfi', '-i', 'color=c=black:s=160x160:d=1:r=10',
            '-frames:v', '10', '-c:v', 'libx264', '-preset', 'ultrafast',
            '-f', 'h264', 'pipe:1'
        ], capture_output=True, text=False, timeout=20)
        if r.returncode != 0:
            result['hint'] = 'ffmpeg 无法生成测试流'
            return result

        # NVDEC decode from H.264 pipe (stdin)
        r2 = subprocess.run([
            ffmpeg_bin, '-hide_banner', '-hwaccel', 'cuda',
            '-threads', str(_clamp_decode_threads()),  # [FIX-NVDEC-THREAD-CAP] 解码线程钳位 ≤8
            '-f', 'h264', '-i', 'pipe:0', '-frames:v', '1', '-f', 'null', '-'
        ], input=r.stdout, capture_output=True, text=False, timeout=20)
        if r2.returncode != 0:
            result['hint'] = 'NVDEC 自检失败（驱动/CUDA 可能不可用）'
            return result
        result['cuda'] = True
    except (subprocess.TimeoutExpired, FileNotFoundError):
        result['hint'] = 'NVDEC 自检超时或 ffmpeg 缺失'
        return result

    # 3. nvenc self-test: 直接测试 h264_nvenc（lavfi → NVENC → null）
    #    步骤 2 已验证 NVDEC，此处仅需验证 NVENC 编码器即可覆盖全 GPU 管线。
    #    单条命令，无 pipe、无临时文件 — 比原 rawvideo pipe 方案更简洁可靠。
    #    NVIDIA NVENC 硬件编码器对视频分辨率有明确的最小限制
    #    10 帧可稳定触发编码器初始化。
    try:
        r = subprocess.run([
            ffmpeg_bin, '-hide_banner',
            '-f', 'lavfi', '-i', 'color=c=black:s=160x160:d=1:r=10',
            '-frames:v', '10', '-c:v', 'h264_nvenc', '-f', 'null', '-'
        ], capture_output=True, text=True, timeout=20)
        if r.returncode == 0:
            result['nvenc'] = True
        else:
            stderr_tail = r.stderr.strip().split('\n')[-3:] if r.stderr else []
            stderr_info = '; '.join(stderr_tail[-2:]) if len(stderr_tail) >= 2 else ''
            if stderr_info:
                result['hint'] = f"NVENC encoder self-test failed ({stderr_info})"
            else:
                result['hint'] = 'NVENC encoder self-test failed (check driver/nvidia-smi)'
    except (subprocess.TimeoutExpired, FileNotFoundError):
        pass

    if result['cuda'] and not result['hint']:
        result['hint'] = 'CUDA' + ('+NVENC' if result['nvenc'] else '') + ' GPU acceleration available'
    return result


def _to_int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _parse_elapsed_to_sec(elapsed_str: str) -> float:
    """
    解析耗时字符串为总秒数，用于没有 [性能统计] 的步骤推算 FPS/平均耗时。
    支持格式: 'HH:MM:SS.ms' / 'MM:SS.ms' / 'SS.ms'
    """
    if not elapsed_str:
        return None
    try:
        parts = elapsed_str.strip().split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + float(parts[1])
        else:
            return float(parts[0])
    except (ValueError, TypeError):
        return None


def collect_ffprobe_packets_tasks(info: dict) -> list:
    """
    收集 packets 模式（快速）ffprobe 统计任务 — 第一阶段扫描。
    仅读容器 header，不逐帧解码，O(1) 成本获取各分段总帧数。
    任务 key: ('ffprobe', path, 'packets')
    """
    tasks = []
    seen = set()

    def add(video_path, mode, timeout):
        key = ('ffprobe', str(video_path), mode)
        if key in seen or not Path(video_path).exists():
            return
        seen.add(key)
        tasks.append({'kind': 'ffprobe', 'key': key, 'path': str(video_path),
                      'mode': mode, 'timeout': timeout})

    for step in info['steps']:
        in_dir = Path(step['input_dir']) if step['input_dir'] else None
        out_dir = Path(step['output_dir']) if step['output_dir'] else None
        for seg in step['segments']:
            if in_dir:
                f = _find_segment_file(in_dir, step['input_prefix'], seg['index'])
                add(f, 'packets', 60)
            if out_dir:
                f = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
                add(f, 'packets', 60)

    final_video = info.get('output_video', '')
    if final_video and Path(final_video).exists():
        add(final_video, 'packets', 120)

    return tasks


def collect_ffprobe_frames_tasks(info: dict, hwaccel: str = 'off') -> list:
    """
    收集 frames 模式（全量解码）ffprobe 统计任务 — 第二阶段重任务。
    需逐帧解码全部画面，成本 O(帧数)，由第一阶段 packets 结果估算后调度。
    任务 key: ('ffprobe', path, 'frames')
    hwaccel: 'cuda' GPU 加速 (NVDEC 解码), 'off' 纯 CPU。
    """
    tasks = []
    seen = set()

    def add(video_path, mode, timeout):
        key = ('ffprobe', str(video_path), mode)
        if key in seen or not Path(video_path).exists():
            return
        seen.add(key)
        tasks.append({'kind': 'ffprobe', 'key': key, 'path': str(video_path),
                      'mode': mode, 'timeout': timeout, 'hwaccel': hwaccel})

    for step in info['steps']:
        in_dir = Path(step['input_dir']) if step['input_dir'] else None
        out_dir = Path(step['output_dir']) if step['output_dir'] else None
        for seg in step['segments']:
            if in_dir:
                f = _find_segment_file(in_dir, step['input_prefix'], seg['index'])
                add(f, 'frames', 900)
            if out_dir:
                f = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index'])
                add(f, 'frames', 900)

    # 最终合并输出视频：仅文件存在时加入（避免不存在的路径被 ffprobe stderr 数字误读为帧数）
    final_video = info.get('output_video', '')
    if final_video:
        add(final_video, 'frames', 3600)

    return tasks


def generate_ffprob_output(info: dict, output_path: str, prefetched: dict = None,
                           skip_frames: bool = False, hwaccel: str = 'off'):
    """
    对初始分段 + 各步骤输出分段 + 最终输出视频执行 ffprobe 统计。
    中间分段（前步骤输出=后步骤输入）只统计一次。
    返回 {step_order: {'in_packets': [], 'in_frames': [], 'out_packets': [], 'out_frames': []},
           'final': {'packets': X, 'frames': Y}}

    v3: prefetched 为并行预取结果 {('ffprobe', path, mode): count_str}，
        命中时直接填充缓存（保持 v1 输出文件格式与分段顺序），缺项回退串行统计。
    skip_frames: 跳过 -count_frames（全量解码），-count_packets 快速扫描不受影响。
    hwaccel: 'cuda' → ffprobe/frames 用等效 ffmpeg NVDEC null 管线；
             'off' → 原生 ffprobe 命令。
    """
    data = {}
    cache = {}  # path -> {'packets': X, 'frames': Y}，避免重复统计中间分段
    inline = prefetched is None  # 无预取结果时为 v1 串行模式（逐条打印进度）

    def probe(video_path: Path, mode: str, timeout: int) -> str:
        key = str(video_path)
        if key not in cache:
            cache[key] = {}
        if mode not in cache[key]:
            hit = prefetched.get(('ffprobe', key, mode)) if prefetched is not None else None
            cache[key][mode] = hit if hit is not None else _ffprobe_count(key, mode, timeout,
                                                                          hwaccel=hwaccel)
        return cache[key][mode]

    with open(output_path, 'w', encoding='utf-8') as out:
        out.write("=" * 70 + "\n")
        out.write(" ffprobe 帧数统计 — 视频增强全流程\n")
        out.write(f" 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        out.write("=" * 70 + "\n")

        for step in info['steps']:
            order = step['order']
            sdata = {'in_packets': [], 'in_frames': [], 'out_packets': [], 'out_frames': []}
            out.write(f"\n── Step {order}: {step['name']} ({step['type']}) ──\n")
            out.write(f"   input_dir : {step['input_dir']}  (prefix: {step['input_prefix']})\n")
            out.write(f"   output_dir: {step['output_dir']}  (prefix: {step['output_prefix']})\n")

            in_dir = Path(step['input_dir']) if step['input_dir'] else None
            out_dir = Path(step['output_dir']) if step['output_dir'] else None

            for mode, key, timeout in (('packets', 'in_packets', 60), ('frames', 'in_frames', 900)):
                if skip_frames and mode == 'frames':
                    continue
                out.write(f"\n  [input]  nb_read_{mode}:\n")
                for seg in step['segments']:
                    f = _find_segment_file(in_dir, step['input_prefix'], seg['index']) if in_dir else None
                    if f and f.exists():
                        c = probe(f, mode, timeout)
                        if inline:
                            print(f"  [ffprobe] Step{order} in/{mode} {f.name}: {c}")
                        sdata[key].append(c)
                        out.write(f"    {f.name}: {c}\n")
                    else:
                        sdata[key].append("[NOT FOUND]")
                        out.write(f"    {step['input_prefix']}_{seg['index']:03d}.mp4: [FILE NOT FOUND]\n")

            for mode, key, timeout in (('packets', 'out_packets', 60), ('frames', 'out_frames', 900)):
                if skip_frames and mode == 'frames':
                    continue
                out.write(f"\n  [output] nb_read_{mode}:\n")
                for seg in step['segments']:
                    f = _find_segment_file(out_dir, step['output_prefix'] or 'segment', seg['index']) if out_dir else None
                    if f and f.exists():
                        c = probe(f, mode, timeout)
                        if inline:
                            print(f"  [ffprobe] Step{order} out/{mode} {f.name}: {c}")
                        sdata[key].append(c)
                        out.write(f"    {f.name}: {c}\n")
                    else:
                        sdata[key].append("[NOT FOUND]")
                        out.write(f"    {(step['output_prefix'] or 'segment')}_{seg['index']:03d}.mp4: [FILE NOT FOUND]\n")

            data[order] = sdata

        # ── 最终输出视频 ──
        final_video = info.get('output_video', '')
        data['final'] = {'packets': None, 'frames': None}
        if final_video and Path(final_video).exists():
            out.write(f"\n── 最终输出视频: {Path(final_video).name} ──\n")
            c = probe(Path(final_video), 'packets', 120)
            if inline:
                print(f"  [ffprobe] final/packets {Path(final_video).name}: {c}")
            data['final']['packets'] = c
            out.write(f"  nb_read_packets: {c}\n")
            if not skip_frames:
                c = probe(Path(final_video), 'frames', 3600)
                if inline:
                    print(f"  [ffprobe] final/frames  {Path(final_video).name}: {c}")
                data['final']['frames'] = c
                out.write(f"  nb_read_frames : {c}\n")
        elif final_video:
            out.write(f"\n── 最终输出视频不存在: {final_video} ──\n")

        out.write("\n" + "=" * 70 + "\n")

    print(f"✅ ffprob_output 已生成: {output_path}")
    return data


# ═══════════════════════════════════════════════
# 4. 系统资源探测与并行执行引擎（v3）
# ═══════════════════════════════════════════════

def _read_text_strip(path: str) -> str:
    """读取文件并 strip，失败返回 None。"""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception:
        return None


def _read_int(path: str) -> int:
    text = _read_text_strip(path)
    if text is None:
        return None
    try:
        return int(text)
    except Exception:
        return None


def _parse_cpuset_count(cpuset_text: str) -> int:
    """解析 cpuset.cpus 如 '0-3,5,7-9'，返回核心数。"""
    if not cpuset_text:
        return 0
    count = 0
    for part in cpuset_text.split(','):
        part = part.strip()
        if '-' in part:
            try:
                a, b = part.split('-', 1)
                count += int(b) - int(a) + 1
            except Exception:
                continue
        elif part:
            try:
                count += 1
            except Exception:
                continue
    return count


def detect_cpu_count() -> int:
    """
    容器感知的 CPU 核数探测。
      - cgroup v2: /sys/fs/cgroup/cpu.max
      - cgroup v1: /sys/fs/cgroup/cpu/cpu.cfs_quota_us / cpu.cfs_period_us
      - 无 quota 时回退 cpuset（v1/v2 路径），最后再回退 os.cpu_count()
    os.cpu_count() 在容器里通常返回宿主机核心数，因此优先读 cgroup。
    """
    # cgroup v2
    cpu_max = _read_text_strip('/sys/fs/cgroup/cpu.max')
    if cpu_max:
        parts = cpu_max.split()
        if len(parts) == 2 and parts[0].lower() != 'max':
            try:
                quota = int(parts[0])
                period = int(parts[1])
                if period > 0 and quota > 0:
                    return max(1, math.ceil(quota / period))
            except Exception:
                pass
    # cgroup v1
    quota = _read_int('/sys/fs/cgroup/cpu/cpu.cfs_quota_us')
    period = _read_int('/sys/fs/cgroup/cpu/cpu.cfs_period_us')
    if quota is not None and period and period > 0 and quota > 0:
        return max(1, math.ceil(quota / period))
    # cpuset 回退
    for cpuset_path in (
        '/sys/fs/cgroup/cpuset/cpuset.cpus',
        '/sys/fs/cgroup/cpuset.cpus.effective',
        '/sys/fs/cgroup/cpuset.cpus',
    ):
        cpuset = _read_text_strip(cpuset_path)
        if cpuset:
            cnt = _parse_cpuset_count(cpuset)
            if cnt > 0:
                return cnt
    return os.cpu_count() or 1


def _detect_cgroup_memory_gb():
    """
    读取 cgroup 内存限制与已用量（GB），返回 (limit_gb, usage_gb)。
    如果未设置限制，limit_gb 为 None；'max' 或极大值也表示无限制。
    """
    limit_gb = None
    usage_gb = None
    # cgroup v2
    limit_text = _read_text_strip('/sys/fs/cgroup/memory.max')
    if limit_text is not None:
        if limit_text.lower() != 'max':
            try:
                limit_gb = int(limit_text) / 1e9
            except Exception:
                pass
        usage_text = _read_text_strip('/sys/fs/cgroup/memory.current')
        if usage_text:
            try:
                usage_gb = int(usage_text) / 1e9
            except Exception:
                pass
        return limit_gb, usage_gb
    # cgroup v1
    limit_text = _read_text_strip('/sys/fs/cgroup/memory/memory.limit_in_bytes')
    if limit_text is not None:
        try:
            limit_gb = int(limit_text) / 1e9
        except Exception:
            pass
        usage_text = _read_text_strip('/sys/fs/cgroup/memory/memory.usage_in_bytes')
        if usage_text:
            try:
                usage_gb = int(usage_text) / 1e9
            except Exception:
                pass
    return limit_gb, usage_gb


def _detect_host_ram_gb():
    """
    跨平台探测物理内存，返回 (total_gb, avail_gb)；失败返回 (0.0, 0.0)。
    优先 psutil；Windows 回退 GlobalMemoryStatusEx；Linux 回退 /proc/meminfo → sysconf。
    """
    try:
        import psutil
        vm = psutil.virtual_memory()
        return vm.total / 1e9, vm.available / 1e9
    except ImportError:
        pass
    except Exception:
        pass
    if sys.platform == 'win32':
        try:
            import ctypes

            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [("dwLength", ctypes.c_ulong), ("dwMemoryLoad", ctypes.c_ulong),
                            ("ullTotalPhys", ctypes.c_ulonglong), ("ullAvailPhys", ctypes.c_ulonglong),
                            ("ullTotalPageFile", ctypes.c_ulonglong), ("ullAvailPageFile", ctypes.c_ulonglong),
                            ("ullTotalVirtual", ctypes.c_ulonglong), ("ullAvailVirtual", ctypes.c_ulonglong),
                            ("sullAvailExtendedVirtual", ctypes.c_ulonglong)]

            stat = MEMORYSTATUSEX()
            stat.dwLength = ctypes.sizeof(stat)
            ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(stat))
            return stat.ullTotalPhys / 1e9, stat.ullAvailPhys / 1e9
        except Exception:
            return 0.0, 0.0
    else:
        try:
            with open('/proc/meminfo', 'r', encoding='utf-8') as f:
                meminfo = f.read()
            total = int(re.search(r'MemTotal:\s+(\d+)\s*kB', meminfo).group(1)) / 1e6
            avail = int(re.search(r'MemAvailable:\s+(\d+)\s*kB', meminfo).group(1)) / 1e6
            return total, avail
        except Exception:
            try:
                page_size = os.sysconf('SC_PAGE_SIZE')
                return (os.sysconf('SC_PHYS_PAGES') * page_size / 1e9,
                        os.sysconf('SC_AVPHYS_PAGES') * page_size / 1e9)
            except Exception:
                return 0.0, 0.0


def detect_ram_gb():
    """
    跨平台 + 容器感知探测内存，返回 (total_gb, avail_gb)；失败返回 (0.0, 0.0)。
      - 宿主机：psutil 优先；Windows 回退 GlobalMemoryStatusEx；Linux 回退 /proc/meminfo/sysconf
      - 容器：再用 cgroup v1/v2 的 memory.limit/usage 修正，避免读到宿主机全部内存
    """
    total, avail = _detect_host_ram_gb()
    cg_limit, cg_usage = _detect_cgroup_memory_gb()
    if cg_limit is not None:
        # 如果 limit 极大（如 2^63-1），说明未限制，仍用宿主机值
        total = min(total, cg_limit) if total > 0 else cg_limit
        if cg_usage is not None:
            avail = max(0.0, total - cg_usage)
        else:
            avail = min(avail, total) if avail > 0 else total
    return total, avail


def compute_auto_workers(task_ram_mb: int = 300, reserve_ratio: float = 0.10) -> int:
    """
    按容器感知后的系统资源自动计算并行 workers：
      workers = min(容器可用 CPU, RAM总计*(1-预留比例) / 单任务内存估计)
    ffprobe/ffmpeg 解码子进程本身吃 CPU，故不超过 CPU 核数；RAM 不足时自动缩减。
    默认预留 10% 给系统/其他进程（参考 subtitle_generator_whisper_v2_process.py 思路，
    但按容器 cgroup 修正 total/avail）。
    """
    cpu = detect_cpu_count()
    total_gb, avail_gb = detect_ram_gb()
    by_ram = cpu
    if total_gb > 0:
        usable_gb = total_gb * (1 - reserve_ratio)
    elif avail_gb > 0:
        usable_gb = avail_gb * 0.9
    else:
        usable_gb = 0.0
    usable_mb = max(0.0, usable_gb * 1024)
    if usable_mb > 0:
        by_ram = max(1, int(usable_mb // max(1, task_ram_mb)))
    return max(1, min(cpu, by_ram))


def _extract_segment_frame_counts(prefetched: dict) -> dict:
    """
    从第一阶段 packets 结果中提取各分段文件的帧数。
    返回 {path_str: frame_count_int}。
    无法解析的忽略，缺项后续排序时 cost=0 放在队尾。
    """
    costs = {}
    for (kind, path, mode), val in prefetched.items():
        if kind == 'ffprobe' and mode == 'packets':
            try:
                costs[path] = int(val)
            except (ValueError, TypeError):
                pass
    return costs


def _path_step_priority(path_str: str, step_dir_priority: dict):
    """根据文件路径返回 (step_order, type_priority)。type: output=1(高), input=0(低)。"""
    ps = str(Path(path_str).resolve())
    best = (-1, 0)
    for sd, (so, tp) in step_dir_priority.items():
        if ps.startswith(sd) or sd in ps:
            if (so, tp) > best:
                best = (so, tp)
    return best


def _enrich_and_sort_heavy_tasks(tasks: list, frame_costs: dict,
                                   step_dir_priority: dict = None,
                                   final_video_paths: set = None) -> list:
    """
    Phase 2 重任务排序（与阶段 1 一致）：
      1. 最终视频 (cost 最大)
      2. Step N 输出 → Step N 输入 → Step N-1 输出 → ...
      3. 同组内：ffprobe 先于 ffmpeg（ffprobe 慢，先跑）
      4. 同组同类型内：cost 降序（大文件优先）
    """
    step_dir_priority = step_dir_priority or {}
    final_video_paths = final_video_paths or set()
    for t in tasks:
        cost = frame_costs.get(t['path'], 0)
        t['cost'] = cost
        # 根据帧数自适应超时：基础 300s，每 1000 帧加 60s，上限 3600s，下限 300s
        if t['kind'] in ('ffmpeg', 'ffprobe'):
            adaptive = min(3600, max(300, 300 + cost // 1000 * 60))
            t['timeout'] = max(t.get('timeout', 0), adaptive)

    def sort_key(t):
        path = t['path']
        # 最终视频：最高优先级
        if path in final_video_paths:
            return (999, 0, 0, -t['cost'])
        # Step 优先级
        sp = _path_step_priority(path, step_dir_priority)
        order = sp[0]  # 越大越靠后 = 越优先
        typ = sp[1]    # output=1(高), input=0(低)
        # 任务类型优先级：ffprobe(1) 先, ffmpeg(0) 后 — ffprobe 慢，先跑
        kind_p = 1 if t['kind'] == 'ffprobe' else 0
        return (order, typ, kind_p, -t['cost'])

    return sorted(tasks, key=sort_key, reverse=True)


def _execute_task(task: dict):
    """执行单个探测/解码任务（模块级函数，ProcessPool 下可 pickle）。

    返回 (result, actual_start_ts) —— 开始时间在 worker 内记录，
    随返回值跨进/线程回传，保证 thread/process 双模式计时准确。

    GPU 任务受 _GPU_SEMAPHORE 闸门限制 (thread 模式)；
    process 模式下信号量不可跨进程共享，GPU 并发由 --gpu-workers 上限约束
    (ProcessPool 本身 max_workers 即上限)。
    """
    gpu_sem = _GPU_SEMAPHORE
    gpu_task = task.get('hwaccel', 'off') == 'cuda'
    if gpu_task and gpu_sem is not None:
        gpu_sem.acquire()

    start = time.time()
    try:
        hw = task.get('hwaccel', 'off')
        if task['kind'] in ('ffprobe', 'ffprobe_recheck'):
            return _ffprobe_count(task['path'], task['mode'], task['timeout'],
                                  hwaccel=hw), start
        # ffmpeg / ffmpeg_recheck
        return _ffmpeg_decode(task['path'],
                              timeout=task.get('timeout', 900),
                              hwaccel=hw,
                              pts_drop_threshold=task.get('pts_drop_threshold', 2.0)), start
    finally:
        if gpu_task and gpu_sem is not None:
            gpu_sem.release()



def _task_error_result(task: dict, exc: Exception):
    """任务异常时生成与正常结果同结构的兜底值，保证后续生成流程不中断"""
    if task['kind'] == 'ffprobe':
        return f"[ERROR: {exc}]"
    return {'frame_count': None, 'errors': [f"[ERROR] 并行执行异常: {exc}"],
            'dup_count': None, 'drop_count': None}


def _print_task_progress(done: int, total: int, task: dict, result,
                         task_duration: float, work_done: int = 0,
                         total_work: int = 0, total_elapsed: float = 0):
    """逐任务进度打印。task_duration=本任务用时，total_elapsed=批次累计（助 ETA 估算）。"""
    name = Path(task['path']).name
    work_str = ""
    if total_work and total_work > 0:
        pct = min(100, work_done * 100 // total_work)
        eta_s = (total_elapsed / max(1, work_done)) * max(0, total_work - work_done) if work_done > 0 else 0
        eta = f" ETA {eta_s/60:.0f}分{eta_s%60:.0f}秒" if eta_s > 0 else ""
        work_str = f" [工作量 {pct}%{eta}]"
    if task['kind'] == 'ffprobe':
        print(f"  [{done}/{total}] ffprobe/{task['mode']} {name}: {result}{work_str}  (用时 {task_duration:.1f}s)")
    elif isinstance(result, dict):
        print(f"  [{done}/{total}] ffmpeg {name}: frame={result.get('frame_count')}, "
              f"dup={result.get('dup_count')}, drop={result.get('drop_count')}, "
              f"errors={len(result.get('errors', []))}{work_str}  (用时 {task_duration:.1f}s)")
    else:
        print(f"  [{done}/{total}] ffmpeg {name}: {result}{work_str}  (用时 {task_duration:.1f}s)")


def run_tasks_parallel(tasks: list, workers: int, parallel_mode: str = 'thread',
                       total_work: int = 0, phase_label: str = '') -> dict:
    """
    并行执行任务列表，返回 {task['key']: result}。
      - workers<=1 时退化为 v1 串行语义
      - parallel_mode: 'thread' / 'process'
      - total_work: 任务总帧数（用于工作量百分比 + ETA），0 则仅显示任务计数
      - phase_label: 阶段标识字符串，如 '阶段 1/2: 快速扫描'
    """
    results = {}
    total = len(tasks)
    if total == 0:
        return results
    workers = max(1, min(workers, total))
    t_batch = time.time()

    if workers == 1:
        work_acc = 0
        for i, t in enumerate(tasks, 1):
            t0 = time.time()
            real_start = t0
            try:
                res, real_start = _execute_task(t)
                results[t['key']] = res
            except Exception as e:
                results[t['key']] = _task_error_result(t, e)
            task_dur = time.time() - real_start
            total_elapsed = time.time() - t_batch
            work_acc += t.get('cost', 0)
            _print_task_progress(i, total, t, results[t['key']], task_dur,
                                 work_done=work_acc, total_work=total_work,
                                 total_elapsed=total_elapsed)
        return results

    executor_cls = ProcessPoolExecutor if parallel_mode == 'process' else ThreadPoolExecutor
    done = 0
    work_acc = 0
    with executor_cls(max_workers=workers) as ex:
        fut2task = {}
        for t in tasks:
            fut2task[ex.submit(_execute_task, t)] = t
        for fut in as_completed(fut2task):
            t = fut2task[fut]
            total_elapsed = time.time() - t_batch
            try:
                res, real_start = fut.result()
                results[t['key']] = res
                task_dur = time.time() - real_start
            except Exception as e:
                results[t['key']] = _task_error_result(t, e)
                task_dur = time.time() - t_batch  # 异常回落：批次已用时间
            done += 1
            work_acc += t.get('cost', 0)
            _print_task_progress(done, total, t, results[t['key']], task_dur,
                                 work_done=work_acc, total_work=total_work,
                                 total_elapsed=total_elapsed)
    return results


def build_chain(info: dict, ffprobe_data: dict, ffmpeg_results: dict):
    """
    计算每段跨步骤帧数链。
    返回 rows: [{'index', 'chain': [{step_order, in_frames, expected, log_out, actual, diff,
                                     log_diff, ffmpeg_val, ffmpeg_diff, dup, drop, errors}], }]
    expected 规则:
      interpolate: (输入帧 - 1) * 倍数 + 1
      upscale:     输入帧（= 前一步骤实际输出，ffprobe 优先）
    """
    rows = []
    n = info['num_segments']
    prev_actual = [None] * n  # 前一步骤实际输出帧数

    for step in info['steps']:
        order = step['order']
        sdata = ffprobe_data.get(order, {})
        in_pk = sdata.get('in_packets', [])
        in_fr = sdata.get('in_frames', [])
        out_pk = sdata.get('out_packets', [])
        out_fr = sdata.get('out_frames', [])
        vals = ffmpeg_results.get(order, ([], [], [], []))
        ff_frames, ff_errors, ff_dd = vals[:3]

        for i, seg in enumerate(step['segments']):
            while len(rows) <= i:
                rows.append({'index': rows.__len__(), 'chain': []})

            # 输入帧数：ffprobe frames > packets > 前一步骤实际输出 > 日志 raw
            in_frames = (_to_int(in_fr[i]) if i < len(in_fr) else None) or \
                        (_to_int(in_pk[i]) if i < len(in_pk) else None)
            if in_frames is None:
                in_frames = prev_actual[i] if prev_actual[i] is not None else (seg['raw_frames'] or 0)

            if step['type'] == 'interpolate':
                mult = step['factor'] or 2
                expected = (in_frames - 1) * mult + 1
            else:
                expected = in_frames

            actual = (_to_int(out_fr[i]) if i < len(out_fr) else None) or \
                     (_to_int(out_pk[i]) if i < len(out_pk) else None)
            log_out = seg['output_frames'] or 0

            ff_val = ff_frames[i] if i < len(ff_frames) else None
            dd = ff_dd[i] if i < len(ff_dd) else None
            errs = ff_errors[i] if i < len(ff_errors) else []

            entry = {
                'step_order': order,
                'step_type': step['type'],
                'in_frames': in_frames,
                'expected': expected,
                'log_out': log_out,
                'log_diff': log_out - expected,
                'actual': actual,
                'diff': (actual - expected) if actual is not None else None,
                'ff_val': ff_val,
                'ff_diff': (ff_val - expected) if ff_val is not None else None,
                'dup': dd.get('dup') if dd else None,
                'drop': dd.get('drop') if dd else None,
                'pts_dups': dd.get('pts_dups', 0) if dd else 0,
                'pts_drops': dd.get('pts_drops', 0) if dd else 0,
                'pts_drop_frames_est': dd.get('pts_drop_frames_est', 0) if dd else 0,
                'pts_backward': dd.get('pts_backward', 0) if dd else 0,
                'repeat_pict_frames': dd.get('repeat_pict_frames', 0) if dd else 0,
                'repeat_pict_sum': dd.get('repeat_pict_sum', 0) if dd else 0,
                'errors': errs,
                'seg': seg,
            }
            rows[i]['chain'].append(entry)
            prev_actual[i] = actual if actual is not None else (log_out or prev_actual[i])

    return rows


def detect_anomalies(entry: dict):
    """返回 (anomaly_str, is_bad, is_warn)"""
    errors = entry['errors']
    has_illegal = any('illegal' in e.lower() for e in errors)
    has_missing = any('missing' in e.lower() for e in errors)
    has_duplicate = any('more than 1000 frames duplicated' in e.lower() for e in errors)
    has_no_frame = any('no frame' in e.lower() for e in errors)
    has_decode_error = any('decode_slice_header error' in e.lower() for e in errors)
    has_pps_error = any('non-existing pps' in e.lower() for e in errors)

    parts = []
    if has_illegal:
        parts.append('⚠ illegal buffer(花屏)')
    if has_pps_error:
        parts.append('⚠ PPS丢失')
    if has_decode_error:
        parts.append('⚠ 解码错误')
    if has_no_frame:
        parts.append('❌ no frame!(丢帧)')
    if has_missing:
        parts.append('⚠ missing ref(丢帧)')
    if has_duplicate:
        parts.append('🔁 大量重复帧(>1000)')
    # 新增: 基于 showinfo pts 解析的原始时间戳异常
    if entry.get('pts_dups', 0) > 0:
        parts.append(f"🔁 原始重复帧(pts_dup={entry['pts_dups']})")
    if entry.get('pts_drops', 0) > 0:
        est = entry.get('pts_drop_frames_est', 0)
        parts.append(f"❌ 原始丢帧(pts_drop={entry['pts_drops']}, est_missing={est})")
    if entry.get('pts_backward', 0) > 0:
        parts.append(f"⚠ pts 回退({entry['pts_backward']})")
    if entry.get('repeat_pict_frames', 0) > 0:
        parts.append(f"🎬 telecine(repeat_pict={entry['repeat_pict_frames']})")
    # vsync dup/drop（-vsync 0 下应为 0，非 0 说明时间戳被修改）
    if entry['dup'] is not None and entry['dup'] > 0:
        parts.append(f"⚠ vsync_dup={entry['dup']}(-vsync0下异常)")
    if entry['drop'] is not None and entry['drop'] > 0:
        parts.append(f"⚠ vsync_drop={entry['drop']}(-vsync0下异常)")
    # 超分场景偏差应严格为 0，插帧容忍 ±2
    tol = 0 if entry['step_type'] == 'upscale' else 2
    if entry['log_diff'] is not None and abs(entry['log_diff']) > tol:
        parts.append(f"📉 日志偏差={entry['log_diff']}")
    if entry['diff'] is not None and abs(entry['diff']) > max(2, tol):
        parts.append(f"📉 ffprobe偏差={entry['diff']}")
    if entry['ff_diff'] is not None and abs(entry['ff_diff']) > max(2, tol):
        parts.append(f"📉 拆帧偏差={entry['ff_diff']}")

    joined = ''.join(parts)
    is_bad = any(kw in joined for kw in ['❌', '⚠ illegal', '⚠ PPS'])
    is_warn = (not is_bad) and len(parts) > 0
    return ('\n'.join(parts) if parts else '✅ 正常'), is_bad, is_warn



def generate_xlsx(info: dict, ffprobe_data: dict, ffmpeg_results: dict, chain_rows: list,
                  output_path: str):
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl 未安装，跳过 xlsx 生成")
        return

    wb = openpyxl.Workbook()

    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    title_font = Font(name='微软雅黑', size=14, bold=True, color='1F3864')
    normal_font = Font(name='微软雅黑', size=10)
    warn_font = Font(name='微软雅黑', size=10, color='CC6600')
    bad_font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
    ok_font = Font(name='微软雅黑', size=10, color='008000')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    bad_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
    sum_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    sum_font = Font(name='微软雅黑', size=10, bold=True)

    steps = info['steps']
    n_steps = len(steps)

    def style_cell(cell, font=normal_font, align=center_align, fill=None):
        cell.font = font
        cell.alignment = align
        cell.border = thin_border
        if fill:
            cell.fill = fill

    def diff_font(dv):
        if dv is None:
            return normal_font, None
        if abs(dv) > 2:
            return bad_font, bad_fill
        if abs(dv) > 0.5:
            return warn_font, warn_fill
        return ok_font, None

    # ══ Sheet 1: 全流程汇总 ══
    ws = wb.active
    ws.title = "全流程汇总"

    # 列: 分段序号 | 分段编号 | (每步: 输入/期望/日志输出/实际输出/差值) | 异常 | 最终比对
    headers = ['分段序号', '分段编号']
    for s in steps:
        headers += [f"Step{s['order']}\n输入帧", f"Step{s['order']}\n期望输出",
                    f"Step{s['order']}\n实际输出", f"Step{s['order']}\n差值"]
    headers += ['异常检测\n(全部步骤)']

    num_cols = len(headers)
    widths = [8, 16] + [12, 12, 12, 10] * n_steps + [40]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = '视频增强全流程 — 帧数链统计汇总'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # 概览
    row = 4
    step_desc = ' → '.join(f"Step{s['order']}:{s['name']}" for s in steps)
    # 处理模式列展示完整模式字符串（基础模式 + 跳过标志，幂等兜底）
    mode_display = _compose_mode_display(
        info.get('mode', ''),
        skip_upscale=info.get('skip_upscale', False),
        skip_interpolate=info.get('skip_interpolate', False),
    ) or 'N/A'
    info_items = [
        ('源视频', info.get('source_video', 'N/A')),
        ('最终输出', info.get('output_video', 'N/A')),
        ('处理模式', mode_display),
        ('步骤链', step_desc),
        ('初始分段目录', steps[0]['input_dir'] if steps else 'N/A'),
        ('分段数', str(info.get('num_segments', 0))),
        ('分段时长', f"{info.get('segment_duration', 300)}秒"),
        ('输出大小', info.get('total_output_size', 'N/A')),
        ('总用时', info.get('total_elapsed', 'N/A')),
    ]
    final_probe = ffprobe_data.get('final', {})
    if final_probe.get('frames'):
        info_items.append(('最终视频帧数', f"packets={final_probe.get('packets')} frames={final_probe.get('frames')}"))
    for key, val in info_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws[f'A{row}'] = key
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=min(8, num_cols))
        ws[f'C{row}'] = val
        ws[f'C{row}'].font = normal_font
        row += 1
    row += 1

    header_row = row
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=ci, value=h)
        style_cell(cell, header_font, center_align, header_fill)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 48
    row += 1

    total_bad = 0
    total_warn = 0
    sum_expected = [0] * n_steps
    sum_actual = [0] * n_steps

    for i, r in enumerate(chain_rows):
        anomaly_all = []
        row_vals = [i + 1, f"segment_{r['index']:03d}"]
        for si, entry in enumerate(r['chain']):
            a_str, is_bad, is_warn = detect_anomalies(entry)
            if is_bad:
                total_bad += 1
                anomaly_all.append(f"[Step{entry['step_order']}] " + a_str.replace('\n', ' / '))
            elif is_warn:
                total_warn += 1
                anomaly_all.append(f"[Step{entry['step_order']}] " + a_str.replace('\n', ' / '))
            row_vals += [entry['in_frames'], entry['expected'],
                         entry['actual'] if entry['actual'] is not None else 'N/A',
                         entry['diff'] if entry['diff'] is not None else 'N/A']
            sum_expected[si] += entry['expected'] or 0
            if entry['actual'] is not None:
                sum_actual[si] += entry['actual']
        anomaly_str = '\n'.join(anomaly_all) if anomaly_all else '✅ 正常'
        row_vals.append(anomaly_str)

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            if ci == num_cols:
                font = bad_font if anomaly_all and any('❌' in a or 'illegal' in a or 'PPS' in a for a in anomaly_all) else (warn_font if anomaly_all else ok_font)
                fill = bad_fill if '❌' in anomaly_str or 'illegal' in anomaly_str or 'PPS' in anomaly_str else (warn_fill if anomaly_all else None)
                style_cell(cell, font, left_align, fill)
            elif ci >= 3 and (ci - 3) % 4 == 3:  # 差值列
                dv = val if isinstance(val, (int, float)) else None
                f, fill = diff_font(dv)
                style_cell(cell, f, center_align, fill)
            else:
                style_cell(cell)
        ws.row_dimensions[row].height = max(24, len(anomaly_all) * 16 + 10)
        row += 1

    # 汇总行
    sum_vals = ['汇总', '']
    for si in range(n_steps):
        d = sum_actual[si] - sum_expected[si]
        sum_vals += ['', sum_expected[si], sum_actual[si], d]
    summary_txt = f"❌ 异常段 {total_bad} | ⚠ 警告 {total_warn}" if (total_bad or total_warn) else '✅ 全部正常'
    sum_vals.append(summary_txt)
    for ci, val in enumerate(sum_vals, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        style_cell(cell, sum_font, center_align if ci != num_cols else left_align, sum_fill)
    row += 1

    # 最终视频比对行
    if final_probe.get('frames') and sum_actual:
        expected_final = sum_actual[-1]
        final_frames = _to_int(final_probe['frames']) or 0
        final_diff = final_frames - expected_final
        fvals = ['最终视频', Path(info.get('output_video', '')).name] + [''] * (num_cols - 4) + [final_diff, '']
        fvals[2 + (n_steps - 1) * 4 + 2] = final_frames  # 放到最后一步"实际输出"列
        for ci, val in enumerate(fvals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            if val == final_diff and ci > 2:
                f, fill = diff_font(final_diff)
                style_cell(cell, f, center_align, fill)
            else:
                style_cell(cell, sum_font, center_align, sum_fill)

    ws.freeze_panes = f'A{header_row + 1}'

    # ══ Sheet 2..N: 各步骤明细 ══
    for step in steps:
        order = step['order']
        ws_s = wb.create_sheet(f"Step{order}-{step['type']}")
        sdata = ffprobe_data.get(order, {})
        in_pk = sdata.get('in_packets', [])
        in_fr = sdata.get('in_frames', [])
        out_pk = sdata.get('out_packets', [])
        out_fr = sdata.get('out_frames', [])
        ff_frames, ff_errors, ff_dd, _ff_decoders = ffmpeg_results.get(order, ([], [], [], []))

        headers_s = ['分段序号', '分段编号', '输入帧数', '期望输出', '日志输出', '差值1\n(日志-期望)',
                     'ffprobe\npackets', '差值2', 'ffprobe\nframes', '差值3',
                     'ffmpeg\n拆帧', '差值4', 'dup', 'drop',
                     'pts_dups', 'pts_drops', 'pts_drop_est', 'pts_backward', 'repeat_pict',
                     '异常检测', '输出时长', '处理耗时', '性能统计\n(总s/平均ms/FPS)',
                     'GFPGAN', '输出大小', '输出文件名']
        widths_s = [8, 22, 11, 11, 11, 12, 12, 9, 12, 9, 11, 9, 8, 8, 9, 9, 11, 11, 10, 30, 12, 12, 22, 10, 11, 40]
        ncols_s = len(headers_s)

        ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_s)
        ws_s['A1'] = f"Step {order}/{n_steps} — {step['name']} ({step['type']})"
        ws_s['A1'].font = title_font
        ws_s['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_s.row_dimensions[1].height = 28

        ws_s['A2'] = f"输入目录: {step['input_dir']}"
        ws_s['A3'] = f"输出目录: {step['output_dir']}"
        factor_desc = f"插帧倍数: {step['factor']}×" if step['type'] == 'interpolate' else f"超分倍数: {step['factor']}×"
        ws_s['A4'] = f"{factor_desc} | 期望规则: " + \
                     ("(输入-1)×倍数+1" if step['type'] == 'interpolate' else "输入帧（超分不改变帧数）")
        for rr in (2, 3, 4):
            ws_s[f'A{rr}'].font = Font(name='微软雅黑', size=9, color='666666')

        hr = 6
        for ci, (h, w) in enumerate(zip(headers_s, widths_s), 1):
            cell = ws_s.cell(row=hr, column=ci, value=h)
            style_cell(cell, header_font, center_align, header_fill)
            ws_s.column_dimensions[get_column_letter(ci)].width = w
        ws_s.row_dimensions[hr].height = 44

        rr = hr + 1
        step_bad = 0
        step_warn = 0
        for i, seg in enumerate(step['segments']):
            chain_entry = chain_rows[i]['chain'][order - 1] if i < len(chain_rows) and len(chain_rows[i]['chain']) >= order else None
            if chain_entry is None:
                continue
            a_str, is_bad, is_warn = detect_anomalies(chain_entry)
            if is_bad:
                step_bad += 1
            elif is_warn:
                step_warn += 1

            perf_str = ''
            if seg.get('perf_total_sec') is not None:
                perf_str = (f"{seg['perf_total_sec']:.1f}s / {seg.get('perf_avg_ms', 0):.1f}ms / "
                            f"FPS {seg.get('perf_fps', 0):.2f}")
            else:
                # IFRNet: 无 [性能统计] 行，基于 GPU 原始帧/s + 耗时 + 输出帧推算
                out_frames = seg.get('output_frames')
                total_sec = _parse_elapsed_to_sec(seg.get('elapsed', ''))
                ifrnet_gpu_fps = seg.get('ifrnet_raw_fps')
                if total_sec and total_sec > 0 and out_frames:
                    out_fps = out_frames / total_sec
                    avg_ms = total_sec * 1000 / out_frames
                    if ifrnet_gpu_fps:
                        perf_str = (f"{total_sec:.1f}s / {avg_ms:.1f}ms / "
                                    f"出{out_fps:.1f}fps GPU{seg['raw_frames']}帧@{ifrnet_gpu_fps:.1f}raw/s")
                    else:
                        perf_str = (f"(推算) {total_sec:.1f}s / {avg_ms:.1f}ms / "
                                    f"FPS {out_fps:.2f}")
                elif total_sec:
                    perf_str = f"(耗时) {total_sec:.1f}s"
            out_name = Path(seg['output_file']).name if seg.get('output_file') else ''
            vals = [
                i + 1, f"{step['output_prefix']}_{seg['index']:03d}",
                chain_entry['in_frames'], chain_entry['expected'],
                chain_entry['log_out'], chain_entry['log_diff'],
                _to_int(out_pk[i]) if i < len(out_pk) and _to_int(out_pk[i]) is not None else 'N/A',
                chain_entry['diff'] if chain_entry['actual'] is not None and _to_int(out_pk[i]) is None else
                ((_to_int(out_pk[i]) - chain_entry['expected']) if i < len(out_pk) and _to_int(out_pk[i]) is not None else 'N/A'),
                _to_int(out_fr[i]) if i < len(out_fr) and _to_int(out_fr[i]) is not None else 'N/A',
                chain_entry['diff'] if chain_entry['diff'] is not None else 'N/A',
                chain_entry['ff_val'] if chain_entry['ff_val'] is not None else 'N/A',
                chain_entry['ff_diff'] if chain_entry['ff_diff'] is not None else 'N/A',
                chain_entry['dup'] if chain_entry['dup'] is not None else '',
                chain_entry['drop'] if chain_entry['drop'] is not None else '',
                chain_entry['pts_dups'],
                chain_entry['pts_drops'],
                chain_entry['pts_drop_frames_est'],
                chain_entry['pts_backward'],
                chain_entry['repeat_pict_frames'],
                a_str, seg.get('output_duration', ''), seg.get('elapsed', ''),
                perf_str, step.get('gfpgan_mode', '') if step['type'] == 'upscale' else '',
                seg.get('output_size', ''), out_name,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws_s.cell(row=rr, column=ci, value=val)
                if ci == 20:
                    font = bad_font if is_bad else (warn_font if is_warn else ok_font)
                    fill = bad_fill if is_bad else (warn_fill if is_warn else None)
                    style_cell(cell, font, left_align, fill)
                elif ci in (6, 8, 10, 12):
                    dv = val if isinstance(val, (int, float)) else None
                    f, fill = diff_font(dv)
                    style_cell(cell, f, center_align, fill)
                else:
                    style_cell(cell)
            ws_s.row_dimensions[rr].height = max(26, a_str.count('\n') * 16 + 12)
            rr += 1

        # 步骤汇总行
        sum_line = f"异常段 {step_bad} | 警告 {step_warn}" if (step_bad or step_warn) else '✅ 全部正常'
        svals = ['汇总', ''] + [''] * 17 + [sum_line] + [''] * 6
        for ci, val in enumerate(svals, 1):
            cell = ws_s.cell(row=rr, column=ci, value=val)
            style_cell(cell, sum_font, center_align if ci != 20 else left_align, sum_fill)
        ws_s.freeze_panes = f'A{hr + 1}'

    # ══ 末 Sheet: ffmpeg 错误详情 ══
    ws2 = wb.create_sheet("ffmpeg错误详情")
    ws2.merge_cells('A1:L1')
    ws2['A1'] = ('ffmpeg 解码 — 逐段错误详情（全部步骤）| '
                 '-vsync 0 下 vsync_dup/drop 应为 0; pts_dups/drops/est/repeat_pict 为原始时间戳/场异常')
    ws2['A1'].font = title_font
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28
    headers2 = ['步骤', '分段', 'ffmpeg拆帧帧数', 'vsync_dup', 'vsync_drop',
                'pts_dups', 'pts_drops', 'pts_drop_est', 'pts_backward', 'repeat_pict',
                '解码器', '错误/警告信息']
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        style_cell(cell, header_font, center_align, header_fill)
    for col, w in zip('ABCDEFGHIJKL', (10, 34, 16, 10, 10, 10, 10, 12, 12, 10, 12, 60)):
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[3].height = 26

    rr = 4
    for step in steps:
        order = step['order']
        ff_frames, ff_errors, ff_dd, ff_decoders = ffmpeg_results.get(order, ([], [], [], []))
        for i, seg in enumerate(step['segments']):
            errs = ff_errors[i] if i < len(ff_errors) else []
            dd = ff_dd[i] if i < len(ff_dd) else None
            decoder = ff_decoders[i] if i < len(ff_decoders) else 'unknown'
            anomalies = dd.get('pts_anomalies', []) if dd else []
            info_lines = list(errs[:30])
            if anomalies:
                info_lines.append('[pts 异常明细]')
                for n, pts, pts_time, typ, note in anomalies[:20]:
                    ts = f" pts_time={pts_time:.6f}" if pts_time is not None else ""
                    info_lines.append(f"  n={n} pts={pts}{ts} type={typ} {note}")

            vals = [f"Step{order}",
                    f"{step['output_prefix']}_{seg['index']:03d}",
                    ff_frames[i] if i < len(ff_frames) else 'N/A',
                    dd.get('dup', '') if dd else '',
                    dd.get('drop', '') if dd else '',
                    dd.get('pts_dups', '') if dd else '',
                    dd.get('pts_drops', '') if dd else '',
                    dd.get('pts_drop_frames_est', '') if dd else '',
                    dd.get('pts_backward', '') if dd else '',
                    dd.get('repeat_pict_frames', '') if dd else '',
                    decoder,
                    '\n'.join(info_lines) if info_lines else '✅ 无错误']
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=rr, column=ci, value=val)
                style_cell(cell, bad_font if (info_lines and ci == 12) else normal_font,
                           left_align if ci == 12 else center_align)
            ws2.row_dimensions[rr].height = max(26, min(300, len(info_lines) * 16))
            rr += 1
    ws2.freeze_panes = 'A4'

    wb.save(output_path)
    print(f"✅ xlsx 已生成: {output_path}")


# ═══════════════════════════════════════════════
# 7. 主函数
# ═══════════════════════════════════════════════

def main():
    t_start = time.time()
    parser = argparse.ArgumentParser(
        description='视频增强流水线 全流程检测分析脚本 v3（多日志 + 自动资源探测 + 多进程/多线程并行 + GPU硬件加速）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('log_files', nargs='+',
                        help='流水线输出日志文件，支持多个 (如 output28-9.txt output28-10.txt)')
    parser.add_argument('--suffix', default=None,
                        help='输出文件后缀 (仅单日志时有效; 默认: 从日志文件名推导, 如 output28-9.txt -> 28-9)')
    parser.add_argument('--segments-dir', default=None,
                        help='初始分段目录 (默认: 由第一个步骤输出目录反推 ../segments; 多日志时应用于全部日志)')
    parser.add_argument('--output-video', default=None,
                        help='最终输出视频路径 (仅单日志时有效; 默认: 从日志自动解析)')
    parser.add_argument('--output-dir', default=None,
                        help='分析文件输出目录 (默认: 各日志同目录)')

    parser.add_argument('--skip-ffmpeg', action='store_true', help='跳过 ffmpeg 解码')
    parser.add_argument('--skip-ffprobe-frames', action='store_true',
                        help='跳过 ffprobe -count_frames (全量解码，慢)；-count_packets 快速扫描不受影响')
    parser.add_argument('--skip-xlsx', action='store_true', help='跳过 xlsx 生成')
    parser.add_argument('--workers', type=int, default=None,
                        help='并行 worker 数 (默认: 按 CPU 核数/可用RAM 自动计算)')
    parser.add_argument('--parallel', choices=['auto', 'thread', 'process'], default='auto',
                        help='并行模式: auto=自动(子进程型任务默认 thread) | thread=多线程 | process=多进程')
    parser.add_argument('--task-ram-mb', type=int, default=300,
                        help='自动 workers 时每任务内存估计 MB (默认 300)')
    # ── GPU 硬件加速 (v3) ──
    parser.add_argument('--hwaccel', choices=['auto', 'cuda', 'off'], default='auto',
                        help='GPU 硬件加速: auto=自动检测 (可用则 cuda) | cuda=强制 CUDA (失败回退) | off=纯 CPU')
    parser.add_argument('--gpu-workers', type=int, default=4,
                        help='GPU 任务并发闸门 (默认 4, GeForce 须防 NVDEC 会话限制; T4/A100 可调大)')
    parser.add_argument('--pts-drop-threshold', type=float, default=2.0,
                        help='pts 跳变判定阈值：diff > threshold * interval 计为丢帧 (默认 2.0，'
                             'VFR/录屏源可酌情加大到 3.0~5.0)')


    args = parser.parse_args()

    # ── 环境检查（跨平台: shutil.which 定位 ffmpeg/ffprobe）──
    if not args.skip_ffmpeg and shutil.which('ffmpeg') is None:
        print("[WARN] PATH 中未找到 ffmpeg，ffmpeg 解码任务将全部标记 [SKIP]")
    if shutil.which('ffprobe') is None:
        print("[WARN] PATH 中未找到 ffprobe，ffprobe 统计任务将全部标记 [SKIP]")

    # ── GPU 硬件加速自检 ──
    gpu_caps = {'cuda': False, 'nvenc': False, 'hint': '未检测'}
    if args.hwaccel != 'off':
        gpu_caps = detect_gpu_hwaccel()
    if args.hwaccel == 'cuda' and not gpu_caps['cuda']:
        print(f"[WARN] --hwaccel cuda 指定，但自检失败: {gpu_caps['hint']}，强制回退 CPU")
        print(f"[WARN]   可试 --hwaccel off 跳过自检 (但 GPU 也会完全禁用)")
    hwaccel_mode = 'cuda' if (args.hwaccel == 'auto' and gpu_caps['cuda'] or
                              args.hwaccel == 'cuda' and gpu_caps['cuda']) else 'off'
    gpu_enabled = (hwaccel_mode == 'cuda')
    gpu_desc = ('CUDA' if gpu_caps['cuda'] else '不可用') + (
        '+NVENC' if gpu_caps['nvenc'] else '')


    # ── 系统资源探测与并行参数 ──
    cpu_count = detect_cpu_count()
    cpu_host = os.cpu_count() or 1
    ram_total, ram_avail = detect_ram_gb()
    ram_desc = f"总计 {ram_total:.1f}GB / 可用 {ram_avail:.1f}GB" if ram_total > 0 else "未知"
    # auto: 任务为子进程型（ffprobe/ffmpeg 自身独立进程解码），Python 侧仅等待，thread 开销最小
    parallel_mode = 'thread' if args.parallel == 'auto' else args.parallel
    # ── GPU 保卫：process 模式信号量不跨进程共享 → --gpu-workers 闸门失效
    #    GPU 启用时自动禁用 process，强制 thread（覆盖用户显式指定的 --parallel process）
    if gpu_enabled and parallel_mode == 'process':
        print(f"[WARN] GPU {gpu_desc} 已启用，--parallel process 不兼容 GPU 并发闸门 "
              f"(--gpu-workers={args.gpu_workers})，已自动切换为 thread")
        parallel_mode = 'thread'
    task_ram = args.task_ram_mb
    workers = args.workers if args.workers and args.workers > 0 else compute_auto_workers(task_ram)

    # ── 初始化 GPU 并发信号量 (thread 模式生效) ──
    global _GPU_SEMAPHORE
    gpu_workers = max(1, min(args.gpu_workers, 8)) if gpu_enabled else 0
    if gpu_enabled and parallel_mode == 'thread':
        _GPU_SEMAPHORE = threading.BoundedSemaphore(gpu_workers)

    cpu_note = f" (宿主机 {cpu_host})" if cpu_count != cpu_host else ""
    print(f"🖥️  系统资源: CPU×{cpu_count}{cpu_note} | RAM {ram_desc}")
    gpu_info = ""
    if gpu_enabled:
        gpu_info = f" | GPU {gpu_desc} (hwaccel={hwaccel_mode}, gpu_workers={gpu_workers})"
    print(f"⚙️  并行配置: workers={workers} | 模式={parallel_mode} | 单任务内存估计={task_ram}MB{gpu_info}")

    # ── 解析全部日志：步骤顺序 / 分段位置 / 最终输出（CPU 轻量，串行即可）──
    jobs = []
    single = len(args.log_files) == 1
    if args.suffix and not single:
        print("[WARN] 多日志模式下 --suffix 无效，各日志自动推导后缀")
    for log_arg in args.log_files:
        log_path = Path(log_arg).resolve()
        if not log_path.exists():
            print(f"[ERROR] 日志文件不存在: {log_path}，已跳过")
            continue

        log_dir = log_path.parent
        output_dir = Path(args.output_dir).resolve() if args.output_dir else log_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        suffix = args.suffix if (args.suffix and single) else (re.sub(r'^output', '', log_path.stem) or 'pipeline')

        print(f"\n📖 解析日志: {log_path}")
        info = parse_pipeline_log(str(log_path))
        if args.output_video and single:
            info['output_video'] = args.output_video

        resolve_step_dirs(info, log_dir, args.segments_dir)

        print(f"   处理模式: {info['mode'] or 'N/A'} | 分段数: {info['num_segments']}")
        for step in info['steps']:
            print(f"   Step {step['order']}: {step['name']} [{step['type']}] "
                  f"factor={step['factor']} 段数={len(step['segments'])}")
            print(f"      输入: {step['input_dir']}")
            print(f"      输出: {step['output_dir']}")
        print(f"   最终输出: {info.get('output_video', 'N/A')}")

        if not info['steps']:
            print("[WARN] 未能从日志中解析出任何处理步骤，已跳过")
            continue

        jobs.append({'log_path': log_path, 'output_dir': output_dir,
                     'suffix': suffix, 'info': info})

    if not jobs:
        print("[ERROR] 没有可分析的有效日志")
        sys.exit(1)

    # ── 两阶段智能调度 ──
    #    阶段 1: 快速扫描 ffprobe/packets（仅读包头），获取各分段帧数（cost）
    #    阶段 2: 根据帧数排序后执行 ffprobe/frames + ffmpeg 解码（重任务, GPU 加速）
    prefetched = {}

    # 预先计算共享排序数据：step 目录优先级 + 最终视频路径
    # typ: output=1(高), input=0(低) → reverse=True 时 output 排前
    step_dir_priority = {}
    for job in jobs:
        for step in job['info']['steps']:
            order = step['order']
            for key, typ in [('output_dir', 1), ('input_dir', 0)]:
                d = step.get(key, '')
                if d:
                    step_dir_priority[str(Path(d).resolve())] = (order, typ)
    final_video_paths = set()
    for job in jobs:
        fv = job['info'].get('output_video', '')
        if fv:
            final_video_paths.add(str(fv))

    # === 阶段 1：快速扫描 (ffprobe packets，始终执行，--skip-ffprobe-frames 不影响) ===
    fast_tasks = []
    for job in jobs:
        fast_tasks.extend(collect_ffprobe_packets_tasks(job['info']))
    if fast_tasks:
        uniq = {}
        for t in fast_tasks:
            uniq.setdefault(t['key'], t)
        fast_list = list(uniq.values())
        # 排序：最终视频 → Step N 输出 → Step N 输入 → ...
        fv_tasks = [t for t in fast_list if t['path'] in final_video_paths]
        other_tasks = [t for t in fast_list if t['path'] not in final_video_paths]
        other_tasks.sort(key=lambda t: _path_step_priority(t['path'], step_dir_priority),
                         reverse=True)
        fast_list = fv_tasks + other_tasks
        phase1_label = '阶段 1/2: 快速扫描'
        print(f"\n📊 {phase1_label} ffprobe/packets ({len(fast_list)} 个文件, workers={workers}) ...")
        t_phase1 = time.time()
        fast_results = run_tasks_parallel(fast_list, workers, parallel_mode,
                                          phase_label=phase1_label)
        prefetched.update(fast_results)
        frame_costs = _extract_segment_frame_counts(prefetched)
        # 最终输出视频 cost = 最后步骤输出分段帧数之和（保证排序时排第一位）
        for job in jobs:
            fv = job['info'].get('output_video', '')
            if fv and fv not in frame_costs:
                steps = job['info'].get('steps', [])
                if steps:
                    last_step = steps[-1]
                    out_dir = Path(last_step['output_dir']) if last_step['output_dir'] else None
                    fv_cost = 0
                    if out_dir:
                        for seg_data in last_step['segments']:
                            sf = _find_segment_file(out_dir,
                                                    last_step['output_prefix'] or 'segment',
                                                    seg_data['index'])
                            fv_cost += frame_costs.get(str(sf), 0)
                    if fv_cost > 0:
                        frame_costs[fv] = fv_cost
        if frame_costs:
            total_frames = sum(frame_costs.values())
            fv_has = sum(1 for j in jobs if j['info'].get('output_video', '') in frame_costs)
            fv_note = f" (含 {fv_has} 个最终视频估算)" if fv_has else ""
            print(f"   已获取 {len(frame_costs)} 个文件的帧数，总计 {total_frames} 帧{fv_note}")
        else:
            print("   未获取到帧数（packets 结果不可解析），第二阶段将按任务数调度")
        print(f"   ✅ 阶段 1 完成，用时 {time.time() - t_phase1:.1f}s")
    else:
        frame_costs = {}

    # === 阶段 2：重任务 (ffprobe frames + ffmpeg 解码, GPU 加速) ===
    heavy_tasks = []
    if not args.skip_ffprobe_frames:
        for job in jobs:
            heavy_tasks.extend(collect_ffprobe_frames_tasks(job['info'], hwaccel=hwaccel_mode))
    if not args.skip_ffmpeg:
        for job in jobs:
            heavy_tasks.extend(collect_ffmpeg_tasks(job['info'],
                                                     hwaccel=hwaccel_mode,
                                                     pts_drop_threshold=args.pts_drop_threshold))
    if heavy_tasks:
        uniq = {}
        for t in heavy_tasks:
            uniq.setdefault(t['key'], t)
        heavy_list = list(uniq.values())
        # 排序：最终视频 → Step N 输出 → Step N 输入 → ...
        #   同组内: ffprobe 先, ffmpeg 后
        heavy_list = _enrich_and_sort_heavy_tasks(heavy_list, frame_costs,
                                                   step_dir_priority=step_dir_priority,
                                                   final_video_paths=final_video_paths)
        total_work = sum(t.get('cost', 0) for t in heavy_list)
        n_desc = f"{len(heavy_list)} 任务, 总 {total_work} 帧" if total_work else f"{len(heavy_list)} 任务"
        phase_label = '阶段 2/2: GPU 深度检测' if gpu_enabled else '阶段 2/2: 深度检测'
        print(f"\n🔬 {phase_label} ({n_desc}, workers={workers}) ...")
        t_phase2 = time.time()
        heavy_results = run_tasks_parallel(heavy_list, workers, parallel_mode,
                                           total_work=total_work,
                                           phase_label=phase_label)
        prefetched.update(heavy_results)
        print(f"   ✅ 阶段 2 完成，用时 {time.time() - t_phase2:.1f}s")
        print(f"✅ 两阶段并行完成: {len(prefetched)} 个结果，"
              f"用时 {time.time() - t_start:.1f}s")
    else:
        print(f"✅ 阶段 1 完成，无重任务")

    # ── 逐日志生成分析文件（I/O 轻量串行，输出格式与 v1 完全一致）──
    for job in jobs:
        info, output_dir, suffix = job['info'], job['output_dir'], job['suffix']
        print(f"\n📦 生成分析文件: {job['log_path'].name} (suffix={suffix})")

        # ── ffprobe 帧数统计 ──
        ffprobe_data = {}
        # ffprobe 输出：packets 始终生成，frames 受 --skip-ffprobe-frames 控制
        ffprob_path = output_dir / f"ffprob_output{suffix}.txt"
        print(f"🔍 生成 ffprobe 输出: {ffprob_path}")
        try:
            ffprobe_data = generate_ffprob_output(info, str(ffprob_path),
                                                  prefetched=prefetched,
                                                  skip_frames=args.skip_ffprobe_frames,
                                                  hwaccel=hwaccel_mode)
        except Exception as e:
            print(f"[WARN] ffprobe 生成异常: {e}，继续后续步骤")

        # ── ffmpeg 解码 ──
        ffmpeg_results = {}
        if not args.skip_ffmpeg:
            ffmpeg_path = output_dir / f"ffmpeg_output{suffix}.txt"
            gpu_tag = f",GPU" if gpu_enabled else ""
            mode_label = "showinfo" + gpu_tag
            print(f"🎬 生成 ffmpeg 输出 ({mode_label} 解码模式): {ffmpeg_path}")
            try:
                ffmpeg_results = generate_ffmpeg_output(info, str(ffmpeg_path),
                                                        prefetched=prefetched,
                                                        hwaccel=hwaccel_mode,
                                                        timeout=900,
                                                        pts_drop_threshold=args.pts_drop_threshold)
            except Exception as e:
                print(f"[WARN] ffmpeg 生成异常: {e}，继续后续步骤")

        # ── 帧数链计算 ──
        chain_rows = build_chain(info, ffprobe_data, ffmpeg_results)

        # ── xlsx ──
        if not args.skip_xlsx:
            xlsx_path = output_dir / f"帧数差值统计汇总_output{suffix}.xlsx"
            print(f"📊 生成 xlsx: {xlsx_path}")
            try:
                generate_xlsx(info, ffprobe_data, ffmpeg_results, chain_rows, str(xlsx_path))
            except Exception as e:
                print(f"[ERROR] xlsx 生成失败: {e}")
                import traceback
                traceback.print_exc()

    print("\n" + "=" * 60)
    print(f"  ✅ 全流程检测分析文件生成完毕！共 {len(jobs)} 个日志，总用时 {time.time() - t_start:.1f}s")
    print("=" * 60)


if __name__ == '__main__':
    main()
